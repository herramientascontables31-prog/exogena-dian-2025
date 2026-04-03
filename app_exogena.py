import os
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime
from io import BytesIO
import difflib

st.set_page_config(page_title="Exógena DIAN 2025", page_icon="📊", layout="wide")

# === PROTECCIÓN CON CONTRASEÑA (Google Sheets) ===
GOOGLE_SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/e/TU_ID_AQUI/pub?output=csv"

# === DIRECTORIO CENTRALIZADO DE TERCEROS ===
DIRECTORIO_CENTRAL_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQAr6NT6bJmum4GUAPuPmrJ2m-pybG8Nyve1Nv8s_MyuUOFQZ42gIPpKt_dm0efy1J8kqGcE8AwNxdk/pub?gid=0&single=true&output=csv"

CLAVE_ADMIN = os.environ.get("EXODIAN_ADMIN_KEY", "")

@st.cache_data(ttl=300)
def cargar_clientes():
    try:
        df = pd.read_csv(GOOGLE_SHEET_CSV_URL, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        clientes = {}
        for _, row in df.iterrows():
            clave = str(row.get('clave', '')).strip()
            nombre = str(row.get('nombre', '')).strip()
            estado = str(row.get('estado', '')).strip().lower()
            if clave and clave.lower() != 'nan':
                clientes[clave] = {
                    'nombre': nombre if nombre.lower() != 'nan' else '',
                    'activo': estado in ('activo', 'si', 'sí', '1', 'true', 'yes'),
                }
        return clientes, None
    except Exception as e:
        return {}, str(e)

def verificar_clave(clave_ingresada):
    if clave_ingresada == CLAVE_ADMIN:
        return True, "🔑 Administrador", None
    clientes, error = cargar_clientes()
    if error:
        return False, "", f"⚠️ No se pudo verificar. Contacta al administrador. ({error[:60]})"
    if clave_ingresada in clientes:
        cliente = clientes[clave_ingresada]
        if cliente['activo']:
            return True, cliente['nombre'], None
        else:
            return False, "", "🚫 Tu acceso está **desactivado**. Contacta al administrador."
    return False, "", "❌ Contraseña incorrecta. Verifica tu compra en exogenadian.com"

# === PRO VALIDATION ===
PRO_SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQc7cTur8DOZ_Kqkpqf7WmbzFT4im5efh0gYzIix4HE9pYp5B24OSDaOCKWjuU5YVXAMZeMGYkVE1eH/pub?gid=0&single=true&output=csv"

@st.cache_data(ttl=300)
def cargar_claves_pro():
    try:
        df = pd.read_csv(PRO_SHEET_URL, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        claves = {}
        for _, row in df.iterrows():
            clave = str(row.get('clave', '')).strip()
            estado = str(row.get('estado', '')).strip().lower()
            nombre = str(row.get('nombre', '')).strip()
            if clave and clave.lower() != 'nan':
                claves[clave] = {'activo': estado in ('activo', 'si', 'sí', '1', 'true'), 'nombre': nombre}
        return claves
    except:
        return {}

def verificar_pro(clave):
    if clave == CLAVE_ADMIN:
        return True, "Administrador"
    claves = cargar_claves_pro()
    if clave in claves and claves[clave]['activo']:
        return True, claves[clave]['nombre']
    return False, ""

if "es_pro" not in st.session_state:
    st.session_state.es_pro = False
    st.session_state.pro_nombre = ""

es_pro = st.session_state.es_pro

st.sidebar.markdown("### 📊 Exógena DIAN 2025")
st.sidebar.markdown("---")
if es_pro:
    st.sidebar.success(f"✅ PRO activo — {st.session_state.pro_nombre}")
    if st.sidebar.button("Cerrar PRO"):
        st.session_state.es_pro = False
        st.session_state.pro_nombre = ""
        st.rerun()
else:
    st.sidebar.info("🆓 Versión gratuita")
    st.sidebar.markdown("F1001 y F2276 requieren PRO")
    with st.sidebar.expander("🔑 Activar PRO"):
        clave_pro = st.text_input("Clave PRO", type="password", placeholder="Ej: PRO-XXXXX")
        if st.button("Activar", use_container_width=True):
            valida, nombre = verificar_pro(clave_pro)
            if valida:
                st.session_state.es_pro = True
                st.session_state.pro_nombre = nombre
                st.rerun()
            else:
                st.error("❌ Clave inválida o inactiva")
    st.sidebar.markdown("[💳 Suscribirse PRO →](https://exogenadian.com/#planes)")
st.sidebar.markdown("[← Volver a ExógenaDIAN](https://exogenadian.com)")

# === ESTILOS ===
st.markdown("""
<style>
    .main-header { font-size: 2.2rem; font-weight: 700; color: #1F4E79; margin-bottom: 0; }
    .sub-header { font-size: 1.1rem; color: #666; margin-top: 0; }
    .metric-card { background: #f8f9fa; border-radius: 10px; padding: 1rem; border-left: 4px solid #1F4E79; }
    .success-box { background: #d4edda; border-radius: 10px; padding: 1rem; border-left: 4px solid #28a745; }
    .privacy-box { background: linear-gradient(135deg, #e8f4fd 0%, #f0f7ee 100%); border-radius: 12px; padding: 1.2rem 1.5rem; border: 1px solid #b8d4e8; margin: 0.8rem 0 1.2rem 0; }
    .privacy-box h4 { color: #1a5276; margin: 0 0 0.5rem 0; font-size: 1rem; }
    .privacy-box p { color: #2c3e50; font-size: 0.88rem; line-height: 1.5; margin: 0.3rem 0; }
    .privacy-box .privacy-icon { font-size: 1.3rem; margin-right: 0.3rem; }
    .stDownloadButton > button { background-color: #1F4E79 !important; color: white !important; font-size: 1.1rem !important; padding: 0.5rem 2rem !important; }
    /* Mejorar el file uploader */
    [data-testid="stFileUploader"] { border: 2px solid #2E75B6 !important; border-radius: 12px !important; padding: 8px !important; background: #f8fafc !important; }
    [data-testid="stFileUploader"]:hover { border-color: #1B3A5C !important; background: #EFF6FF !important; }
    [data-testid="stFileUploader"] button { background: #2E75B6 !important; color: white !important; border-radius: 8px !important; font-weight: 600 !important; }
    [data-testid="stFileUploader"] small { color: #6B7280 !important; }
    [data-testid="stFileUploader"] section { padding: 12px !important; }
</style>
""", unsafe_allow_html=True)

# === DIRECTORIO CENTRALIZADO ===
@st.cache_data(ttl=600)
def cargar_directorio_central():
    try:
        df = pd.read_csv(DIRECTORIO_CENTRAL_URL, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        directorio = {}
        for _, row in df.iterrows():
            nit = str(row.get('nit', '')).strip()
            if not nit or nit.lower() == 'nan':
                continue
            nit = nit.replace('.', '').replace('-', '').strip()
            if '.' in nit:
                try: nit = str(int(float(nit)))
                except: pass
            directorio[nit] = {
                'razon': str(row.get('razón social', row.get('razon social', ''))).strip(),
                'dir': str(row.get('dirección', row.get('direccion', ''))).strip(),
                'depto': str(row.get('cod depto', row.get('depto', ''))).strip(),
                'mpio': str(row.get('cod municipio', row.get('municipio', ''))).strip(),
                'pais': str(row.get('cod país', row.get('pais', '169'))).strip(),
                'td': str(row.get('tipo doc', '')).strip(),
                'dv': str(row.get('dv', '')).strip(),
            }
            for k in directorio[nit]:
                if directorio[nit][k].lower() == 'nan':
                    directorio[nit][k] = ''
        return directorio, None
    except Exception as e:
        return {}, str(e)

# === CONSTANTES ===
UVT = 49799
C3UVT = 149397
C12UVT = 597588
NM = "222222222"
TDM = "43"

# =====================================================================
# CORRECCIÓN 1: Nuevas constantes — NIT DIAN + Bancos colombianos
# =====================================================================
BANCOS_COLOMBIANOS = {
    "bancolombia":      ("890903938", "BANCOLOMBIA S.A."),
    "davivienda":       ("860034313", "BANCO DAVIVIENDA S.A."),
    "bogota":           ("860002964", "BANCO DE BOGOTA S.A."),
    "occidente":        ("890300279", "BANCO DE OCCIDENTE S.A."),
    "popular":          ("860007738", "BANCO POPULAR S.A."),
    "bbva":             ("860003020", "BBVA COLOMBIA S.A."),
    "scotiabank":       ("890903937", "SCOTIABANK COLPATRIA S.A."),
    "colpatria":        ("890903937", "SCOTIABANK COLPATRIA S.A."),
    "itau":             ("890903937", "ITAU CORPBANCA COLOMBIA S.A."),
    "av villas":        ("860035827", "BANCO AV VILLAS S.A."),
    "avvillas":         ("860035827", "BANCO AV VILLAS S.A."),
    "caja social":      ("860007335", "BANCO CAJA SOCIAL S.A."),
    "agrario":          ("800037800", "BANCO AGRARIO DE COLOMBIA S.A."),
    "bancamia":         ("900215071", "BANCAMIA S.A."),
    "nequi":            ("890903938", "BANCOLOMBIA S.A."),
    "nu colombia":      ("901654565", "NU COLOMBIA S.A."),
    "nubank":           ("901654565", "NU COLOMBIA S.A."),
    "daviplata":        ("860034313", "BANCO DAVIVIENDA S.A."),
    "rappipay":         ("901346953", "RAPPIPAY DAVIPLATA S.A."),
    "gnb sudameris":    ("860050750", "GNB SUDAMERIS S.A."),
    "gnb":              ("860050750", "GNB SUDAMERIS S.A."),
    "pichincha":        ("890200756", "BANCO PICHINCHA S.A."),
    "falabella":        ("900047981", "BANCO FALABELLA S.A."),
    "serfinanza":       ("860043186", "BANCO SERFINANZA S.A."),
    "mundo mujer":      ("900211468", "BANCO MUNDO MUJER S.A."),
    "coopcentral":      ("890203088", "BANCO COOPCENTRAL"),
    "jp morgan":        ("900114346", "JP MORGAN CORPORACION FINANCIERA S.A."),
    "citibank":         ("860051135", "CITIBANK COLOMBIA S.A."),
    "helm":             ("860050750", "GNB SUDAMERIS S.A."),
    "corpbanca":        ("890903937", "ITAU CORPBANCA COLOMBIA S.A."),
    "ban100":           ("900943055", "LULO BANK S.A."),
    "lulo":             ("900943055", "LULO BANK S.A."),
}

NIT_DIAN = "800197268"
CUENTAS_IMPUESTOS_DIAN = ["2365", "2367", "2404", "2408", "2412"]

# Conceptos F1001 que solo deben reportar entidades (NIT jurídico)
# Los pagos de seguridad social deben coincidir con la planilla PILA
CONCEPTOS_SOLO_ENTIDADES = {"5011", "5012", "5013", "5023", "5024", "5025", "5027"}

# === FUNCIONES CORE ===
def calc_dv(n):
    n = str(n).replace(".", "").replace("-", "").strip()
    if not n or not n.isdigit() or n == NM:
        return ""
    pesos = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    np = n.zfill(15)
    s = sum(int(np[i]) * pesos[i] for i in range(15))
    r = s % 11
    return str(11 - r) if r >= 2 else str(r)

def detectar_tipo_doc(nit):
    if not nit or nit == NM:
        return TDM
    nit = str(nit).strip()
    if '.' in nit:
        try: nit = str(int(float(nit)))
        except: pass
    if not nit.isdigit():
        letras = sum(1 for c in nit if c.isalpha())
        if letras <= 3:
            return "41"
        else:
            return "42"
    if len(nit) == 9 and nit[0] in ('8', '9'):
        return "31"
    if len(nit) >= 9 and nit[0] in ('8', '9'):
        return "31"
    return "13"

def safe_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        import math
        if math.isnan(v):
            return 0.0
        return float(v)
    try:
        s = str(v).strip()
        if s.lower() == 'nan' or s == '':
            return 0.0
        return float(s.replace(",", "").replace("$", "").replace(" ", ""))
    except:
        return 0.0

def safe_str(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        import math
        if v == 0 or (isinstance(v, float) and math.isnan(v)):
            return ""
    s = str(v).strip()
    if s.lower() == 'nan' or s == '0.0':
        return ""
    return s

def detectar_columnas(df):
    import unicodedata
    def normalizar(texto):
        if not texto: return ""
        texto = str(texto).lower().strip()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
        for ch in '.,;:-_/\\()[]{}#"\'':
            texto = texto.replace(ch, ' ')
        return ' '.join(texto.split())
    columnas = {}
    nombres = [normalizar(str(c)) for c in df.columns]
    mapeo = {
        'cuenta': ['cuenta', 'codigo cuenta', 'cod cuenta', 'cuenta contable', 'codigo', 'account'],
        'nombre': ['descripcion cuenta', 'descripcion', 'nombre cuenta', 'nombre', 'detalle',
                    'concepto', 'account name', 'desc cuenta'],
        'nit': ['tercero', 'nit', 'identificacion', 'documento', 'id tercero', 'nro documento',
                'num documento', 'cedula', 'numero identificacion'],
        'razon_social': ['razon social', 'nombre tercero', 'tercero nombre', 'razon', 'beneficiario',
                         'proveedor', 'cliente', 'nombre razon social'],
        'debito': ['debitos', 'debito', 'debe', 'movimiento debito', 'mov debito', 'cargos',
                   'debits', 'debit'],
        'credito': ['creditos', 'credito', 'haber', 'movimiento credito', 'mov credito', 'abonos',
                    'credits', 'credit'],
        'saldo': ['saldo final', 'saldo', 'saldo actual', 'balance', 'saldo cierre',
                  'saldo a diciembre', 'saldo dic'],
    }
    for campo, keywords in mapeo.items():
        for i, nom in enumerate(nombres):
            if i in columnas.values(): continue
            for kw in keywords:
                if kw == nom or kw in nom:
                    columnas[campo] = i
                    break
            if campo in columnas: break
    for i, nom in enumerate(nombres):
        if ('saldo inicial' in nom or 'saldo anterior' in nom) and columnas.get('saldo') == i:
            del columnas['saldo']
            for j, nom2 in enumerate(nombres):
                if j not in columnas.values() and j != i:
                    if 'saldo final' in nom2 or (nom2 == 'saldo' and 'inicial' not in nom2):
                        columnas['saldo'] = j
                        break
    return columnas

def validar_columnas(columnas_detectadas):
    requeridas = ['cuenta', 'nit', 'debito', 'credito']
    faltantes = [c for c in requeridas if c not in columnas_detectadas]
    return len(faltantes) == 0, faltantes

def en_rango(cta, d, h):
    n = len(d)
    return cta[:n] >= d and cta[:n] <= h

def pad_dpto(v):
    if not v: return ""
    v = str(v).strip()
    if '.' in v:
        try: v = str(int(float(v)))
        except: pass
    if v.lower() == 'nan': return ""
    return v.zfill(2) if v.isdigit() else v

def pad_mpio(v):
    if not v: return ""
    v = str(v).strip()
    if '.' in v:
        try: v = str(int(float(v)))
        except: pass
    if v.lower() == 'nan': return ""
    return v.zfill(3) if v.isdigit() else v

def buscar_info_terceros(nits_list, progress_bar=None, log_fn=None):
    import requests
    from time import sleep
    import re

    def log(msg):
        if log_fn: log_fn(msg)

    encontrados = {}
    total = len(nits_list)
    errores_seguidos = 0

    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': '*/*',
        'Accept-Language': 'es-CO,es;q=0.9',
    }

    def buscar_rues(nit):
        try:
            resp = requests.get('https://www.rues.org.co/RM/ConsultaNit_Api',
                params={'nit': str(nit), 'tipo': 'N'},
                headers={**HEADERS, 'Referer': 'https://www.rues.org.co/',
                         'X-Requested-With': 'XMLHttpRequest', 'Accept': 'application/json'}, timeout=12)
            if resp.status_code == 200:
                data = resp.json()
                registros = data if isinstance(data, list) else \
                            data.get('registros', data.get('data', [])) if isinstance(data, dict) else []
                if registros and len(registros) > 0:
                    return extraer_info_dict(registros[0]), None
                return None, "Sin resultados"
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def buscar_datos_gov_lote(nits_batch):
        resultados = {}
        datasets = ["c82q-fe7j", "8yz5-t3jw"]
        for ds_id in datasets:
            try:
                nits_str = "','".join(str(n) for n in nits_batch)
                resp = requests.get(f"https://www.datos.gov.co/resource/{ds_id}.json",
                    params={'$where': f"nit in ('{nits_str}')", '$limit': 5000},
                    headers=HEADERS, timeout=20)
                if resp.status_code == 200:
                    data = resp.json()
                    if data and isinstance(data, list):
                        for emp in data:
                            nit_val = str(emp.get('nit', emp.get('NIT', ''))).strip()
                            if nit_val:
                                info = extraer_info_dict(emp)
                                if info:
                                    resultados[nit_val] = info
                                    resultados[nit_val]['_fuente'] = 'Datos.gov.co'
                        if resultados: return resultados, None
                    return {}, f"Dataset {ds_id}: sin datos"
                else: continue
            except Exception as e: continue
        return {}, "Ningún dataset respondió"

    def buscar_einforma(nit):
        try:
            resp = requests.get(f'https://www.einforma.co/servlet/app/portal/ENTP/prod/ETIQUETA_EMPRESA_498/nif/{nit}',
                headers=HEADERS, timeout=12, allow_redirects=True)
            if resp.status_code == 200:
                info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
                texto = resp.text
                rs_match = re.findall(r'<h1[^>]*class="[^"]*nombre[^"]*"[^>]*>([^<]+)</h1>', texto, re.IGNORECASE)
                if not rs_match: rs_match = re.findall(r'<title>([^<]+?)[\s\-|]', texto)
                if rs_match:
                    rs = rs_match[0].strip()
                    if len(rs) > 3 and str(nit) not in rs.lower(): info['razon_social'] = rs.upper()
                dir_match = re.findall(r'(?:Direcci[oó]n|Domicilio)[:\s]*</[^>]+>\s*<[^>]+>([^<]+)', texto, re.IGNORECASE)
                if dir_match: info['dir'] = dir_match[0].strip()
                if info.get('razon_social') or info.get('dir'): return info, None
            return None, f"HTTP {resp.status_code}"
        except Exception as e: return None, str(e)[:80]

    def buscar_web_ddg(nit):
        try:
            resp = requests.get('https://html.duckduckgo.com/html/',
                params={'q': f'NIT {nit} Colombia empresa direccion'}, headers=HEADERS, timeout=12)
            if resp.status_code == 200: return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e: return None, str(e)[:80]

    def buscar_web_bing(nit):
        try:
            resp = requests.get(f'https://www.bing.com/search?q=NIT+{nit}+Colombia+empresa+direccion&setlang=es',
                headers=HEADERS, timeout=12)
            if resp.status_code == 200: return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e: return None, str(e)[:80]

    def buscar_web_google(nit):
        try:
            query = f"NIT+{nit}+Colombia+empresa+direccion"
            resp = requests.get(f'https://www.google.com/search?q={query}&hl=es&gl=co',
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html', 'Accept-Language': 'es-CO,es;q=0.9'}, timeout=12)
            if resp.status_code == 200: return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e: return None, str(e)[:80]

    def extraer_info_web(nit, html):
        info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
        nit_str = str(nit)
        texto = re.sub(r'<[^>]+>', ' ', html)
        texto = re.sub(r'\s+', ' ', texto)
        patrones_rs = [
            r'(?:NIT|Nit|nit)[\s.:]*' + re.escape(nit_str) + r'[\s\-–—:,.]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]+)',
            r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]{5,50}?)[\s\-–—:,.]+(?:NIT|Nit|nit)[\s.:]*' + re.escape(nit_str),
            r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]{5,50}?)\s*[-–—]\s*NIT[\s.:]*' + re.escape(nit_str),
        ]
        for patron in patrones_rs:
            matches = re.findall(patron, texto)
            if matches:
                rs = matches[0].strip().rstrip('.,;:-– ')
                if 3 < len(rs) < 120:
                    info['razon_social'] = rs.upper()
                    break
        patrones_dir = [
            r'(?:Direcci[oó]n|Dir\.?|Ubicaci[oó]n)[:\s]+([A-Za-z]{2,3}[\s.]*(?:No\.?\s*)?\d+[\w\s#\-.,No°]+?\d)',
            r'((?:CL|CR|KR|TV|DG|CALLE|CARRERA|AV|AVENIDA|TRANSVERSAL|DIAGONAL)[\s.]*(?:No\.?\s*)?\d+[\w\s#\-.,No°]*\d)',
        ]
        for patron in patrones_dir:
            matches = re.findall(patron, texto, re.IGNORECASE)
            if matches:
                dir_candidata = matches[0].strip()[:100]
                if len(dir_candidata) > 5:
                    info['dir'] = dir_candidata
                    break
        if info.get('razon_social') or info.get('dir'): return info
        return None

    def extraer_info_dict(emp):
        if not isinstance(emp, dict): return None
        info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
        for campo in ['razon_social', 'Razon_Social', 'nombre', 'Nombre', 'razonSocial', 'RazonSocial',
                       'nombre_razon_social', 'NombreEstablecimiento', 'organizacion', 'nombre_empresa']:
            val = emp.get(campo, '')
            if val:
                info['razon_social'] = str(val).strip().upper()
                break
        for campo in ['digito_verificacion', 'Digito_Verificacion', 'dv', 'DV', 'digitoVerificacion']:
            val = emp.get(campo, '')
            if val is not None and str(val).strip():
                info['dv'] = str(val).strip()
                break
        for campo in ['direccion', 'Direccion', 'direccion_comercial', 'DireccionComercial', 'dir_comercial']:
            val = emp.get(campo, '')
            if val:
                info['dir'] = str(val).strip()
                break
        for campo in ['codigo_departamento', 'departamento', 'cod_departamento', 'CodigoDepartamento', 'dep_codigo', 'cod_depto']:
            val = emp.get(campo, '')
            if val:
                info['dp'] = pad_dpto(str(val).strip())
                break
        for campo in ['codigo_municipio', 'municipio', 'cod_municipio', 'CodigoMunicipio', 'mun_codigo', 'ciudad', 'cod_ciudad']:
            val = emp.get(campo, '')
            if val:
                info['mp'] = pad_mpio(str(val).strip())
                break
        return info if (info.get('razon_social') or info.get('dir')) else None

    # FLUJO PRINCIPAL
    log("📡 **Paso 1:** Consultando datos.gov.co (lote completo)...")
    try:
        lote_result, lote_error = buscar_datos_gov_lote(nits_list)
        if lote_result:
            encontrados.update(lote_result)
            log(f"  ✅ datos.gov.co: {len(lote_result)} terceros encontrados")
        else:
            log(f"  ❌ datos.gov.co: {lote_error}")
    except Exception as e:
        log(f"  ❌ datos.gov.co: Error — {str(e)[:80]}")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    rues_funciona = False
    if nits_faltantes:
        log(f"📡 **Paso 2:** Probando RUES ({len(nits_faltantes)} NITs pendientes)...")
        test_nit = nits_faltantes[0]
        try:
            resultado, error = buscar_rues(test_nit)
            if resultado:
                resultado['_fuente'] = 'RUES'
                encontrados[test_nit] = resultado
                rues_funciona = True
                log(f"  ✅ RUES funciona (NIT {test_nit} encontrado)")
            else:
                log(f"  ❌ RUES: {error}")
        except Exception as e:
            log(f"  ❌ RUES no disponible: {str(e)[:80]}")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    buscador_web = None
    buscadores = [('DuckDuckGo', buscar_web_ddg), ('Bing', buscar_web_bing), ('Google', buscar_web_google)]
    if nits_faltantes:
        log(f"📡 **Paso 3:** Probando buscadores web...")
        test_nit = nits_faltantes[0]
        for nombre_b, fn_b in buscadores:
            try:
                resultado, error = fn_b(test_nit)
                if resultado:
                    resultado['_fuente'] = nombre_b
                    encontrados[test_nit] = resultado
                    buscador_web = (nombre_b, fn_b)
                    log(f"  ✅ {nombre_b} funciona")
                    break
                else:
                    log(f"  ⚠️ {nombre_b}: {error}")
            except Exception as e:
                log(f"  ❌ {nombre_b}: {str(e)[:60]}")
        if not buscador_web:
            log("  ❌ Ningún buscador web funcionó")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    if nits_faltantes and (rues_funciona or buscador_web):
        fuentes_activas = []
        if rues_funciona: fuentes_activas.append(('RUES', buscar_rues))
        if buscador_web: fuentes_activas.append(buscador_web)
        nombres = " → ".join(f[0] for f in fuentes_activas)
        log(f"🔍 **Paso 4:** Buscando {len(nits_faltantes)} NITs restantes [{nombres}]...")
        for i, nit in enumerate(nits_faltantes):
            if progress_bar:
                progress_bar.progress((i + 1) / len(nits_faltantes),
                    text=f"🔍 {nit} ({i+1}/{len(nits_faltantes)}) — Encontrados: {len(encontrados)}")
            if errores_seguidos >= 15:
                log(f"  ⛔ Detenido tras {errores_seguidos} errores seguidos")
                break
            encontro = False
            for nombre_f, fn_f in fuentes_activas:
                try:
                    resultado, error = fn_f(nit)
                    if resultado:
                        resultado['_fuente'] = nombre_f
                        encontrados[nit] = resultado
                        encontro = True
                        errores_seguidos = 0
                        break
                except Exception: continue
            if not encontro: errores_seguidos += 1
            sleep(0.8)

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    if nits_faltantes and len(nits_faltantes) < 30:
        log(f"📡 **Paso 5:** Probando einforma.co ({len(nits_faltantes)} NITs pendientes)...")
        errores_ein = 0
        for nit in nits_faltantes:
            if errores_ein >= 5: break
            try:
                resultado, error = buscar_einforma(nit)
                if resultado:
                    resultado['_fuente'] = 'einforma.co'
                    encontrados[nit] = resultado
                    errores_ein = 0
                else: errores_ein += 1
            except Exception: errores_ein += 1
            sleep(1.0)

    n_dir = sum(1 for d in encontrados.values() if d.get('dir'))
    n_rs = sum(1 for d in encontrados.values() if d.get('razon_social'))
    log(f"\n📊 **Resumen:** {len(encontrados)}/{total} terceros — {n_dir} con dirección, {n_rs} con razón social")
    return encontrados, len(encontrados) == 0 and total > 0


def normalizar_texto(t):
    import unicodedata
    if not t: return ""
    t = str(t).upper().strip()
    t = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    for ch in '.,;:-_/\\()[]{}#"\'':
        t = t.replace(ch, ' ')
    t = ' '.join(t.split())
    reemplazos = [
        ('S A S', 'SAS'), ('S A  S', 'SAS'), ('S A', 'SA'), ('S  A', 'SA'),
        ('E S P', 'ESP'), ('E  S  P', 'ESP'), ('E U', 'EU'), ('E  U', 'EU'),
        ('C I', 'CI'), ('C  I', 'CI'), ('N I T', 'NIT'),
        ('SOCIEDAD ANONIMA', 'SA'), ('SOCIEDAD POR ACCIONES SIMPLIFICADA', 'SAS'),
        ('EMPRESA DE SERVICIOS PUBLICOS', 'ESP'), ('LIMITADA', 'LTDA'),
    ]
    for viejo, nuevo in reemplazos:
        t = t.replace(viejo, nuevo)
    return ' '.join(t.split())


def similitud_textos(a, b):
    a = normalizar_texto(a)
    b = normalizar_texto(b)
    if not a or not b: return 0.0
    if a == b: return 1.0
    pa = set(a.split())
    pb = set(b.split())
    if not pa or not pb: return 0.0
    comunes = pa & pb
    return len(comunes) / max(len(pa), len(pb))

# ======================================================================
# === PARAMETRIZACION ===
PARAM_1001_NOMINA_SUB = {
    "5001": ["06", "07", "08", "09", "10", "15", "27", "30", "33", "36", "39", "42", "45"],
    "5002": ["01", "03", "05"],
    "5024": ["02"],
    "5025": ["03"],
    "5027": ["04"],
    "5023": ["68", "72", "75"],
}

PARAM_1001_RANGOS = [
    ("5005", "5120", "5120", True), ("5005", "5220", "5220", True),
    ("5011", "5230", "5230", True), ("5011", "5130", "5130", True),
    ("5055", "5115", "5115", True), ("5004", "5110", "5110", True),
    ("5016", "5125", "5125", True), ("5016", "5135", "5139", True),
    ("5016", "5140", "5199", True), ("5016", "5210", "5219", True),
    ("5016", "5235", "5249", True), ("5016", "5295", "5299", True),
    ("5055", "5115", "5115", True), ("5006", "5305", "5305", True),
    ("5101", "530540", "530540", True), ("5007", "1435", "1499", True),
    ("5010", "1504", "1699", True), ("5010", "1520", "1540", True),
]

PARAM_1003 = [
    ("1301", "13551505", "13551509"),
    ("1301", "13551510", "13551514"),
    ("1302", "13551515", "13551519"),
    ("1303", "13551520", "13551524"),
    ("1305", "13551525", "13551529"),
    ("1304", "13551530", "13551534"),
    ("1306", "13551535", "13551539"),
    ("1308", "13551540", "13551594"),
    ("1308", "13551595", "13551599"),
    ("1307", "135518", "135518"),
    ("1311", "135599", "135599"),
]

# =====================================================================
# CORRECCIÓN 2: PARAM_1007 — Rangos específicos PRIMERO
# =====================================================================
PARAM_1007 = [
    ("4003", "4210", "4210"),    # Financieros (específico → primero)
    ("4001", "4101", "4199"),
    ("4001", "4135", "4135"),
    ("4002", "4201", "4299"),    # No operacionales (genérico → después)
]

# =====================================================================
# CORRECCIÓN 3: PARAM_1008 — Excluir 1355 (ya va en F1003)
# =====================================================================
PARAM_1008 = [
    ("1315", "1305", "1305"),    # Clientes
    ("1316", "1380", "1399"),    # Deudores varios
    ("1317", "1330", "1340"),    # Anticipos y avances (hasta 1340, excluye 1355)
    ("1317", "1365", "1365"),    # CxC trabajadores
]

# =====================================================================
# PARAM_1009 — Cuentas por pagar: TODAS las de clase 2 (pasivo)
# Impuestos y retenciones → se agrupan a DIAN en el procesamiento
# =====================================================================
PARAM_1009 = [
    ("2202", "2105", "2199"),    # Obligaciones financieras
    ("2201", "2205", "2295"),    # Proveedores nacionales y del exterior
    ("2204", "2305", "2334"),    # Cuentas corrientes, costos por pagar
    ("2210", "2335", "2364"),    # Costos por pagar, dividendos, acreedores oficiales
    ("2206", "2365", "2369"),    # Retenciones y aportes de nómina → DIAN
    ("2207", "2380", "2399"),    # Acreedores varios
    ("2208", "2404", "2499"),    # Impuestos por pagar → DIAN
    ("2209", "2505", "2599"),    # Obligaciones laborales
    ("2210", "2605", "2699"),    # Pasivos estimados y provisiones
    ("2210", "2705", "2799"),    # Diferidos
    ("2210", "2805", "2899"),    # Otros pasivos
    ("2210", "2905", "2999"),    # Bonos y papeles comerciales
]

MAPEO_1012 = [
    ("8302", "1105", "1105"), ("8301", "1110", "1110"),
    ("8305", "1201", "1204"), ("8303", "1205", "1205"),
    ("8304", "1210", "1210"), ("8306", "1225", "1225"),
    ("8309", "1265", "1265"),
]

# === CLASIFICADOR INTELIGENTE POR NOMBRE DE CUENTA ===
KEYWORDS_1001 = [
    ("5001", True, ["sueldo", "salario", "basico", "jornal", "horas extra", "recargo",
                     "auxilio transporte", "auxilio de transporte", "rodamiento"]),
    ("5002", True, ["honorario", "honorarios"]),
    ("5002", True, ["comision", "comisiones"]),
    ("5024", True, ["aporte salud", "aporte eps", "aportes a eps", "aporte a eps",
                     "aporte a salud", "cotizacion salud", "aportes eps"]),
    ("5025", True, ["aporte pension", "aporte a pension", "aportes a pension",
                     "pension obligatoria", "cotizacion pension", "fondo pension", "aportes pension"]),
    ("5027", True, ["arl", "riesgo laboral", "riesgos laborales", "riesgos profesionales",
                     "aporte arl", "aporte riesgo", "aportes arl"]),
    ("5023", True, ["parafiscal", "parafiscales", "icbf", "sena", "caja de compensacion",
                     "compensacion familiar", "comfama", "compensar", "cafam", "colsubsidio", "comfenalco"]),
    ("5001", True, ["vacacion", "vacaciones"]),
    ("5001", True, ["cesantia", "cesantias", "interes sobre cesantia", "intereses cesantia",
                     "intereses sobre cesantias"]),
    ("5001", True, ["prima de servicio", "prima servicio", "prima legal"]),
    ("5001", True, ["dotacion", "suministro a trabajador"]),
    ("5001", True, ["incapacidad", "incapacidades"]),
    ("5001", True, ["bonificacion", "bonificaciones"]),
    ("5011", True, ["seguro", "poliza", "prima de seguro", "todo riesgo", "cumplimiento"]),
    ("5005", True, ["arriendo", "arrendamiento", "arrendamientos", "canon", "alquiler"]),
    ("5004", True, ["acueducto", "alcantarillado", "energia", "electrica", "telefono",
                     "telecomunicacion", "internet", "gas", "servicio publico", "servicios publicos",
                     "vigilancia", "correo", "portes"]),
    ("5004", True, ["transporte", "flete", "acarreo", "taxi", "taxis", "buses", "envio", "mensajeria"]),
    ("5055", True, ["impuesto", "ica", "industria y comercio", "predial", "vehiculo",
                     "timbre", "estampilla", "estampillas"]),
    ("5101", True, ["gmf", "4x1000", "4 x 1000", "gravamen a los movimientos", "gravamen movimiento"]),
    ("5006", True, ["interes bancario", "intereses bancarios", "interes mora",
                     "comision bancaria", "comisiones bancarias", "gasto financiero", "gastos financieros",
                     "diferencia en cambio", "gravamen", "rendimiento financiero"]),
    ("5010", True, ["gastos de personal", "personal admn", "bienestar", "medicina prepagada",
                     "auxilio funerario", "auxilio educativo", "capacitacion empleado"]),
    ("5016", True, ["viaje", "viatico", "pasaje", "tiquete", "hospedaje", "hotel"]),
    ("5016", True, ["mantenimiento", "reparacion", "adecuacion", "instalacion electrica"]),
    ("5016", True, ["legal", "notarial", "registro", "licencia"]),
    ("5016", True, ["depreciacion", "amortizacion", "agotamiento", "provision"]),
    ("5016", True, ["aseo y cafeteria", "cafeteria", "papeleria", "utiles", "fotocopia", "parqueadero",
                     "casino", "restaurante", "representacion", "suscripcion", "afiliacion",
                     "publicidad", "propaganda", "seminario", "elemento de aseo", "diversos"]),
    ("5007", True, ["inventario", "compra de", "mercancia", "materia prima", "material", "insumo", "repuesto"]),
]

KEYWORDS_1007 = [
    ("4001", ["ingreso operacional", "consultoria", "asesoria", "soporte tecnico",
              "capacitacion", "outsourcing", "desarrollo software", "servicio",
              "honorario recibido", "ingreso actividad"]),
    ("4001", ["venta", "comercio", "producto", "mercancia"]),
    ("4002", ["ingreso no operacional", "extraordinario", "recuperacion"]),
    ("4003", ["arrendamiento recibido", "arriendo recibido", "canon recibido"]),
]

KEYWORDS_1003 = [
    ("1301", ["retencion honorario", "retfte honorario", "retefuente honorario", "rete fuente honorario"]),
    ("1302", ["retencion comision", "retfte comision", "retefuente comision"]),
    ("1303", ["retencion servicio", "retfte servicio", "retefuente servicio"]),
    ("1304", ["retencion arriendo", "retfte arriendo", "retefuente arriendo", "retencion arrendamiento"]),
    ("1305", ["retencion rendimiento", "retfte rendimiento", "retefuente rendimiento",
              "retencion financiero", "rendimientos financieros", "rendimiento financiero"]),
    ("1306", ["retencion compra", "retfte compra", "retefuente compra", "retencion enajenacion"]),
    ("1307", ["retencion ica", "rete ica", "reteica", "industria y comercio retenido",
              "ica retenido", "impuesto de industria y comercio"]),
    ("1308", ["otras retencion", "otra retencion", "retencion otro", "retencion otros",
              "retenciones por cobrar"]),
    ("1311", ["autorretencion", "auto retencion", "autoretefte", "autoretfte",
              "anticipo autorretencion"]),
]

def normalizar_nombre(nom):
    import unicodedata
    if not nom: return ""
    nom = str(nom).lower().strip()
    nom = ''.join(c for c in unicodedata.normalize('NFD', nom) if unicodedata.category(c) != 'Mn')
    for ch in '.,;:-_/\\()[]{}#"\'&$%@!¿?':
        nom = nom.replace(ch, ' ')
    return ' '.join(nom.split())


def stem_es(palabra):
    if not palabra or len(palabra) < 4: return palabra
    if palabra.endswith('es') and len(palabra) > 5:
        base = palabra[:-2]
        if base.endswith('ion') or base.endswith('dad') or base.endswith('idad'):
            return base
        return base
    if palabra.endswith('s') and not palabra.endswith('ss'):
        return palabra[:-1]
    return palabra


def clasificar_por_nombre(nom, tabla_keywords):
    nom_n = normalizar_nombre(nom)
    if not nom_n: return None
    palabras_nom = set(nom_n.split())
    stems_nom = set(stem_es(p) for p in palabras_nom)
    nom_stemmed = ' '.join(stem_es(p) for p in nom_n.split())
    for item in tabla_keywords:
        if len(item) == 3:
            conc, ded, keywords = item
        else:
            conc, keywords = item
            ded = True
        for kw in keywords:
            if ' ' not in kw: continue
            kw_stemmed = ' '.join(stem_es(p) for p in kw.split())
            if kw_stemmed in nom_stemmed or kw in nom_n:
                return (conc, ded) if len(item) == 3 else conc
    for item in tabla_keywords:
        if len(item) == 3:
            conc, ded, keywords = item
        else:
            conc, keywords = item
            ded = True
        for kw in keywords:
            if ' ' in kw: continue
            kw_stem = stem_es(kw)
            if kw in palabras_nom or kw_stem in stems_nom:
                return (conc, ded) if len(item) == 3 else conc
    return None


def concepto_1001(cta, nom_cta=""):
    if en_rango(cta, "5105", "5105"):
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado
        sc = cta[4:6] if len(cta) >= 6 else cta[4:] if len(cta) > 4 else ""
        for conc, subs in PARAM_1001_NOMINA_SUB.items():
            if sc in subs: return conc, True
        return "5001", True
    if cta[:2] in ("51", "52", "53"):
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado
        for conc, d, h, ded in PARAM_1001_RANGOS:
            if en_rango(cta, d, h): return conc, ded
        return "5016", True
    if cta[:2] == "14":
        for conc, d, h, ded in PARAM_1001_RANGOS:
            if en_rango(cta, d, h): return conc, ded
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado
    return None, True


def buscar_concepto(cta, params, nom_cta="", tabla_keywords=None):
    for c, d, h in params:
        if en_rango(cta, d, h): return c
    if tabla_keywords and nom_cta:
        resultado = clasificar_por_nombre(nom_cta, tabla_keywords)
        if resultado:
            return resultado if isinstance(resultado, str) else resultado[0]
    return ""

# === PROCESAMIENTO PRINCIPAL (CON TODAS LAS CORRECCIONES) ===
def procesar_balance(df_balance, df_directorio=None, col_map=None, cierra_impuestos=True, dir_central=None, es_pro=False):
    if col_map is None:
        col_map = detectar_columnas(df_balance)

    CI = col_map.get('cuenta', 0)
    NI = col_map.get('nombre', 1)
    TI = col_map.get('nit', 2)
    RI = col_map.get('razon_social', 3)
    DI = col_map.get('debito', 4)
    KI = col_map.get('credito', 5)
    SI = col_map.get('saldo', 6)

    dir_externo = {}
    if df_directorio is not None:
        for _, row in df_directorio.iterrows():
            nit_d = safe_str(row.iloc[0])
            if not nit_d: continue
            if '.' in nit_d:
                try: nit_d = str(int(float(nit_d)))
                except: pass
            dir_externo[nit_d] = {
                'dir': safe_str(row.iloc[1]) if len(row) > 1 else "",
                'dp': pad_dpto(safe_str(row.iloc[2])) if len(row) > 2 else "",
                'mp': pad_mpio(safe_str(row.iloc[3])) if len(row) > 3 else "",
                'pais': safe_str(row.iloc[4]) if len(row) > 4 else "169",
            }

    # =====================================================================
    # CORRECCIÓN 8: Leer balance — incluir filas sin tercero para bancos
    # =====================================================================
    bal = []

    for _, row in df_balance.iterrows():
        cta = safe_str(row.iloc[CI])
        nit = safe_str(row.iloc[TI]) if TI < len(row) else ""
        if not cta:
            continue
        if '.' in nit:
            try: nit = str(int(float(nit)))
            except: pass

        # Permitir filas sin tercero para bancos (F1012 las necesita)
        if not nit:
            cuentas_sin_tercero_ok = ["1110", "1105"]
            if not any(cta.startswith(p) for p in cuentas_sin_tercero_ok):
                continue

        bal.append({
            'cta': cta,
            'nom_cta': safe_str(row.iloc[NI]) if NI < len(row) else "",
            'td': detectar_tipo_doc(nit) if nit else "",
            'nit': nit,
            'razon': safe_str(row.iloc[RI]) if RI < len(row) else "",
            'deb': safe_num(row.iloc[DI]) if DI < len(row) else 0,
            'cred': safe_num(row.iloc[KI]) if KI < len(row) else 0,
            'saldo': safe_num(row.iloc[SI]) if SI is not None and SI < len(row) else 0,
        })

    def valor_impuesto(f, tipo='activo'):
        if cierra_impuestos:
            return abs(f['saldo'])
        else:
            if tipo == 'activo':
                return max(f['deb'] - f['cred'], 0)
            else:
                return max(f['cred'] - f['deb'], 0)

    if dir_central is None:
        dir_central = {}

    direc = {}
    nits_nuevos = {}

    for f in bal:
        if not f['nit']:
            continue
        if f['nit'] not in direc:
            td = f['td'] if f['td'] else detectar_tipo_doc(f['nit'])
            r = f['razon']
            dv = calc_dv(f['nit'])
            d = {'td': td, 'dv': dv,
                 'a1': '', 'a2': '', 'n1': '', 'n2': '',
                 'rs': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
            if td == "13":
                p = r.split()
                if len(p) >= 1: d['a1'] = p[0]
                if len(p) >= 2: d['a2'] = p[1]
                if len(p) >= 3: d['n1'] = p[2]
                if len(p) >= 4: d['n2'] = ' '.join(p[3:])
            else:
                d['rs'] = r

            if f['nit'] in dir_central:
                dc = dir_central[f['nit']]
                if dc.get('dir'): d['dir'] = dc['dir']
                if dc.get('depto'): d['dp'] = pad_dpto(dc['depto'])
                if dc.get('mpio'): d['mp'] = pad_mpio(dc['mpio'])
                if dc.get('pais'): d['pais'] = dc['pais']
                if dc.get('td'): d['td'] = dc['td']
                if dc.get('dv'): d['dv'] = dc['dv']
                if dc.get('razon') and not d['rs'] and td != "13": d['rs'] = dc['razon']

            if f['nit'] in dir_externo:
                ext = dir_externo[f['nit']]
                if ext['dir']: d['dir'] = ext['dir']
                if ext['dp']: d['dp'] = pad_dpto(ext['dp'])
                if ext['mp']: d['mp'] = pad_mpio(ext['mp'])
                if ext.get('pais'): d['pais'] = ext['pais']

            if d['dir'] and f['nit'] not in dir_central and f['nit'] != NM:
                nits_nuevos[f['nit']] = {
                    'razon': d['rs'] or r, 'dir': d['dir'],
                    'dp': d.get('dp', ''), 'mp': d.get('mp', ''),
                    'pais': d.get('pais', '169'), 'td': d['td'], 'dv': d['dv'],
                }

            direc[f['nit']] = d

    direc[NM] = {'td': TDM, 'dv': '', 'a1': '', 'a2': '', 'n1': '', 'n2': '',
                 'rs': 'CUANTIAS MENORES', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}

    def t(nit):
        return direc.get(nit, {'td': detectar_tipo_doc(nit), 'dv': calc_dv(nit),
                                'a1': '', 'a2': '', 'n1': '', 'n2': '',
                                'rs': nit, 'dir': '', 'dp': '', 'mp': '', 'pais': '169'})

    # === CREAR WORKBOOK ===
    wb = openpyxl.Workbook()
    hf = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    thin = Side(style='thin', color='808080')

    def nueva_hoja(nombre, headers):
        ws = wb.create_sheet(nombre)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = hfont; cell.fill = hf
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        return ws

    def escribir_tercero(ws, fila, col, nit, con_pais=False):
        d = t(nit)
        c = col
        td_val = d.get('td', '') or detectar_tipo_doc(nit)
        dv_val = d.get('dv', '') or calc_dv(nit)
        valores = [td_val, nit, dv_val, d['a1'], d['a2'], d['n1'], d['n2'], d['rs'], d['dir'], d['dp'], d['mp']]
        for v in valores:
            cell = ws.cell(fila, c)
            cell.value = str(v) if v else ""
            cell.number_format = '@'
            c += 1
        if con_pais:
            cell = ws.cell(fila, c)
            cell.value = str(d.get('pais', '169') or "169")
            cell.number_format = '@'
            c += 1
        return c

    def fmt(ws, fila, cols):
        for c in cols:
            ws.cell(fila, c).number_format = '#,##0'

    def zebra(ws, fila):
        for c in range(1, ws.max_column + 1):
            ws.cell(fila, c).font = Font(size=10, name='Arial')
            ws.cell(fila, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        if (fila - 2) % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(fila, c).fill = PatternFill('solid', fgColor='F2F7FB')

    resultados = {}

    # ========== PRE-CALCULAR RETENCIONES Y BASES POR NIT ==========
    ret_fte_por_nit = defaultdict(float)
    ret_iva_por_nit = defaultdict(float)
    gastos_por_nit = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        cta = f['cta']
        if en_rango(cta, "236505", "236530"):
            ret_fte_por_nit[f['nit']] += valor_impuesto(f, 'pasivo')
        elif en_rango(cta, "2367", "2367"):
            ret_iva_por_nit[f['nit']] += valor_impuesto(f, 'pasivo')
        if cta[:2] in ("51", "52", "53") and abs(f['saldo']) > 0:
            gastos_por_nit[f['nit']] += abs(f['saldo'])

    # ========== F1001 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Pago Deducible", "Pago No Deducible",
         "IVA Ded", "IVA No Ded", "Ret Fte Renta", "Ret Fte Asumida", "Ret IVA R.Comun", "Ret IVA No Dom"]
    ws = nueva_hoja("F1001 Pagos", h)

    def clasificar_deducibilidad(cta, nom_cta=""):
        nom = normalizar_nombre(nom_cta)
        if 'gmf' in nom or '4x1000' in nom or '4 x 1000' in nom or 'gravamen movimiento' in nom:
            return 'gmf'
        if 'interes moratorio' in nom or 'interes de mora' in nom or 'intereses mora' in nom:
            return 'no_ded'
        if cta.startswith('53050504'): return 'no_ded'
        if 'no deducible' in nom or 'no deduci' in nom: return 'no_ded'
        if cta.startswith('53950520'): return 'no_ded'
        if 'multa' in nom or 'sancion' in nom or 'litigio' in nom: return 'no_ded'
        if cta.startswith('539520'): return 'no_ded'
        if 'donacion' in nom or 'donaciones' in nom: return 'no_ded'
        if cta.startswith('539525'): return 'no_ded'
        return 'ded'

    dic = defaultdict(lambda: [0.0] * 5)
    nits_en_1001 = set()
    CONCEPTOS_NOMINA = {"5001", "5024", "5025", "5027", "5023"}
    nits_pila_persona = []  # Para alertar sobre personas en conceptos de entidad
    for f in bal:
        if not f['nit']: continue
        valor = abs(f['saldo'])
        if valor == 0: continue
        conc, ded = concepto_1001(f['cta'], f.get('nom_cta', ''))
        if not conc or conc in CONCEPTOS_NOMINA: continue

        # Conceptos 5011, 5012, 5013, etc. solo para entidades (NIT jurídico)
        if conc in CONCEPTOS_SOLO_ENTIDADES:
            td = detectar_tipo_doc(f['nit'])
            if td == '13':  # Persona natural → reclasificar a 5016
                nits_pila_persona.append((f['nit'], conc, valor))
                conc = '5016'

        k = (conc, f['nit'])
        tipo_ded = clasificar_deducibilidad(f['cta'], f.get('nom_cta', ''))
        if tipo_ded == 'gmf':
            dic[k][0] += valor * 0.5
            dic[k][1] += valor * 0.5
        elif tipo_ded == 'no_ded':
            dic[k][1] += valor
        else:
            if ded: dic[k][0] += valor
            else: dic[k][1] += valor
        nits_en_1001.add(f['nit'])

    nit_conceptos = defaultdict(list)
    for (conc, nit), v in dic.items():
        nit_conceptos[nit].append((conc, v[0] + v[1]))
    for nit in nit_conceptos:
        ret_total = ret_fte_por_nit.get(nit, 0)
        ret_iva = ret_iva_por_nit.get(nit, 0)
        if ret_total == 0 and ret_iva == 0: continue
        total_pagado = sum(v for _, v in nit_conceptos[nit])
        if total_pagado == 0: continue
        for conc, pago in nit_conceptos[nit]:
            proporcion = pago / total_pagado if total_pagado > 0 else 0
            dic[(conc, nit)][2] += ret_total * proporcion
            dic[(conc, nit)][4] += ret_iva * proporcion

    final = {}
    menores = defaultdict(lambda: [0.0] * 5)
    for (c, n), v in dic.items():
        total_pago = v[0] + v[1]
        if total_pago == 0: continue
        tiene_retencion = v[2] > 0 or v[4] > 0
        if total_pago < C3UVT and not tiene_retencion:
            for i in range(5): menores[(c, NM)][i] += v[i]
        else:
            final[(c, n)] = v
    for k, v in menores.items():
        if v[0] + v[1] > 0:
            if k not in final: final[k] = v

    fila = 2
    for (conc, nit), v in sorted(final.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit, True)
        ws.cell(fila, 14).value = int(v[0])
        ws.cell(fila, 15).value = int(v[1])
        ws.cell(fila, 16).value = 0
        ws.cell(fila, 17).value = 0
        ws.cell(fila, 18).value = int(v[2])
        ws.cell(fila, 19).value = 0
        ws.cell(fila, 20).value = int(v[4])
        ws.cell(fila, 21).value = 0
        fmt(ws, fila, range(14, 22))
        zebra(ws, fila)
        fila += 1
    resultados['F1001 Pagos'] = len(final)

    # Si no es PRO, reemplazar hoja F1001 con mensaje
    if not es_pro:
        wb.remove(wb["F1001 Pagos"])
        ws_pro = wb.create_sheet("F1001 Pagos (PRO)", 0)
        ws_pro.append(["⚠️ El formato F1001 Pagos requiere suscripción PRO"])
        ws_pro.append([""])
        ws_pro.append(["Suscríbete en: https://exogenadian.com/#planes"])
        ws_pro.append(["Precio: $14.500/mes — Acceso a todas las herramientas"])
        ws_pro.append([""])
        ws_pro.append([f"Se detectaron {len(final)} registros para el F1001 que se generarán con PRO."])
        resultados['F1001 Pagos'] = f'🔒 PRO ({len(final)} registros)'

    # =====================================================================
    # F1003 — Retenciones que le practicaron
    # Solo cta 1355 con terceros reales (excluye DIAN, ICA, saldos a favor)
    # Si saldo = 0: usar débitos (la retención se cruzó en el año)
    # Si saldo > 0: usar saldo (pendiente de cruzar)
    # =====================================================================
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Base Retencion", "Retencion Acumulada"]
    ws = nueva_hoja("F1003 Retenciones", h)

    # NITs a excluir del F1003 (DIAN, entes territoriales)
    NITS_EXCLUIR_1003 = {NIT_DIAN, "899999090", "899999063"}  # DIAN, Bogotá, etc.

    dic3 = defaultdict(lambda: [0.0, 0.0])
    for f in bal:
        if not f['nit']: continue
        cta = f['cta']
        if not cta.startswith('1355'): continue
        # Solo subcuentas detalle (mín 6 dígitos para nivel de concepto)
        if len(cta) < 6: continue
        # Excluir: ICA (135518), saldos a favor (135595), autorretenciones (135599)
        if cta.startswith('135518'): continue   # ReteICA → no va en F1003
        if cta.startswith('135595'): continue   # Saldos a favor
        if cta.startswith('135599'): continue   # Autorretenciones
        # Excluir terceros institucionales (DIAN, entes territoriales)
        if f['nit'] in NITS_EXCLUIR_1003: continue

        # Lógica de valor:
        saldo = abs(f['saldo'])
        deb = f['deb']
        if saldo > 0:
            val = saldo    # Tiene saldo pendiente → reportar saldo
        elif deb > 0:
            val = deb      # Saldo en 0 pero tuvo movimiento → reportar débitos
        else:
            continue       # Sin saldo ni movimiento → no reportar

        # Buscar concepto por rango de subcuenta
        conc = buscar_concepto(cta, PARAM_1003, f.get('nom_cta', ''), KEYWORDS_1003)
        if not conc: conc = "1308"  # Otras retenciones como default

        dic3[(conc, f['nit'])][1] += val

    ingresos_por_nit = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        if f['cta'][:1] == "4" and abs(f['saldo']) > 0:
            ingresos_por_nit[f['nit']] += abs(f['saldo'])
    for (conc, nit), v in dic3.items():
        if v[1] > 0:
            v[0] = ingresos_por_nit.get(nit, 0)
    dic3 = {k: v for k, v in dic3.items() if v[1] > 0}

    fila = 2
    for (conc, nit), v in sorted(dic3.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(v[0])
        ws.cell(fila, 14).value = int(v[1])
        fmt(ws, fila, [13, 14])
        zebra(ws, fila)
        fila += 1
    resultados['F1003 Retenciones'] = len(dic3)

    # ========== F1005 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "IVA Descontable", "IVA Devol Ventas"]
    ws = nueva_hoja("F1005 IVA Descontable", h)

    dic5 = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        if en_rango(f['cta'], "2408", "2408"):
            nom = normalizar_nombre(f.get('nom_cta', ''))
            es_descontable = 'descontable' in nom or f['cta'][:6] >= '240810'
            if es_descontable:
                val = valor_impuesto(f, 'activo')
                if val > 0:
                    dic5[f['nit']] += val

    fila = 2
    for nit, val in sorted(dic5.items()):
        escribir_tercero(ws, fila, 1, nit)
        ws.cell(fila, 12).value = int(val)
        ws.cell(fila, 13).value = 0
        fmt(ws, fila, [12, 13])
        zebra(ws, fila)
        fila += 1
    resultados['F1005 IVA Descontable'] = len(dic5)

    # ========== F1006 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "IVA Generado", "IVA Devol Compras", "Imp Consumo"]
    ws = nueva_hoja("F1006 IVA Generado", h)

    dic6 = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        if en_rango(f['cta'], "2408", "2408"):
            nom = normalizar_nombre(f.get('nom_cta', ''))
            es_descontable = 'descontable' in nom or f['cta'][:6] >= '240810'
            if not es_descontable:
                val = valor_impuesto(f, 'pasivo')
                if val > 0:
                    dic6[f['nit']] += val

    fila = 2
    for nit, val in sorted(dic6.items()):
        escribir_tercero(ws, fila, 1, nit)
        ws.cell(fila, 12).value = int(val)
        ws.cell(fila, 13).value = 0
        ws.cell(fila, 14).value = 0
        fmt(ws, fila, [12, 13, 14])
        zebra(ws, fila)
        fila += 1
    resultados['F1006 IVA Generado'] = len(dic6)

    # ========== F1007 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Ingresos Brutos", "Devoluciones"]
    ws = nueva_hoja("F1007 Ingresos", h)

    dic7 = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        conc = buscar_concepto(f['cta'], PARAM_1007, f.get('nom_cta', ''), KEYWORDS_1007)
        if not conc: continue
        valor = abs(f['saldo'])
        if valor > 0:
            dic7[(conc, f['nit'])] += valor

    final7 = {}
    men7 = defaultdict(float)
    for (c, n), v in dic7.items():
        if v < C3UVT: men7[(c, NM)] += v
        else: final7[(c, n)] = v
    for k, v in men7.items():
        if k not in final7: final7[k] = v

    fila = 2
    for (conc, nit), val in sorted(final7.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit, True)
        ws.cell(fila, 14).value = int(val)
        ws.cell(fila, 15).value = 0
        fmt(ws, fila, [14, 15])
        zebra(ws, fila)
        fila += 1
    resultados['F1007 Ingresos'] = len(final7)

    # ========== F1008 (CORREGIDO: sin 1355) ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Saldo CxC Dic31"]
    ws = nueva_hoja("F1008 CxC", h)

    dic8 = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        conc = buscar_concepto(f['cta'], PARAM_1008, f.get('nom_cta', ''))
        if not conc: continue
        s = abs(f['saldo'])
        if s == 0: continue
        dic8[(conc, f['nit'])] += s

    final8 = {}
    men8 = defaultdict(float)
    for (c, n), v in dic8.items():
        if v < C12UVT: men8[(c, NM)] += v
        else: final8[(c, n)] = v
    for k, v in men8.items():
        if k not in final8: final8[k] = v

    fila = 2
    for (conc, nit), val in sorted(final8.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(val)
        fmt(ws, fila, [13])
        zebra(ws, fila)
        fila += 1
    resultados['F1008 CxC'] = len(final8)

    # =====================================================================
    # F1009 — Cuentas por Pagar (TODO el pasivo — clase 2)
    # DIAN: usa saldo del resumen 4 dígitos (ya neteado con pagos/anticipos)
    # Resto: detalle con tercero, solo saldos crédito (< 0)
    # Saldo = 0 → no se reporta
    # =====================================================================
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Saldo CxP Dic31"]
    ws = nueva_hoja("F1009 CxP", h)

    PREFIJOS_DIAN_F1009 = ("2365", "2367", "2368", "2370", "2404", "2408", "2412")

    if NIT_DIAN not in direc:
        direc[NIT_DIAN] = {
            'td': '31', 'dv': calc_dv(NIT_DIAN),
            'a1': '', 'a2': '', 'n1': '', 'n2': '',
            'rs': 'DIRECCIÓN DE IMPUESTOS Y ADUANAS NACIONALES - DIAN',
            'dir': 'CRA 8 # 6C-38', 'dp': '11', 'mp': '11001', 'pais': '169',
        }

    # Paso 1: DIAN — Leer saldos de cuentas resumen (4 dígitos, sin NIT)
    # Estos ya están neteados: retención causada - pagos realizados = saldo real
    dian_total_f1009 = 0
    for _, row in df_balance.iterrows():
        cta_raw = safe_str(row.iloc[CI]).replace('.', '').strip()
        nit_raw = safe_str(row.iloc[TI]).strip()
        if nit_raw and nit_raw != 'nan' and nit_raw != '0': continue  # Solo filas resumen
        if len(cta_raw) != 4: continue
        if cta_raw not in ('2365', '2367', '2370', '2404', '2408', '2412'): continue
        saldo_v = safe_num(row.iloc[SI]) if SI is not None and SI < len(row.index) else 0
        if saldo_v < 0:  # Saldo crédito = pasivo pendiente
            dian_total_f1009 += abs(saldo_v)

    # Paso 2: No-DIAN — Detalle con tercero, saldos con signo, netear por NIT
    dic9_signed = defaultdict(float)
    for f in bal:
        cta = f['cta']
        if cta[:1] != '2': continue
        s = f['saldo']
        if s == 0: continue
        if not f['nit']: continue
        # Excluir cuentas DIAN (ya van del resumen)
        es_dian = any(cta.startswith(p) for p in PREFIJOS_DIAN_F1009)
        if es_dian: continue
        conc = buscar_concepto(cta, PARAM_1009, f.get('nom_cta', ''))
        if not conc: conc = "2210"
        dic9_signed[(conc, f['nit'])] += s

    # Solo reportar saldos netos crédito (< 0)
    dic9 = {}
    for k, v in dic9_signed.items():
        if v < 0:
            dic9[k] = abs(v)

    # Agregar DIAN consolidado
    if dian_total_f1009 > 0:
        dic9[("2206", NIT_DIAN)] = dian_total_f1009

    final9 = {}
    men9 = defaultdict(float)
    for (c, n), v in dic9.items():
        if n == NIT_DIAN:
            final9[(c, n)] = final9.get((c, n), 0) + v
        elif v < C12UVT:
            men9[(c, NM)] += v
        else:
            final9[(c, n)] = v
    for k, v in men9.items():
        if k not in final9: final9[k] = v

    fila = 2
    for (conc, nit), val in sorted(final9.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(val)
        fmt(ws, fila, [13])
        zebra(ws, fila)
        fila += 1
    resultados['F1009 CxP'] = len(final9)

    # ========== F1010 (INFO: mejor solicitar como adicional) ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Valor Patrimonial", "% Participacion", "Valor Porcentual"]
    ws = nueva_hoja("F1010 Socios", h)

    dic10 = defaultdict(float)
    for f in bal:
        if not f['nit']: continue
        if en_rango(f['cta'], "3105", "3115") or en_rango(f['cta'], "3110", "3110"):
            dic10[f['nit']] += abs(f['saldo'])

    capital_total = sum(dic10.values())
    fila = 2
    for nit, val in sorted(dic10.items()):
        escribir_tercero(ws, fila, 1, nit, True)
        ws.cell(fila, 13).value = int(val)
        pct = round(val / capital_total * 100, 2) if capital_total > 0 else 0
        ws.cell(fila, 14).value = pct / 100
        ws.cell(fila, 14).number_format = '0.00%'
        ws.cell(fila, 15).value = int(val)
        fmt(ws, fila, [13, 15])
        zebra(ws, fila)
        fila += 1
    resultados['F1010 Socios'] = len(dic10)

    # =====================================================================
    # CORRECCIÓN 7: F1012 — Detectar bancos por nombre de subcuenta
    # =====================================================================
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Saldo Dic31", "Valor Patrimonial"]
    ws = nueva_hoja("F1012 Inversiones", h)

    dic12 = defaultdict(float)
    for f in bal:
        cta = f['cta']
        saldo = abs(f['saldo'])
        if saldo == 0: continue

        # Bancos nacionales (1110): detectar por nombre si no tiene tercero
        if en_rango(cta, "1110", "1110"):
            nit = f['nit']
            if not nit:
                nom_lower = normalizar_nombre(f.get('nom_cta', ''))
                for keyword, (nit_banco, rs_banco) in BANCOS_COLOMBIANOS.items():
                    if keyword in nom_lower:
                        nit = nit_banco
                        if nit not in direc:
                            direc[nit] = {
                                'td': '31', 'dv': calc_dv(nit),
                                'a1': '', 'a2': '', 'n1': '', 'n2': '',
                                'rs': rs_banco, 'dir': '', 'dp': '',
                                'mp': '', 'pais': '169'
                            }
                        break
                if not nit:
                    # Banco sin tercero identificado → incluir con NM para diligenciar después
                    nit = NM
            dic12[("8301", nit)] += saldo
            continue

        # Caja (1105): incluir aunque no tenga tercero
        if en_rango(cta, "1105", "1105"):
            nit = f['nit'] if f['nit'] else NM
            dic12[("8302", nit)] += saldo
            continue

        # Resto de inversiones
        if not f['nit']: continue
        for conc, d, h2 in MAPEO_1012:
            if en_rango(cta, d, h2):
                dic12[(conc, f['nit'])] += saldo
                break

    fila = 2
    for (conc, nit), val in sorted(dic12.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 10).value = int(val)
        ws.cell(fila, 11).value = int(val)
        fmt(ws, fila, [10, 11])
        zebra(ws, fila)
        fila += 1
    resultados['F1012 Inversiones'] = len(dic12)

    # ========== F2276 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Direccion", "Dpto", "Mpio", "Pais", "Salarios", "Emol Ecles", "Honor 383",
         "Serv 383", "Comis 383", "Pensiones", "Vacaciones", "Cesantias e Int",
         "Incapacidades", "Otros Pag Lab", "Total Bruto", "Aporte Salud", "Aporte Pension",
         "Sol Pensional", "Vol Empleador", "Vol Trabajador", "AFC", "Ret Fte", "Total Pagos"]
    ws = nueva_hoja("F2276 Rentas Trabajo", h)

    dic26 = defaultdict(lambda: [0.0] * 19)
    # Índices del array → columnas Excel:
    # 0=Salarios, 1=EmolEcles, 2=Honor383, 3=Serv383, 4=Comis383,
    # 5=Pensiones, 6=Vacaciones, 7=CesantíaseInt, 8=Incapacidades,
    # 9=OtrosPagLab, 10=TotalBruto, 11=AporteSalud, 12=AportePension,
    # 13=SolPensional, 14=VolEmpleador, 15=VolTrabajador, 16=AFC,
    # 17=RetFte, 18=TotalPagos
    for f in bal:
        if not f['nit']: continue
        if not en_rango(f['cta'], "5105", "5105"): continue
        valor = abs(f['saldo'])
        if valor == 0: continue
        nit = f['nit']
        nom = normalizar_nombre(f.get('nom_cta', ''))
        sc = f['cta'][4:6] if len(f['cta']) >= 6 else ""
        clasificado = False

        # Subcuentas por código PUC
        if sc in ("03",):
            # 510503 = Salario integral → va en Salarios
            dic26[nit][0] += valor; clasificado = True
        elif sc in ("06", "07", "08", "09", "10", "15"):
            # Sueldos, horas extra, recargos, auxilio transporte
            dic26[nit][0] += valor; clasificado = True
        elif sc in ("27",):
            # Auxilio de transporte → Salarios (hace parte del ingreso laboral)
            dic26[nit][0] += valor; clasificado = True
        elif sc in ("30", "33"):
            # Cesantías e intereses → [7]
            dic26[nit][7] += valor; clasificado = True
        elif sc in ("36",):
            # Prima de servicios → [9] Otros pagos laborales
            dic26[nit][9] += valor; clasificado = True
        elif sc in ("39",):
            # Vacaciones → [6]
            dic26[nit][6] += valor; clasificado = True
        elif sc in ("42", "45"):
            # Bonificaciones, dotación → [9] Otros pagos laborales
            dic26[nit][9] += valor; clasificado = True
        elif sc in ("01", "05"):
            # Honorarios a personas naturales → [2] Honor 383
            d2 = t(nit)
            if d2['td'] == "13":
                dic26[nit][2] += valor; clasificado = True
        elif sc in ("02",):
            # Aportes a salud (EPS) → [11] Aporte Salud
            dic26[nit][11] += valor; clasificado = True
        elif sc in ("04",):
            # Aportes a pensión → [12] Aporte Pension
            dic26[nit][12] += valor; clasificado = True
        elif sc in ("68", "72", "75"):
            # Parafiscales (ICBF, SENA, Cajas) → no van en F2276
            clasificado = True

        if not clasificado and nom:
            palabras = set(nom.split())
            if any(kw in nom for kw in ["salario integral", "integral"]):
                dic26[nit][0] += valor
            elif any(kw in palabras for kw in ["sueldo", "salario", "basico", "jornal"]) or \
               any(kw in nom for kw in ["hora extra", "horas extra", "recargo"]):
                dic26[nit][0] += valor
            elif any(kw in nom for kw in ["cesantia", "interes sobre cesantia", "intereses cesantia"]):
                dic26[nit][7] += valor
            elif any(kw in palabras for kw in ["vacacion", "vacaciones"]):
                dic26[nit][6] += valor
            elif any(kw in nom for kw in ["prima de servicio", "prima servicio"]):
                dic26[nit][9] += valor
            elif any(kw in palabras for kw in ["incapacidad", "incapacidades"]):
                dic26[nit][8] += valor
            elif any(kw in nom for kw in ["aporte salud", "aporte eps", "aportes eps", "aportes a eps"]):
                dic26[nit][11] += valor
            elif any(kw in nom for kw in ["aporte pension", "aportes pension", "aportes a pension"]):
                dic26[nit][12] += valor
            elif any(kw in palabras for kw in ["dotacion", "bonificacion", "auxilio"]):
                dic26[nit][9] += valor
            elif any(kw in palabras for kw in ["honorario", "honorarios"]):
                d2 = t(nit)
                if d2['td'] == "13": dic26[nit][2] += valor
                else: dic26[nit][9] += valor
            elif any(kw in palabras for kw in ["parafiscal", "parafiscales", "icbf", "sena",
                                                "compensar", "comfama", "cafam"]):
                pass
            else:
                dic26[nit][9] += valor

    for f in bal:
        if not f['nit']: continue
        if en_rango(f['cta'], "2365", "2365") and f['nit'] in dic26:
            dic26[f['nit']][17] += valor_impuesto(f, 'pasivo')
    for nit in dic26:
        dic26[nit][10] = sum(dic26[nit][:10])   # Total Bruto = sum(Salarios..OtrosPagLab)
        dic26[nit][18] = dic26[nit][10]          # Total Pagos = Total Bruto

    fila = 2
    for nit, v in sorted(dic26.items()):
        d = t(nit)
        for ci, val in [(1, d['td']), (2, nit), (3, d['dv']),
                        (4, d['a1']), (5, d['a2']), (6, d['n1']), (7, d['n2']),
                        (8, d['dir']), (9, d['dp']), (10, d['mp']),
                        (11, d.get('pais', '169') or '169')]:
            cell = ws.cell(fila, ci)
            cell.value = val
            cell.number_format = '@'
        for i in range(19):
            ws.cell(fila, 12 + i).value = int(v[i])
            ws.cell(fila, 12 + i).number_format = '#,##0'
        zebra(ws, fila)
        fila += 1
    resultados['F2276 Rentas Trabajo'] = len(dic26)

    # Si no es PRO, reemplazar hoja F2276 con mensaje
    if not es_pro:
        wb.remove(wb["F2276 Rentas Trabajo"])
        ws_pro2 = wb.create_sheet("F2276 Rentas Trabajo (PRO)")
        ws_pro2.append(["⚠️ El formato F2276 Rentas de Trabajo requiere suscripción PRO"])
        ws_pro2.append([""])
        ws_pro2.append(["Suscríbete en: https://exogenadian.com/#planes"])
        ws_pro2.append(["Precio: $14.500/mes — Acceso a todas las herramientas"])
        ws_pro2.append([""])
        ws_pro2.append([f"Se detectaron {len(dic26)} registros para el F2276 que se generarán con PRO."])
        resultados['F2276 Rentas Trabajo'] = f'🔒 PRO ({len(dic26)} registros)'

    # ========== RESUMEN ==========
    n_con_dir = sum(1 for d in direc.values() if d.get('dir', ''))
    n_de_central = sum(1 for nit in direc if nit in dir_central and dir_central.get(nit, {}).get('dir', ''))
    n_de_cliente = sum(1 for nit in direc if nit in dir_externo and dir_externo.get(nit, {}).get('dir', ''))

    wsr = wb.active
    wsr.title = "Resumen"
    wsr['A1'] = "RESUMEN PROCESAMIENTO EXOGENA AG 2025"
    wsr['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4E79')
    wsr['A3'] = "Fecha:"; wsr['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    wsr['A4'] = "Filas del balance:"; wsr['B4'] = len(bal)
    wsr['A5'] = "Terceros:"; wsr['B5'] = len(direc)
    wsr['A6'] = "Con dirección:"; wsr['B6'] = n_con_dir
    wsr['A7'] = "  → Del directorio centralizado:"; wsr['B7'] = n_de_central
    wsr['A7'].font = Font(size=9, name='Arial', color='666666')
    wsr['A8'] = "  → Del directorio del cliente:"; wsr['B8'] = n_de_cliente
    wsr['A8'].font = Font(size=9, name='Arial', color='666666')
    rr = 9
    wsr.cell(rr, 1).value = "Direcciones nuevas para agregar:"
    wsr.cell(rr, 2).value = len(nits_nuevos)
    wsr.cell(rr, 1).font = Font(size=9, name='Arial', color='0066CC')
    rr += 1

    fila_inicio_formatos = rr + 1
    wsr.cell(fila_inicio_formatos, 1).value = "Formato"
    wsr.cell(fila_inicio_formatos, 2).value = "Registros"
    wsr.cell(fila_inicio_formatos, 1).font = Font(bold=True)
    wsr.cell(fila_inicio_formatos, 2).font = Font(bold=True)
    row = fila_inicio_formatos + 1
    for nombre, n in resultados.items():
        wsr.cell(row, 1).value = nombre; wsr.cell(row, 2).value = n; row += 1
    wsr.cell(row + 1, 1).value = "TOTAL"; wsr.cell(row + 1, 1).font = Font(bold=True)
    wsr.cell(row + 1, 2).value = sum(resultados.values()); wsr.cell(row + 1, 2).font = Font(bold=True)
    wsr.cell(row + 3, 1).value = "F1004, F1011, F1647: requieren datos manuales"
    wsr.cell(row + 4, 1).value = "F1010: se recomienda solicitar listado de socios como adicional"
    wsr.cell(row + 4, 1).font = Font(size=9, name='Arial', color='CC6600')

    # === ESTILOS PARA SECCIONES DE ADVERTENCIAS ===
    ROJO_BANNER  = 'C0392B'; ROJO_CLARO   = 'FADBD8'
    NARANJA_BANNER = 'E67E22'; NARANJA_CLARO  = 'FDF2E9'
    AZUL_BANNER  = '2471A3'; AZUL_CLARO   = 'D6EAF8'
    VERDE_BANNER = '1D8348'; VERDE_CLARO  = 'D5F5E3'
    GRIS_TEXTO   = '2C3E50'; GRIS_CLARO   = 'F8F9F9'
    BLANCO       = 'FFFFFF'

    border_thin = Border(
        left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'))
    border_bottom_thick = Border(bottom=Side(style='medium', color='ABB2B9'))

    def banner_row(ws, r, texto, color_fondo, cols=3):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c); cell.fill = PatternFill('solid', fgColor=color_fondo)
            cell.font = Font(bold=True, size=12, name='Calibri', color=BLANCO)
            cell.alignment = Alignment(vertical='center'); cell.border = border_thin
        ws.cell(r, 1).value = texto
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        ws.row_dimensions[r].height = 30

    def sub_banner(ws, r, texto, color_fondo, color_texto='FFFFFF', cols=3):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c); cell.fill = PatternFill('solid', fgColor=color_fondo)
            cell.font = Font(bold=True, size=10, name='Calibri', color=color_texto)
            cell.alignment = Alignment(vertical='center'); cell.border = border_thin
        ws.cell(r, 1).value = texto
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        ws.row_dimensions[r].height = 24

    def item_row(ws, r, titulo, detalle, color_alt, cols=3):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c); cell.fill = PatternFill('solid', fgColor=color_alt); cell.border = border_thin
        ws.cell(r, 1).value = titulo
        ws.cell(r, 1).font = Font(bold=True, size=10, name='Calibri', color=GRIS_TEXTO)
        ws.cell(r, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        r2 = r + 1
        for c in range(1, cols + 1):
            cell = ws.cell(r2, c); cell.fill = PatternFill('solid', fgColor=BLANCO); cell.border = border_thin
        ws.cell(r2, 1).value = detalle
        ws.cell(r2, 1).font = Font(size=9, name='Calibri', color='566573')
        ws.cell(r2, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=cols)
        ws.row_dimensions[r2].height = 45
        return r2 + 1

    def valor_row(ws, r, titulo, detalle, color_fondo, cols=3):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c); cell.fill = PatternFill('solid', fgColor=color_fondo); cell.border = border_thin
        ws.cell(r, 1).value = titulo
        ws.cell(r, 1).font = Font(bold=True, size=10, name='Calibri', color=GRIS_TEXTO)
        ws.cell(r, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        r2 = r + 1
        for c in range(1, cols + 1):
            cell = ws.cell(r2, c); cell.fill = PatternFill('solid', fgColor=BLANCO); cell.border = border_thin
        ws.cell(r2, 1).value = detalle
        ws.cell(r2, 1).font = Font(size=9, name='Calibri', color='566573')
        ws.cell(r2, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=cols)
        ws.row_dimensions[r2].height = 40
        return r2 + 1

    NUM_COLS = 3
    r = row + 6

    # Sección 0: Confidencialidad
    banner_row(wsr, r, "   COMPROMISO DE CONFIDENCIALIDAD Y PROTECCIÓN DE DATOS", VERDE_BANNER, NUM_COLS)
    r += 1
    confidencialidad = [
        ("🛡️  USO EXCLUSIVO DE LA INFORMACIÓN", "Los archivos suministrados por el cliente son utilizados ÚNICA Y EXCLUSIVAMENTE para la generación de los formatos de información exógena solicitados."),
        ("🗑️  ELIMINACIÓN INMEDIATA", "Una vez generada y entregada la información exógena al cliente, TODOS los archivos recibidos son ELIMINADOS de forma definitiva e irreversible."),
        ("🚫  NO DIVULGACIÓN A TERCEROS", "La información contable, tributaria y de terceros del cliente NO se comparte, vende, cede ni transfiere a ninguna persona natural o jurídica."),
        ("📋  MARCO LEGAL APLICABLE", "Ley 1581 de 2012, Ley 43 de 1990 Art. 63, ET Art. 583 y demás normas concordantes."),
    ]
    for i, (titulo, detalle) in enumerate(confidencialidad):
        alt_color = VERDE_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # Sección 1: Deslinde
    r += 2
    banner_row(wsr, r, "   DESLINDE DE RESPONSABILIDAD", ROJO_BANNER, NUM_COLS)
    r += 1
    disclaimer_lines = [
        "Esta herramienta es un ASISTENTE. Los resultados son un BORRADOR que requiere revisión profesional.",
        "Es responsabilidad del CONTADOR PÚBLICO y/o CONTRIBUYENTE verificar y validar toda la información.",
        "Ni el desarrollador ni la herramienta se hacen responsables por errores, omisiones o sanciones.",
        "El uso implica la aceptación total de estos términos. Art. 631 ET — Resolución DIAN 000227 de 2025.",
    ]
    for i, texto in enumerate(disclaimer_lines):
        bg = ROJO_CLARO if i % 2 == 0 else BLANCO
        for c in range(1, NUM_COLS + 1):
            cell = wsr.cell(r, c); cell.fill = PatternFill('solid', fgColor=bg); cell.border = border_thin
        wsr.cell(r, 1).value = texto
        wsr.cell(r, 1).font = Font(size=9, name='Calibri', color=GRIS_TEXTO)
        wsr.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='center')
        wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
        wsr.row_dimensions[r].height = 35; r += 1

    # Sección 2: Advertencias deducibilidad
    r += 2
    banner_row(wsr, r, "   ADVERTENCIAS — DEDUCIBILIDAD DE GASTOS", NARANJA_BANNER, NUM_COLS)
    r += 1
    sub_banner(wsr, r, "  El contador DEBE verificar los siguientes aspectos.", NARANJA_CLARO, GRIS_TEXTO, NUM_COLS)
    r += 1
    advertencias = [
        ("PAGOS EN EFECTIVO (Art. 771-5 ET)", "No deducibles los pagos en efectivo que superen el MENOR entre 40% del total pagado o 35% de costos y deducciones totales."),
        ("SEGURIDAD SOCIAL NO PAGADA (Art. 664, 108 ET)", "Salarios y prestaciones solo deducibles si el empleador está al día en aportes."),
        ("INDEPENDIENTES SIN SEGURIDAD SOCIAL (Art. 108 par. 2 ET)", "Pagos por honorarios/servicios a independientes NO deducibles si no se verificó el pago de aportes."),
        ("FACTURA ELECTRÓNICA / RADIAN (Art. 771-2, 616-1 ET)", "Costos y gastos deben estar soportados con factura electrónica o DSNO."),
        ("GMF / 4x1000 (Art. 115 ET)", "Deducible SOLO en un 50%. Esta herramienta ya clasifica el GMF automáticamente."),
        ("INTERESES DE MORA Y SANCIONES (Art. 107, 107-1 ET)", "NO son deducibles. Esta herramienta los clasifica automáticamente como no deducibles."),
        ("CAUSALIDAD, NECESIDAD Y PROPORCIONALIDAD (Art. 107 ET)", "La herramienta NO evalúa estos criterios — es responsabilidad del contador."),
    ]
    for i, (titulo, detalle) in enumerate(advertencias):
        alt_color = NARANJA_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {i+1}. {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # Sección 3: Cruces DIAN
    r += 2
    banner_row(wsr, r, "   CRUCES DE INFORMACIÓN — POSIBLES REQUERIMIENTOS DIAN", AZUL_BANNER, NUM_COLS)
    r += 1
    sub_banner(wsr, r, "  La DIAN cruza la información exógena con otras fuentes.", AZUL_CLARO, GRIS_TEXTO, NUM_COLS)
    r += 2
    sub_banner(wsr, r, "  A. VALIDACIONES AUTOMÁTICAS", VERDE_BANNER, BLANCO, NUM_COLS)
    r += 1

    total_ingresos_4 = sum(abs(f['saldo']) for f in bal if f['cta'][:1] == '4' and f.get('nit') and abs(f['saldo']) > 0)
    total_gastos_5 = sum(abs(f['saldo']) for f in bal if f['cta'][:2] in ('51','52','53') and f.get('nit') and abs(f['saldo']) > 0)
    total_costos_6 = sum(abs(f['saldo']) for f in bal if f['cta'][:1] == '6' and f.get('nit') and abs(f['saldo']) > 0)
    total_iva_desc = sum(v for v in dic5.values()) if dic5 else 0
    total_iva_gen = sum(v for v in dic6.values()) if dic6 else 0
    total_ret_fte_2365 = sum(v for v in ret_fte_por_nit.values())
    total_ret_iva_2367 = sum(v for v in ret_iva_por_nit.values())
    total_ret_1355 = sum(v[1] for v in dic3.values()) if dic3 else 0
    total_nomina = sum(abs(f['saldo']) for f in bal if en_rango(f['cta'], '5105', '5105') and f.get('nit') and abs(f['saldo']) > 0)
    total_f1001 = sum(v[0] + v[1] for v in dic.values())

    # Totales de balance para saldos (F1008, F1009, F1012)
    total_bal_cxc = sum(abs(f['saldo']) for f in bal
                        if f['cta'][:2] == '13' and not f['cta'].startswith('1355')
                        and f.get('nit') and abs(f['saldo']) > 0)
    # Total pasivo: leer directamente de la fila resumen "2" del balance
    total_bal_cxp = 0
    for _, row in df_balance.iterrows():
        cta_raw = safe_str(row.iloc[CI]).replace('.', '').strip()
        if cta_raw == '2':
            saldo_v = safe_num(row.iloc[SI]) if SI is not None and SI < len(row.index) else 0
            total_bal_cxp = abs(saldo_v)
            break
    total_bal_inv = sum(abs(f['saldo']) for f in bal
                        if (f['cta'][:4] in ('1105','1110','1115','1120') or f['cta'][:2] == '12')
                        and f.get('nit') and abs(f['saldo']) > 0)

    # Nómina completa que va a F2276 (para explicar diferencia F1001)
    total_nomina_f2276 = sum(abs(f['saldo']) for f in bal
                             if en_rango(f['cta'], '5101', '5110') and f.get('nit') and abs(f['saldo']) > 0)
    # Retención por salarios (Cta 2365 a empleados → diferencia en F1001 vs Cta 2365 total)
    total_ret_salarios = sum(abs(f['saldo']) for f in bal
                             if en_rango(f['cta'], '236505', '236505') and f.get('nit') and abs(f['saldo']) > 0)

    validaciones_auto = []
    if total_gastos_5 + total_costos_6 > 0 and total_iva_desc > 0:
        pct_iva = (total_iva_desc / (total_gastos_5 + total_costos_6)) * 100
        validaciones_auto.append((f"IVA Descontable / (Gastos + Costos) = {pct_iva:.1f}%",
            f"IVA Descontable: ${total_iva_desc:,.0f}  |  Gastos+Costos: ${total_gastos_5 + total_costos_6:,.0f}"))
    if total_f1001 > 0 and total_ret_fte_2365 > 0:
        pct_ret = (total_ret_fte_2365 / total_f1001) * 100
        validaciones_auto.append((f"Retención Fte Practicada / Pagos F1001 = {pct_ret:.1f}%",
            f"Ret. Fte (2365): ${total_ret_fte_2365:,.0f}  |  Pagos F1001: ${total_f1001:,.0f}"))
    if total_ingresos_4 > 0 and total_ret_1355 > 0:
        pct_ret_ing = (total_ret_1355 / total_ingresos_4) * 100
        validaciones_auto.append((f"Retenciones Recibidas / Ingresos = {pct_ret_ing:.1f}%",
            f"Ret. recibidas: ${total_ret_1355:,.0f}  |  Ingresos: ${total_ingresos_4:,.0f}"))
    if total_ingresos_4 > 0:
        validaciones_auto.append((f"Total Ingresos (4xxx): ${total_ingresos_4:,.0f}",
            "Debe coincidir con ingresos brutos de la declaración de renta."))
    if total_gastos_5 + total_costos_6 > 0:
        validaciones_auto.append((f"Total Gastos: ${total_gastos_5:,.0f}  |  Total Costos: ${total_costos_6:,.0f}",
            f"Suma: ${total_gastos_5 + total_costos_6:,.0f}. Debe coincidir con costos y deducciones de renta."))
    if total_nomina > 0:
        validaciones_auto.append((f"Total Nómina (5105): ${total_nomina:,.0f}",
            "Cruza F2276 vs PILA, Nómina electrónica y certificados F220."))

    for i, (titulo, detalle) in enumerate(validaciones_auto):
        alt_color = VERDE_CLARO if i % 2 == 0 else GRIS_CLARO
        r = valor_row(wsr, r, f"  ✓ {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    r += 1
    sub_banner(wsr, r, "  B. CRUCES QUE REALIZA LA DIAN (verificación manual)", AZUL_BANNER, BLANCO, NUM_COLS)
    r += 1
    cruces_dian = [
        ("EXÓGENA vs DECLARACIÓN DE IVA (Form. 300)", "IVA generado (F1006) vs IVA declarado | IVA descontable (F1005) vs IVA solicitado."),
        ("EXÓGENA vs DECLARACIÓN DE RENTA (Form. 110/210)", "Ingresos F1007 = Ingresos brutos renta | Pagos F1001 + Nómina F2276 = Costos y deducciones."),
        ("EXÓGENA vs RETENCIÓN EN LA FUENTE (Form. 350)", "Retenciones 2365 en F1001 deben coincidir mes a mes con el Form. 350."),
        ("EXÓGENA PROPIA vs EXÓGENA DE TERCEROS", "Su F1001 (pagos) debe coincidir con el F1007 (ingresos) del tercero."),
        ("EXÓGENA vs FACTURACIÓN ELECTRÓNICA (RADIAN)", "La DIAN cruza valores de exógena contra facturas en RADIAN."),
        ("EXÓGENA vs ENTIDADES FINANCIERAS", "Bancos reportan saldos, CDTs, inversiones, préstamos, intereses, GMF."),
        ("CUANTÍAS MENORES vs MATERIALIDAD", "Si cuantías menores (222222222) es muy alto vs total, la DIAN puede exigir identificación individual."),
    ]
    for i, (titulo, detalle) in enumerate(cruces_dian):
        alt_color = AZUL_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {i+1}. {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    r += 1
    banner_row(wsr, r, "   RECUERDE: Sanción por errores en exógena — Art. 651 ET", ROJO_BANNER, NUM_COLS)
    r += 1
    nota_sancion = ("La sanción puede ser hasta del 5% de los valores no reportados o reportados incorrectamente, "
                    "sin exceder de 15.000 UVT. Verifique TODA la información antes de presentarla.")
    for c in range(1, NUM_COLS + 1):
        cell = wsr.cell(r, c); cell.fill = PatternFill('solid', fgColor=ROJO_CLARO); cell.border = border_thin
    wsr.cell(r, 1).value = nota_sancion
    wsr.cell(r, 1).font = Font(size=10, name='Calibri', color=ROJO_BANNER, bold=True)
    wsr.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='center')
    wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    wsr.row_dimensions[r].height = 50
    wsr.column_dimensions['A'].width = 40; wsr.column_dimensions['B'].width = 30; wsr.column_dimensions['C'].width = 25

    for ws_name in wb.sheetnames:
        ws2 = wb[ws_name]
        if ws_name.startswith("F"):
            for col in range(1, ws2.max_column + 1):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
            ws2.freeze_panes = 'A2'

    # =====================================================================
    # HOJA: RESUMEN VALORES — CONFRONTACIÓN EXÓGENA vs BALANCE
    # (se mueve a 2da posición al final)
    # =====================================================================
    ws_rv = wb.create_sheet("Resumen Valores")

    rv_hf = PatternFill('solid', fgColor='1F4E79')
    rv_hfont = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    rv_thin = Side(style='thin', color='B0B0B0')
    rv_border = Border(top=rv_thin, bottom=rv_thin, left=rv_thin, right=rv_thin)
    rv_nfmt = '#,##0'
    rv_titulo_fill = PatternFill('solid', fgColor='D6EAF8')
    rv_ok_fill = PatternFill('solid', fgColor='D5F5E3')
    rv_warn_fill = PatternFill('solid', fgColor='FDF2E9')
    rv_err_fill = PatternFill('solid', fgColor='FADBD8')
    rv_nota_fill = PatternFill('solid', fgColor='FEF9E7')
    rv_alt1 = PatternFill('solid', fgColor='FFFFFF')
    rv_alt2 = PatternFill('solid', fgColor='F8F9FA')
    rv_sub_font = Font(size=9, name='Calibri', color='666666', italic=True)
    rv_norm_font = Font(size=10, name='Calibri')
    rv_bold_font = Font(size=10, name='Calibri', bold=True)

    rv_headers = ["Formato", "Concepto", "Regs", "Total Exógena",
                  "Cuenta Balance", "Total Balance", "Diferencia", "Estado"]
    NC_RV = len(rv_headers)
    for c, h_txt in enumerate(rv_headers, 1):
        cell = ws_rv.cell(1, c, h_txt)
        cell.font = rv_hfont; cell.fill = rv_hf
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = rv_border

    # Calcular totales por formato
    total_f1001_pagos = sum(v[0] + v[1] for v in final.values())
    total_f1001_retfte = sum(v[2] for v in final.values())
    total_f1001_retiva = sum(v[4] for v in final.values())
    total_f1003 = sum(v[1] for v in dic3.values()) if dic3 else 0
    total_f1005 = sum(dic5.values()) if dic5 else 0
    total_f1006 = sum(dic6.values()) if dic6 else 0
    total_f1007 = sum(final7.values()) if final7 else 0
    total_f1008 = sum(final8.values()) if final8 else 0
    total_f1009 = sum(final9.values()) if final9 else 0
    total_f1010 = sum(dic10.values()) if dic10 else 0
    total_f1012 = sum(dic12.values()) if dic12 else 0
    total_f2276 = sum(v[10] for v in dic26.values()) if dic26 else 0  # Total Bruto sin retenciones

    rv_row = 2

    def _rv_titulo(texto):
        nonlocal rv_row
        for c in range(1, NC_RV + 1):
            ws_rv.cell(rv_row, c).fill = rv_titulo_fill
            ws_rv.cell(rv_row, c).font = Font(bold=True, size=10, name='Calibri', color='1F4E79')
            ws_rv.cell(rv_row, c).border = rv_border
        ws_rv.cell(rv_row, 1).value = texto
        ws_rv.merge_cells(start_row=rv_row, start_column=1, end_row=rv_row, end_column=NC_RV)
        ws_rv.row_dimensions[rv_row].height = 26
        rv_row += 1

    def _rv_linea(formato, concepto, regs, val_exo, cta_bal, val_bal, es_sub=False):
        nonlocal rv_row
        alt = rv_alt1 if rv_row % 2 == 0 else rv_alt2

        ws_rv.cell(rv_row, 1).value = formato
        ws_rv.cell(rv_row, 2).value = concepto
        ws_rv.cell(rv_row, 3).value = regs if regs else None
        ws_rv.cell(rv_row, 4).value = int(val_exo) if val_exo else None
        ws_rv.cell(rv_row, 5).value = cta_bal
        ws_rv.cell(rv_row, 6).value = int(val_bal) if val_bal else None
        ws_rv.cell(rv_row, 4).number_format = rv_nfmt
        ws_rv.cell(rv_row, 6).number_format = rv_nfmt
        ws_rv.cell(rv_row, 3).alignment = Alignment(horizontal='center')

        # Calcular diferencia y estado
        if val_exo is not None and val_bal is not None and val_bal != "":
            dif = int(val_exo or 0) - int(val_bal or 0)
            ws_rv.cell(rv_row, 7).value = dif
            ws_rv.cell(rv_row, 7).number_format = rv_nfmt
            pct = abs(dif / val_bal * 100) if val_bal and val_bal != 0 else (100 if dif != 0 else 0)
            if abs(dif) <= 1000:
                estado = "✅ OK"; fill_e = rv_ok_fill
            elif pct <= 5:
                estado = "⚠️ Revisar"; fill_e = rv_warn_fill
            else:
                estado = "❌ Diferencia"; fill_e = rv_err_fill
            ws_rv.cell(rv_row, 8).value = estado
            ws_rv.cell(rv_row, 7).fill = fill_e
            ws_rv.cell(rv_row, 8).fill = fill_e
        else:
            ws_rv.cell(rv_row, 7).value = None
            ws_rv.cell(rv_row, 8).value = "ℹ️"

        fnt = rv_sub_font if es_sub else rv_norm_font
        for c in range(1, NC_RV + 1):
            if not ws_rv.cell(rv_row, c).fill or ws_rv.cell(rv_row, c).fill == PatternFill():
                ws_rv.cell(rv_row, c).fill = alt
            ws_rv.cell(rv_row, c).font = fnt
            ws_rv.cell(rv_row, c).border = rv_border
        rv_row += 1

    def _rv_nota(texto, color_fill=None):
        nonlocal rv_row
        fill = color_fill or rv_nota_fill
        for c in range(1, NC_RV + 1):
            ws_rv.cell(rv_row, c).fill = fill
            ws_rv.cell(rv_row, c).border = rv_border
        ws_rv.cell(rv_row, 1).value = texto
        ws_rv.cell(rv_row, 1).font = Font(size=9, name='Calibri', color='996600', italic=True)
        ws_rv.cell(rv_row, 1).alignment = Alignment(wrap_text=True)
        ws_rv.merge_cells(start_row=rv_row, start_column=1, end_row=rv_row, end_column=NC_RV)
        ws_rv.row_dimensions[rv_row].height = 35
        rv_row += 1

    # ========== SECCIÓN 1: INGRESOS, COSTOS Y GASTOS ==========
    _rv_titulo("📋 MOVIMIENTOS DEL AÑO (INGRESOS, COSTOS Y GASTOS)")
    _rv_linea("F1007", "Ingresos recibidos", len(final7),
              total_f1007, "Cta 4xxx", total_ingresos_4)
    dif_1007 = total_f1007 - total_ingresos_4
    if abs(dif_1007) > 1000 and total_ingresos_4 > 0:
        cuantias_menores_7 = total_ingresos_4 - total_f1007
        _rv_nota(f"↳ Diferencia F1007: ${dif_1007:,.0f}. Cta 4 del balance = ${total_ingresos_4:,.0f}, "
                 f"Exógena F1007 = ${total_f1007:,.0f}. Puede deberse a: cuantías menores agrupadas al NIT 222222222, "
                 f"ingresos sin tercero en el balance, o ajustes de cierre.")

    _rv_linea("F1001", "Pagos y abonos en cuenta", len(final),
              total_f1001_pagos, "Cta 5+6", total_gastos_5 + total_costos_6)
    dif_1001 = total_f1001_pagos - (total_gastos_5 + total_costos_6)
    if abs(dif_1001) > 1000:
        _rv_nota(f"↳ Diferencia F1001: ${dif_1001:,.0f}. Es normal porque el gasto de nómina "
                 f"(${total_nomina_f2276:,.0f}) va al F2276 y no al F1001. También las retenciones por "
                 f"salarios y los parafiscales se excluyen del F1001.")

    _rv_linea("F1001", "  → Retención Fte practicada", "",
              total_f1001_retfte, "Cta 2365", total_ret_fte_2365, es_sub=True)
    dif_ret = total_ret_fte_2365 - total_f1001_retfte
    if dif_ret > 1000:
        _rv_nota(f"↳ Cta 2365 (${total_ret_fte_2365:,.0f}) > RetFte F1001 (${total_f1001_retfte:,.0f}). "
                 f"Diferencia ${dif_ret:,.0f}: incluye retenciones por salarios (van en F2276) "
                 f"y posibles autorretenciones (Cta 236575/236540) que se declaran en Form 350 Rng 50-56, "
                 f"NO en exógena por tercero.")

    _rv_linea("F1001", "  → Retención IVA practicada", "",
              total_f1001_retiva, "Cta 2367", total_ret_iva_2367, es_sub=True)

    _rv_linea("F1003", "Retenciones que le practicaron", len(dic3),
              total_f1003, "Cta 1355 (débitos)", total_ret_1355)
    _rv_linea("F1005", "IVA Descontable", len(dic5),
              total_f1005, "Cta 2408 (compras)", total_iva_desc)
    _rv_linea("F1006", "IVA Generado", len(dic6),
              total_f1006, "Cta 2408 (ventas)", total_iva_gen)
    _rv_linea("F2276", "Rentas de trabajo y pensiones", len(dic26),
              total_f2276, "Cta 5105 (nómina)", total_nomina)
    # Calcular retención F2276
    total_ret_f2276 = sum(v[17] for v in dic26.values()) if dic26 else 0
    if total_ret_salarios > 0 or total_ret_f2276 > 0:
        _rv_nota(f"↳ F2276 Total Bruto: ${total_f2276:,.0f} (sin retenciones). "
                 f"Retenciones F2276: ${total_ret_f2276:,.0f}. "
                 f"Debe coincidir con RetFte por salarios (Cta 236505 = ${total_ret_salarios:,.0f}).")

    # Notas de seguridad social / PILA
    _rv_nota("⚠️ CONCEPTOS 5011, 5012, 5013, 5023-5027: Solo reportan entidades (EPS, AFP, ARL, Cajas). "
             "Estos valores deben coincidir con la planilla PILA. Verifique con el formato de pagos a la seguridad social.")

    # Nota autorretenciones
    posible_autoret = max(0, total_ret_fte_2365 - total_f1001_retfte)
    if posible_autoret > total_ret_fte_2365 * 0.1:  # Solo si es significativo
        _rv_nota(f"⚠️ AUTORRETENCIONES: Posibles autorretenciones por ${posible_autoret:,.0f}. "
                 f"Verifique Cta 236575/236540. Se declaran en Form 350 renglones 50-56, NO en exógena.")

    # ========== SECCIÓN 2: SALDOS A DICIEMBRE 31 ==========
    _rv_titulo("📋 SALDOS A DICIEMBRE 31 (PATRIMONIO)")
    _rv_linea("F1008", "Cuentas por cobrar", len(final8),
              total_f1008, "Ctas 13xx (sin 1355)", total_bal_cxc)
    dif_1008 = total_f1008 - total_bal_cxc
    if abs(dif_1008) > 1000 and total_bal_cxc > 0:
        _rv_nota(f"↳ Diferencia F1008: ${dif_1008:,.0f}. Ctas 13xx balance = ${total_bal_cxc:,.0f}, "
                 f"Exógena = ${total_f1008:,.0f}. Puede deberse a cuantías menores agrupadas o "
                 f"cuentas sin tercero.")

    _rv_linea("F1009", "Cuentas por pagar", len(final9),
              total_f1009, "Total Pasivo (cta 2)", total_bal_cxp)
    dif_1009 = total_f1009 - total_bal_cxp
    if abs(dif_1009) > 1000 and total_bal_cxp > 0:
        _rv_nota(f"↳ Diferencia F1009: ${dif_1009:,.0f}. Total pasivo balance = ${total_bal_cxp:,.0f}, "
                 f"Exógena = ${total_f1009:,.0f}. Puede incluir cuentas sin tercero o saldos menores agrupados.")

    _rv_linea("F1010", "Socios y accionistas", len(dic10),
              total_f1010, "Cta 3xxx (capital)", None)
    _rv_linea("F1012", "Inversiones y ctas bancarias", len(dic12),
              total_f1012, "Ctas 11xx + 12xx", total_bal_inv)
    dif_1012 = total_f1012 - total_bal_inv
    if abs(dif_1012) > 1000 and total_bal_inv > 0:
        _rv_nota(f"↳ Diferencia F1012: ${dif_1012:,.0f}. Balance bancos+inversiones = ${total_bal_inv:,.0f}, "
                 f"Exógena = ${total_f1012:,.0f}. Verifique que todos los bancos tengan NIT asignado.")

    # ========== SECCIÓN 3: FORMATOS NO INCLUIDOS ==========
    _rv_titulo("📋 FORMATOS QUE REQUIEREN INFORMACIÓN ADICIONAL")
    _rv_nota("F1004 (Descuentos tributarios): Requiere datos manuales del contador.")
    _rv_nota("F1010 (Socios): Se recomienda solicitar listado detallado de socios como información adicional.")
    _rv_nota("F1011, F1647: Requieren datos manuales del contador.")

    # Nota final
    rv_row += 1
    ws_rv.cell(rv_row, 1).value = (
        "💡 NOTA INFORMATIVA: Los valores del balance corresponden a saldos acumulados con tercero. "
        "Pueden existir diferencias por: cuantías menores agrupadas, ajustes, notas débito/crédito, "
        "o cuentas que no tienen tercero asignado en el balance de prueba."
    )
    ws_rv.cell(rv_row, 1).font = Font(size=9, name='Calibri', color='999999', italic=True)
    ws_rv.cell(rv_row, 1).alignment = Alignment(wrap_text=True)
    ws_rv.merge_cells(start_row=rv_row, start_column=1, end_row=rv_row, end_column=NC_RV)
    ws_rv.row_dimensions[rv_row].height = 40

    rv_anchos = [10, 32, 6, 18, 18, 18, 16, 16]
    for i, a in enumerate(rv_anchos, 1):
        ws_rv.column_dimensions[openpyxl.utils.get_column_letter(i)].width = a
    ws_rv.freeze_panes = 'A2'

    # === MOVER "Resumen Valores" a la posición 2 (después de "Resumen") ===
    sheet_names = wb.sheetnames
    idx_rv = sheet_names.index("Resumen Valores")
    wb.move_sheet("Resumen Valores", offset=(1 - idx_rv))

    # === CRUCES EXÓGENA vs BALANCE (para dashboard) ===
    cruces = {
        'F1007 Ingresos':        (total_f1007, total_ingresos_4),
        'F1001 Pagos':           (total_f1001_pagos, total_gastos_5 + total_costos_6),
        '  → Ret Fte (F1001)':   (total_f1001_retfte, total_ret_fte_2365),
        '  → Ret IVA (F1001)':   (total_f1001_retiva, total_ret_iva_2367),
        'F1003 Retenciones':     (total_f1003, total_ret_1355),
        'F1005 IVA Descontable': (total_f1005, total_iva_desc),
        'F1006 IVA Generado':    (total_f1006, total_iva_gen),
        'F2276 Nómina':          (total_f2276, total_nomina),
        'F1008 CxC':             (total_f1008, total_bal_cxc),
        'F1009 CxP':             (total_f1009, total_bal_cxp),
        'F1012 Inversiones':     (total_f1012, total_bal_inv),
    }

    return wb, resultados, len(bal), len(direc), n_con_dir, nits_nuevos, cruces


# ======================================================================
# === INTERFAZ STREAMLIT ===
# ======================================================================

st.markdown('<p class="main-header">📊 Exógena DIAN 2025</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Generador automático de formatos — Año Gravable 2025</p>', unsafe_allow_html=True)

st.markdown("""
<div class="privacy-box">
    <h4><span class="privacy-icon">🛡️</span> Confidencialidad y Protección de Datos</h4>
    <p>🗑️ Sus archivos se procesan en memoria y se eliminan automáticamente al finalizar.</p>
    <p>🚫 No almacenamos, compartimos ni transferimos su información a terceros.</p>
    <p>📋 Cumplimiento: Ley 1581/2012, Ley 43/1990 Art. 63, ET Art. 583.</p>
</div>
""", unsafe_allow_html=True)

# === CARGAR DIRECTORIO CENTRALIZADO ===
dir_central, err_central = cargar_directorio_central()
if err_central:
    st.sidebar.warning(f"⚠️ Directorio centralizado no disponible: {err_central[:80]}")
else:
    st.sidebar.success(f"📚 Directorio centralizado: {len(dir_central)} terceros")

# === CARGA DE ARCHIVOS ===
st.markdown("""
<div style="background: linear-gradient(135deg, #EFF6FF 0%, #ECFDF5 100%); border-radius: 16px; padding: 32px; margin: 20px 0; border: 2px dashed #2E75B6; text-align: center;">
    <p style="font-size: 2.5rem; margin-bottom: 8px;">📊</p>
    <h3 style="color: #1B3A5C; margin-bottom: 8px; font-size: 1.3rem;">Cargue su Balance de Prueba por Tercero</h3>
    <p style="color: #6B7280; font-size: 0.9rem; margin-bottom: 4px;">Archivo Excel (.xlsx o .xls) con columnas: Cuenta, Nombre, NIT, Razón Social, Débitos, Créditos, Saldo</p>
    <p style="color: #9CA3AF; font-size: 0.78rem;">El archivo se procesa 100% en la nube y se elimina al finalizar</p>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Seleccione o arrastre su archivo aquí",
    type=["xlsx", "xls"],
    help="Balance de prueba por tercero con columnas: Cuenta, Nombre, NIT, Razón Social, Débitos, Créditos, Saldo Final",
    label_visibility="collapsed"
)

if not uploaded_file:
    st.info("👆 **Haga clic en \"Browse files\" o arrastre su archivo Excel** para comenzar a generar los formatos de exógena.")

with st.expander("📋 Opciones adicionales", expanded=False):
    col_a, col_b = st.columns(2)
    with col_a:
        uploaded_dir = st.file_uploader(
            "Directorio de terceros (opcional)",
            type=["xlsx", "xls", "csv"],
            help="Columnas: NIT, Dirección, Cod Depto, Cod Municipio, Cod País"
        )
    with col_b:
        cierra_impuestos = st.toggle(
            "¿El cliente ya cerró impuestos?",
            value=True,
            help="Si las retenciones (1355) ya se cruzaron contra la DIAN y su saldo es $0, la app usa los débitos para F1003."
        )



# === PROCESAMIENTO ===
if uploaded_file:
    st.markdown("---")
    st.markdown("### ⚙️ Paso 2: Procesamiento")

    try:
        df_balance = pd.read_excel(uploaded_file, dtype={0: str, 2: str})
    except Exception as e:
        st.error(f"❌ Error al leer el archivo: {e}")
        st.stop()

    col_map = detectar_columnas(df_balance)
    valido, faltantes = validar_columnas(col_map)

    if not valido:
        st.error(f"❌ No se detectaron las columnas requeridas: {', '.join(faltantes)}")
        st.info("💡 Verifique que su archivo tenga columnas de: Cuenta, NIT/Tercero, Débitos, Créditos")
        st.dataframe(df_balance.head(5))
        st.stop()

    col_nombres = {campo: str(df_balance.columns[idx]) for campo, idx in col_map.items()}
    with st.expander("🔍 Columnas detectadas", expanded=False):
        for campo, nombre in col_nombres.items():
            st.write(f"**{campo}** → `{nombre}`")

    df_directorio = None
    if uploaded_dir:
        try:
            if uploaded_dir.name.endswith('.csv'):
                df_directorio = pd.read_csv(uploaded_dir, dtype=str)
            else:
                df_directorio = pd.read_excel(uploaded_dir, dtype=str)
            st.success(f"✅ Directorio cargado: {len(df_directorio)} registros")
        except Exception as e:
            st.warning(f"⚠️ Error al leer directorio: {e}")

    # Procesar
    with st.spinner("⏳ Generando formatos de exógena..."):
        try:
            wb, resultados, n_filas, n_terceros, n_con_dir, nits_nuevos, cruces = \
                procesar_balance(
                    df_balance,
                    df_directorio=df_directorio,
                    col_map=col_map,
                    cierra_impuestos=cierra_impuestos,
                    dir_central=dir_central,
                    es_pro=es_pro,
                )
        except Exception as e:
            st.error(f"❌ Error al procesar: {e}")
            import traceback
            st.code(traceback.format_exc())
            st.stop()

    # === RESULTADOS ===
    st.markdown("---")
    st.markdown("### ✅ Paso 3: Resultados")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📄 Filas procesadas", f"{n_filas:,}")
    col2.metric("👥 Terceros", f"{n_terceros:,}")
    col3.metric("📍 Con dirección", f"{n_con_dir:,}")
    col4.metric("📊 Total registros", f"{sum(resultados.values()):,}")

    st.markdown("#### 📋 Registros por formato")
    cols = st.columns(4)
    for i, (nombre, n) in enumerate(resultados.items()):
        with cols[i % 4]:
            st.metric(nombre, n)

    # === CRUCE EXÓGENA vs BALANCE ===
    st.markdown("#### 🔍 Verificación: Exógena vs Balance")
    st.caption("Compara los totales generados en cada formato contra las cuentas del balance de prueba.")

    etiquetas_balance = {
        'F1007 Ingresos':        'Cta 4xxx',
        'F1001 Pagos':           'Cta 5+6',
        '  → Ret Fte (F1001)':   'Cta 2365',
        '  → Ret IVA (F1001)':   'Cta 2367',
        'F1003 Retenciones':     'Cta 1355 (déb)',
        'F1005 IVA Descontable': 'Cta 2408',
        'F1006 IVA Generado':    'Cta 2408',
        'F1008 CxC':             'Ctas 13xx',
        'F1009 CxP':             'Total Pasivo (cta 2)',
        'F1012 Inversiones':     'Ctas 11+12xx',
        'F2276 Nómina':          'Cta 5105',
    }

    cruce_data = []
    for nombre, (val_exo, val_bal) in cruces.items():
        fila_cruce = {
            'Formato': nombre,
            'Total Exógena': f"${val_exo:,.0f}" if val_exo else "$0",
        }
        if val_bal is not None:
            fila_cruce['Ref. Balance'] = f"${val_bal:,.0f}"
            dif = val_exo - val_bal
            fila_cruce['Diferencia'] = f"${dif:,.0f}" if dif != 0 else "—"
            if val_bal == 0 and val_exo == 0:
                fila_cruce['Estado'] = "—"
            elif val_bal != 0 and abs(dif / val_bal) <= 0.01:
                fila_cruce['Estado'] = "✅"
            elif val_bal != 0 and abs(dif / val_bal) <= 0.05:
                fila_cruce['Estado'] = "⚠️"
            else:
                fila_cruce['Estado'] = "❌" if dif != 0 else "✅"
        else:
            fila_cruce['Ref. Balance'] = etiquetas_balance.get(nombre, '—')
            fila_cruce['Diferencia'] = "—"
            fila_cruce['Estado'] = "ℹ️"
        cruce_data.append(fila_cruce)

    df_cruce = pd.DataFrame(cruce_data)
    st.dataframe(df_cruce, use_container_width=True, hide_index=True)

    # Alertas de diferencias importantes
    alertas_cruce = []
    for nombre, (val_exo, val_bal) in cruces.items():
        if val_bal is not None and val_bal != 0:
            pct = abs((val_exo - val_bal) / val_bal) * 100
            dif = val_exo - val_bal
            if pct > 5 and abs(dif) > 10000:
                explicacion = ""
                if 'F1001 Pagos' in nombre:
                    explicacion = " → Normal: nómina va al F2276, no al F1001."
                elif 'Ret Fte' in nombre and dif < 0:
                    explicacion = " → Incluye ret. salarios (F2276) y posibles autorretenciones."
                elif 'F1007' in nombre:
                    explicacion = " → Puede incluir cuantías menores agrupadas o ingresos sin tercero."
                alertas_cruce.append(f"**{nombre}**: Exógena ${val_exo:,.0f} vs Balance ${val_bal:,.0f} (dif {pct:.1f}%){explicacion}")
    if alertas_cruce:
        with st.expander(f"📊 {len(alertas_cruce)} diferencia(s) significativas — ver explicación", expanded=True):
            for a in alertas_cruce:
                st.warning(a)

    st.info("💡 **Conceptos 5011, 5012, 5013 (EPS, AFP, ARL, Cajas):** Solo se reportan a entidades. "
            "Estos valores deben coincidir con la planilla de pagos a la seguridad social (PILA). "
            "Ver hoja 'Resumen Valores' para más detalle.")

    if nits_nuevos:
        with st.expander(f"🆕 {len(nits_nuevos)} direcciones nuevas encontradas", expanded=False):
            st.write("Estas direcciones se encontraron en Internet y pueden agregarse al directorio centralizado:")
            data_nuevos = []
            for nit, info in sorted(nits_nuevos.items()):
                data_nuevos.append({
                    'NIT': nit,
                    'Razón Social': info.get('razon', ''),
                    'Dirección': info.get('dir', ''),
                    'Depto': info.get('dp', ''),
                    'Municipio': info.get('mp', ''),
                })
            st.dataframe(pd.DataFrame(data_nuevos), use_container_width=True)

    # === DESCARGAR ===
    st.markdown("---")
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"Exogena_AG2025_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    st.download_button(
        label="📥 Descargar Exógena Completa",
        data=buffer,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    st.markdown("""
    <div style="background: #fff3cd; border-radius: 10px; padding: 1rem; margin-top: 1rem; border-left: 4px solid #ffc107;">
        <strong>⚠️ Recuerde:</strong> Esta herramienta genera un <strong>borrador</strong>.
        El Contador Público debe verificar y validar toda la información antes de presentarla a la DIAN.
        <br><em>Art. 631 ET — Resolución DIAN 000227 de 2025 — Sanción Art. 651 ET.</em>
    </div>
    """, unsafe_allow_html=True)

else:
    st.info("👆 Cargue su archivo de Balance de Prueba por Tercero para comenzar.")
    st.markdown("""
    **Formatos que genera esta herramienta:**

    | Formato | Descripción | Concepto |
    |---------|-------------|----------|
    | F1001 | Pagos y abonos en cuenta | Gastos y costos por tercero |
    | F1003 | Retenciones que le practicaron | Retenciones en la fuente recibidas |
    | F1005 | IVA Descontable | IVA pagado en compras |
    | F1006 | IVA Generado | IVA cobrado en ventas |
    | F1007 | Ingresos recibidos | Ingresos por tercero |
    | F1008 | Cuentas por cobrar | Saldos deudores a Dic 31 |
    | F1009 | Cuentas por pagar | Saldos acreedores a Dic 31 |
    | F1010 | Socios y accionistas | Composición del capital |
    | F1012 | Inversiones y cuentas bancarias | Saldos en bancos e inversiones |
    | F2276 | Rentas de trabajo | Pagos laborales por empleado |

    **Formatos manuales:** F1004, F1011, F1647 requieren datos adicionales no incluidos en el balance.
    """)

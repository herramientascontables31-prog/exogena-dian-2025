import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime
from io import BytesIO
import difflib

# Google Sheets API (escritura automática)
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_OK = True
except ImportError:
    GSPREAD_OK = False

st.set_page_config(page_title="Exógena DIAN 2025", page_icon="📊", layout="wide")

# === PROTECCIÓN CON CONTRASEÑA (Google Sheets) ===
# Instrucciones:
# 1. Crea un Google Sheet con columnas: clave | nombre | estado
# 2. Publícalo como CSV: Archivo → Compartir → Publicar en la web → CSV
# 3. Pega la URL aquí abajo:
GOOGLE_SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/e/TU_ID_AQUI/pub?output=csv"

# === DIRECTORIO CENTRALIZADO DE TERCEROS ===
# Opción 1 (solo lectura): URL CSV publicada
DIRECTORIO_CENTRAL_URL = "https://docs.google.com/spreadsheets/d/e/TU_ID_DIRECTORIO/pub?output=csv"

# Opción 2 (lectura + escritura automática): ID del Google Sheet + cuenta de servicio
# El ID es la parte entre /d/ y /edit en la URL del Sheet
# Ejemplo: https://docs.google.com/spreadsheets/d/ABC123xyz/edit → ID = "ABC123xyz"
DIRECTORIO_SHEET_ID = ""  # ← Pega aquí el ID de tu Google Sheet de directorio
DIRECTORIO_HOJA_NOMBRE = "Directorio"  # Nombre de la hoja dentro del Sheet

# === CONFIGURACIÓN CUENTA DE SERVICIO GOOGLE ===
# Para escritura automática en Google Sheets:
# 1. Ve a console.cloud.google.com → Crear proyecto
# 2. Habilita "Google Sheets API"
# 3. Crea una Cuenta de Servicio → descarga el JSON
# 4. Comparte tu Google Sheet con el email de la cuenta de servicio (permisos de Editor)
# 5. En Streamlit Cloud → Settings → Secrets → pega el JSON así:
#
#    [gcp_service_account]
#    type = "service_account"
#    project_id = "tu-proyecto"
#    private_key_id = "..."
#    private_key = "-----BEGIN PRIVATE KEY-----\n...\n-----END PRIVATE KEY-----\n"
#    client_email = "exogena@tu-proyecto.iam.gserviceaccount.com"
#    client_id = "..."
#    auth_uri = "https://accounts.google.com/o/oauth2/auth"
#    token_uri = "https://oauth2.googleapis.com/token"
#    auth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"
#    client_x509_cert_url = "..."

def conectar_gsheets():
    """Conecta a Google Sheets via cuenta de servicio. Retorna (client, error)."""
    if not GSPREAD_OK:
        return None, "gspread no instalado. Ejecuta: pip install gspread google-auth"
    if not DIRECTORIO_SHEET_ID:
        return None, "DIRECTORIO_SHEET_ID no configurado"
    try:
        creds_dict = dict(st.secrets.get("gcp_service_account", {}))
        if not creds_dict:
            return None, "Secreto gcp_service_account no configurado en Streamlit"
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        return client, None
    except Exception as e:
        return None, str(e)

def agregar_direcciones_a_sheet(nits_nuevos):
    """Agrega direcciones nuevas al Google Sheet centralizado.
    Retorna (n_agregados, error)."""
    client, err = conectar_gsheets()
    if err:
        return 0, err
    try:
        sh = client.open_by_key(DIRECTORIO_SHEET_ID)
        ws = sh.worksheet(DIRECTORIO_HOJA_NOMBRE)
        
        # Leer NITs existentes para no duplicar
        nits_existentes = set()
        col_nits = ws.col_values(1)  # Columna A = NIT
        for n in col_nits[1:]:  # Saltar header
            nit_limpio = str(n).replace('.', '').replace('-', '').strip()
            if nit_limpio:
                nits_existentes.add(nit_limpio)
        
        # Preparar filas nuevas
        filas_nuevas = []
        for nit, info in sorted(nits_nuevos.items()):
            if nit in nits_existentes:
                continue
            filas_nuevas.append([
                nit,
                info.get('razon', ''),
                info.get('dir', ''),
                info.get('dp', ''),
                info.get('mp', ''),
                info.get('pais', '169'),
                info.get('td', ''),
                info.get('dv', ''),
            ])
        
        if filas_nuevas:
            ws.append_rows(filas_nuevas, value_input_option='USER_ENTERED')
        
        return len(filas_nuevas), None
    except Exception as e:
        return 0, str(e)

def sheets_escritura_disponible():
    """Verifica si la escritura automática en Google Sheets está configurada."""
    if not GSPREAD_OK:
        return False
    if not DIRECTORIO_SHEET_ID:
        return False
    try:
        creds_dict = dict(st.secrets.get("gcp_service_account", {}))
        return bool(creds_dict)
    except:
        return False

# Contraseña de respaldo (funciona siempre, por si Google Sheets falla)
CLAVE_ADMIN = "ExoDIAN-2025-ADMIN"

@st.cache_data(ttl=300)  # Cache 5 minutos para no consultar Google en cada click
def cargar_clientes():
    """Carga la lista de clientes desde Google Sheets publicado como CSV."""
    try:
        df = pd.read_csv(GOOGLE_SHEET_CSV_URL, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        clientes = {}
        for _, row in df.iterrows():
            clave = str(row.get('clave', '')).strip()
            nombre = str(row.get('nombre', '')).strip()
            estado = str(row.get('estado', '')).strip().lower()
            if clave and clave.lower() != 'nan':
                clientes[clave] = {
                    'nombre': nombre if nombre.lower() != 'nan' else '',
                    'activo': estado in ('activo', 'si', 'sí', '1', 'true', 'yes'),
                }
        return clientes, None
    except Exception as e:
        return {}, str(e)

def verificar_clave(clave_ingresada):
    """Verifica la clave contra Google Sheets. Retorna (valida, nombre, mensaje_error)."""
    # Clave de administrador (siempre funciona)
    if clave_ingresada == CLAVE_ADMIN:
        return True, "🔑 Administrador", None
    
    clientes, error = cargar_clientes()
    
    if error:
        # Si Google Sheets falla, solo funciona la clave admin
        return False, "", f"⚠️ No se pudo verificar. Contacta al administrador. ({error[:60]})"
    
    if clave_ingresada in clientes:
        cliente = clientes[clave_ingresada]
        if cliente['activo']:
            return True, cliente['nombre'], None
        else:
            return False, "", "🚫 Tu acceso está **desactivado**. Contacta al administrador."
    
    return False, "", "❌ Contraseña incorrecta. Verifica tu compra en exogenadian.com"

if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.nombre_cliente = ""

if not st.session_state.autenticado:
    st.markdown("""
    <div style="max-width: 450px; margin: 80px auto; text-align: center;">
        <h1 style="font-size: 2.5rem; margin-bottom: 0.2rem;">📊</h1>
        <h2 style="font-family: serif; font-size: 1.8rem; color: #0B1D3A; margin-bottom: 0.5rem;">Exógena DIAN 2025</h2>
        <p style="color: #64748B; font-size: 0.95rem; margin-bottom: 2rem;">Ingresa tu contraseña de acceso para continuar</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        clave = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña aquí")
        if st.button("Ingresar", use_container_width=True, type="primary"):
            valida, nombre, msg_error = verificar_clave(clave)
            if valida:
                st.session_state.autenticado = True
                st.session_state.nombre_cliente = nombre
                st.rerun()
            else:
                st.error(msg_error)
        
        st.markdown("""
        <div style="text-align: center; margin-top: 2rem;">
            <p style="color: #94a3b8; font-size: 0.82rem;">
                ¿No tienes contraseña? <a href="https://exogenadian.com/#precios" target="_blank" style="color: #1F4E79;">Compra tu acceso aquí</a>
            </p>
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# === Mostrar nombre del cliente conectado ===
if st.session_state.get('nombre_cliente'):
    st.sidebar.markdown(f"### 👤 {st.session_state.nombre_cliente}")
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Cerrar sesión"):
        st.session_state.autenticado = False
        st.session_state.nombre_cliente = ""
        st.rerun()
    
    # Indicador de sincronización
    st.sidebar.markdown("---")
    if sheets_escritura_disponible():
        st.sidebar.success("🔄 Sincronización automática **activa**")
        st.sidebar.caption("Las direcciones nuevas se agregan al directorio centralizado automáticamente.")
    else:
        st.sidebar.warning("📋 Sincronización **manual**")
        st.sidebar.caption("Las direcciones nuevas se descargan como CSV para pegar manualmente.")

# === ESTILOS ===
st.markdown("""
<style>
    .main-header { font-size: 2.2rem; font-weight: 700; color: #1F4E79; margin-bottom: 0; }
    .sub-header { font-size: 1.1rem; color: #666; margin-top: 0; }
    .metric-card { background: #f8f9fa; border-radius: 10px; padding: 1rem; border-left: 4px solid #1F4E79; }
    .success-box { background: #d4edda; border-radius: 10px; padding: 1rem; border-left: 4px solid #28a745; }
    .privacy-box { background: linear-gradient(135deg, #e8f4fd 0%, #f0f7ee 100%); border-radius: 12px; padding: 1.2rem 1.5rem; border: 1px solid #b8d4e8; margin: 0.8rem 0 1.2rem 0; }
    .privacy-box h4 { color: #1a5276; margin: 0 0 0.5rem 0; font-size: 1rem; }
    .privacy-box p { color: #2c3e50; font-size: 0.88rem; line-height: 1.5; margin: 0.3rem 0; }
    .privacy-box .privacy-icon { font-size: 1.3rem; margin-right: 0.3rem; }
    .stDownloadButton > button { background-color: #1F4E79 !important; color: white !important; font-size: 1.1rem !important; padding: 0.5rem 2rem !important; }
</style>
""", unsafe_allow_html=True)

# === DIRECTORIO CENTRALIZADO ===
@st.cache_data(ttl=600)  # Cache 10 min
def cargar_directorio_central():
    """Carga el directorio centralizado de terceros desde Google Sheets."""
    try:
        df = pd.read_csv(DIRECTORIO_CENTRAL_URL, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        directorio = {}
        for _, row in df.iterrows():
            nit = str(row.get('nit', '')).strip()
            if not nit or nit.lower() == 'nan':
                continue
            # Limpiar NIT
            nit = nit.replace('.', '').replace('-', '').strip()
            if '.' in nit:
                try: nit = str(int(float(nit)))
                except: pass
            directorio[nit] = {
                'razon': str(row.get('razón social', row.get('razon social', ''))).strip(),
                'dir': str(row.get('dirección', row.get('direccion', ''))).strip(),
                'depto': str(row.get('cod depto', row.get('depto', ''))).strip(),
                'mpio': str(row.get('cod municipio', row.get('municipio', ''))).strip(),
                'pais': str(row.get('cod país', row.get('pais', '169'))).strip(),
                'td': str(row.get('tipo doc', '')).strip(),
                'dv': str(row.get('dv', '')).strip(),
            }
            # Limpiar nans
            for k in directorio[nit]:
                if directorio[nit][k].lower() == 'nan':
                    directorio[nit][k] = ''
        return directorio, None
    except Exception as e:
        return {}, str(e)

# === CONSTANTES ===
UVT = 49799
C3UVT = 149397
C12UVT = 597588
NM = "222222222"
TDM = "43"

# === FUNCIONES CORE ===
def calc_dv(n):
    n = str(n).replace(".", "").replace("-", "").strip()
    if not n or not n.isdigit() or n == NM:
        return ""
    pesos = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    np = n.zfill(15)
    s = sum(int(np[i]) * pesos[i] for i in range(15))
    r = s % 11
    return str(11 - r) if r >= 2 else str(r)

def detectar_tipo_doc(nit):
    """Detecta automáticamente el tipo de documento DIAN según el número de identificación.
    
    Tipos soportados:
    - 31: NIT (empresas colombianas, 9 dígitos que inician con 8 o 9)
    - 13: Cédula de ciudadanía (personas naturales colombianas)
    - 12: Tarjeta de identidad (menores de edad) — no distinguible por número
    - 22: Cédula de extranjería — no distinguible por número  
    - 41: Pasaporte (contiene letras)
    - 42: Documento de identificación extranjero (contiene letras)
    - 43: Cuantías menores / sin identificación
    
    Limitación: los tipos 12, 13 y 22 no se pueden distinguir solo por el número.
    Se asume 13 (CC) como default para personas. El contador debe ajustar manualmente
    los casos de TI (12) y CE (22) en el archivo de salida.
    """
    if not nit or nit == NM:
        return TDM  # "43" para cuantías menores
    
    nit = str(nit).strip()
    
    # Limpiar formato decimal (ej: "890904997.0" → "890904997")
    if '.' in nit:
        try:
            nit = str(int(float(nit)))
        except:
            pass
    
    # Si contiene letras → documento extranjero o pasaporte
    if not nit.isdigit():
        # Pasaportes suelen tener formato: 2 letras + números (ej: AB123456)
        # o patrones alfanuméricos cortos
        letras = sum(1 for c in nit if c.isalpha())
        if letras <= 3:
            return "41"  # Pasaporte (pocas letras + números)
        else:
            return "42"  # Documento de identificación extranjero
    
    # NIT de empresa: 9 dígitos, empieza con 8 o 9
    # Ej: 800123456, 890904997, 900555111
    if len(nit) == 9 and nit[0] in ('8', '9'):
        return "31"  # NIT
    
    # NIT de entidad pública o especial: 9+ dígitos con 8 o 9
    if len(nit) >= 9 and nit[0] in ('8', '9'):
        return "31"  # NIT
    
    # Todo lo demás se asume cédula de ciudadanía
    # Nota: TI (12) y CE (22) no se pueden distinguir automáticamente
    return "13"  # Cédula de ciudadanía

def safe_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        import math
        if math.isnan(v):
            return 0.0
        return float(v)
    try:
        s = str(v).strip()
        if s.lower() == 'nan' or s == '':
            return 0.0
        return float(s.replace(",", "").replace("$", "").replace(" ", ""))
    except:
        return 0.0

def safe_str(v):
    if v is None:
        return ""
    # Celdas vacías en Excel a veces se leen como 0 o 0.0
    if isinstance(v, (int, float)):
        import math
        if v == 0 or (isinstance(v, float) and math.isnan(v)):
            return ""
    s = str(v).strip()
    if s.lower() == 'nan' or s == '0.0':
        return ""
    return s

def detectar_columnas(df):
    """Detecta automáticamente qué columna corresponde a cada campo del balance.
    Busca por nombre de columna (flexible, sin importar mayúsculas/tildes/orden).
    Retorna un dict con los índices de columna para cada campo."""
    
    import unicodedata
    
    def normalizar(texto):
        """Quita tildes, pasa a minúsculas, quita espacios extra"""
        if not texto:
            return ""
        texto = str(texto).lower().strip()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
        for ch in '.,;:-_/\\()[]{}#"\'':
            texto = texto.replace(ch, ' ')
        return ' '.join(texto.split())
    
    columnas = {}
    nombres = [normalizar(str(c)) for c in df.columns]
    
    # === Mapeo de nombres posibles para cada campo ===
    mapeo = {
        'cuenta': ['cuenta', 'codigo cuenta', 'cod cuenta', 'cuenta contable', 'codigo', 'account'],
        'nombre': ['descripcion cuenta', 'descripcion', 'nombre cuenta', 'nombre', 'detalle', 
                    'concepto', 'account name', 'desc cuenta'],
        'nit': ['tercero', 'nit', 'identificacion', 'documento', 'id tercero', 'nro documento',
                'num documento', 'cedula', 'numero identificacion'],
        'razon_social': ['razon social', 'nombre tercero', 'tercero nombre', 'razon', 'beneficiario',
                         'proveedor', 'cliente', 'nombre razon social'],
        'debito': ['debitos', 'debito', 'debe', 'movimiento debito', 'mov debito', 'cargos',
                   'debits', 'debit'],
        'credito': ['creditos', 'credito', 'haber', 'movimiento credito', 'mov credito', 'abonos',
                    'credits', 'credit'],
        'saldo': ['saldo final', 'saldo', 'saldo actual', 'balance', 'saldo cierre',
                  'saldo a diciembre', 'saldo dic'],
    }
    
    for campo, keywords in mapeo.items():
        for i, nom in enumerate(nombres):
            if i in columnas.values():
                continue  # Ya asignada
            for kw in keywords:
                if kw == nom or kw in nom:
                    columnas[campo] = i
                    break
            if campo in columnas:
                break
    
    # Si no encontró "saldo" pero hay "saldo final", ya lo detectó arriba
    # Si hay "saldo inicial" asegurarse de NO confundirlo con "saldo"
    for i, nom in enumerate(nombres):
        if ('saldo inicial' in nom or 'saldo anterior' in nom) and columnas.get('saldo') == i:
            # Ups, detectamos saldo inicial como saldo — buscar otra columna
            del columnas['saldo']
            for j, nom2 in enumerate(nombres):
                if j not in columnas.values() and j != i:
                    if 'saldo final' in nom2 or (nom2 == 'saldo' and 'inicial' not in nom2):
                        columnas['saldo'] = j
                        break
    
    return columnas

def validar_columnas(columnas_detectadas):
    """Valida que se encontraron las columnas mínimas necesarias.
    Retorna (ok, campos_faltantes)."""
    requeridas = ['cuenta', 'nit', 'debito', 'credito']
    faltantes = [c for c in requeridas if c not in columnas_detectadas]
    return len(faltantes) == 0, faltantes

def en_rango(cta, d, h):
    n = len(d)
    return cta[:n] >= d and cta[:n] <= h

def pad_dpto(v):
    """Formatea código departamento DIAN: 2 dígitos con cero a la izquierda"""
    if not v:
        return ""
    v = str(v).strip()
    if '.' in v:
        try: v = str(int(float(v)))
        except: pass
    if v.lower() == 'nan': return ""
    return v.zfill(2) if v.isdigit() else v

def pad_mpio(v):
    """Formatea código municipio DIAN: 3 dígitos con ceros a la izquierda"""
    if not v:
        return ""
    v = str(v).strip()
    if '.' in v:
        try: v = str(int(float(v)))
        except: pass
    if v.lower() == 'nan': return ""
    return v.zfill(3) if v.isdigit() else v

def buscar_info_terceros(nits_list, progress_bar=None, log_fn=None):
    """Busca info de terceros usando múltiples fuentes de internet."""
    import requests
    from time import sleep
    import re

    def log(msg):
        if log_fn:
            log_fn(msg)

    encontrados = {}
    total = len(nits_list)
    errores_seguidos = 0

    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': '*/*',
        'Accept-Language': 'es-CO,es;q=0.9',
    }

    def buscar_rues(nit):
        try:
            resp = requests.get(
                'https://www.rues.org.co/RM/ConsultaNit_Api',
                params={'nit': str(nit), 'tipo': 'N'},
                headers={**HEADERS, 'Referer': 'https://www.rues.org.co/',
                         'X-Requested-With': 'XMLHttpRequest',
                         'Accept': 'application/json'},
                timeout=12
            )
            if resp.status_code == 200:
                data = resp.json()
                registros = data if isinstance(data, list) else \
                            data.get('registros', data.get('data', [])) if isinstance(data, dict) else []
                if registros and len(registros) > 0:
                    return extraer_info_dict(registros[0]), None
                return None, "Sin resultados"
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def buscar_datos_gov_lote(nits_batch):
        resultados = {}
        datasets = ["c82q-fe7j", "8yz5-t3jw"]
        for ds_id in datasets:
            try:
                nits_str = "','".join(str(n) for n in nits_batch)
                resp = requests.get(
                    f"https://www.datos.gov.co/resource/{ds_id}.json",
                    params={'$where': f"nit in ('{nits_str}')", '$limit': 5000},
                    headers=HEADERS, timeout=20
                )
                if resp.status_code == 200:
                    data = resp.json()
                    if data and isinstance(data, list):
                        for emp in data:
                            nit_val = str(emp.get('nit', emp.get('NIT', ''))).strip()
                            if nit_val:
                                info = extraer_info_dict(emp)
                                if info:
                                    resultados[nit_val] = info
                                    resultados[nit_val]['_fuente'] = 'Datos.gov.co'
                        if resultados:
                            return resultados, None
                    return {}, f"Dataset {ds_id}: sin datos"
                else:
                    continue
            except Exception as e:
                continue
        return {}, "Ningún dataset respondió"

    def buscar_einforma(nit):
        try:
            resp = requests.get(
                f'https://www.einforma.co/servlet/app/portal/ENTP/prod/ETIQUETA_EMPRESA_498/nif/{nit}',
                headers=HEADERS, timeout=12, allow_redirects=True
            )
            if resp.status_code == 200:
                info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
                texto = resp.text
                rs_match = re.findall(r'<h1[^>]*class="[^"]*nombre[^"]*"[^>]*>([^<]+)</h1>', texto, re.IGNORECASE)
                if not rs_match:
                    rs_match = re.findall(r'<title>([^<]+?)[\s\-|]', texto)
                if rs_match:
                    rs = rs_match[0].strip()
                    if len(rs) > 3 and str(nit) not in rs.lower():
                        info['razon_social'] = rs.upper()
                dir_match = re.findall(r'(?:Direcci[oó]n|Domicilio)[:\s]*</[^>]+>\s*<[^>]+>([^<]+)', texto, re.IGNORECASE)
                if dir_match:
                    info['dir'] = dir_match[0].strip()
                if info.get('razon_social') or info.get('dir'):
                    return info, None
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def buscar_web_ddg(nit):
        try:
            resp = requests.get('https://html.duckduckgo.com/html/',
                params={'q': f'NIT {nit} Colombia empresa direccion'}, headers=HEADERS, timeout=12)
            if resp.status_code == 200:
                return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def buscar_web_bing(nit):
        try:
            resp = requests.get(f'https://www.bing.com/search?q=NIT+{nit}+Colombia+empresa+direccion&setlang=es',
                headers=HEADERS, timeout=12)
            if resp.status_code == 200:
                return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def buscar_web_google(nit):
        try:
            query = f"NIT+{nit}+Colombia+empresa+direccion"
            resp = requests.get(f'https://www.google.com/search?q={query}&hl=es&gl=co',
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html', 'Accept-Language': 'es-CO,es;q=0.9'}, timeout=12)
            if resp.status_code == 200:
                return extraer_info_web(nit, resp.text), None
            return None, f"HTTP {resp.status_code}"
        except Exception as e:
            return None, str(e)[:80]

    def extraer_info_web(nit, html):
        info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
        nit_str = str(nit)
        texto = re.sub(r'<[^>]+>', ' ', html)
        texto = re.sub(r'\s+', ' ', texto)
        patrones_rs = [
            r'(?:NIT|Nit|nit)[\s.:]*' + re.escape(nit_str) + r'[\s\-–—:,.]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]+)',
            r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]{5,50}?)[\s\-–—:,.]+(?:NIT|Nit|nit)[\s.:]*' + re.escape(nit_str),
            r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s&.,]{5,50}?)\s*[-–—]\s*NIT[\s.:]*' + re.escape(nit_str),
        ]
        for patron in patrones_rs:
            matches = re.findall(patron, texto)
            if matches:
                rs = matches[0].strip().rstrip('.,;:-– ')
                if 3 < len(rs) < 120:
                    info['razon_social'] = rs.upper()
                    break
        patrones_dir = [
            r'(?:Direcci[oó]n|Dir\.?|Ubicaci[oó]n)[:\s]+([A-Za-z]{2,3}[\s.]*(?:No\.?\s*)?\d+[\w\s#\-.,No°]+?\d)',
            r'((?:CL|CR|KR|TV|DG|CALLE|CARRERA|AV|AVENIDA|TRANSVERSAL|DIAGONAL)[\s.]*(?:No\.?\s*)?\d+[\w\s#\-.,No°]*\d)',
        ]
        for patron in patrones_dir:
            matches = re.findall(patron, texto, re.IGNORECASE)
            if matches:
                dir_candidata = matches[0].strip()[:100]
                if len(dir_candidata) > 5:
                    info['dir'] = dir_candidata
                    break
        if info.get('razon_social') or info.get('dir'):
            return info
        return None

    def extraer_info_dict(emp):
        if not isinstance(emp, dict):
            return None
        info = {'razon_social': '', 'dv': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
        for campo in ['razon_social', 'Razon_Social', 'nombre', 'Nombre', 'razonSocial', 'RazonSocial',
                       'nombre_razon_social', 'NombreEstablecimiento', 'organizacion', 'nombre_empresa']:
            val = emp.get(campo, '')
            if val:
                info['razon_social'] = str(val).strip().upper()
                break
        for campo in ['digito_verificacion', 'Digito_Verificacion', 'dv', 'DV', 'digitoVerificacion']:
            val = emp.get(campo, '')
            if val is not None and str(val).strip():
                info['dv'] = str(val).strip()
                break
        for campo in ['direccion', 'Direccion', 'direccion_comercial', 'DireccionComercial', 'dir_comercial']:
            val = emp.get(campo, '')
            if val:
                info['dir'] = str(val).strip()
                break
        for campo in ['codigo_departamento', 'departamento', 'cod_departamento', 'CodigoDepartamento', 'dep_codigo', 'cod_depto']:
            val = emp.get(campo, '')
            if val:
                info['dp'] = pad_dpto(str(val).strip())
                break
        for campo in ['codigo_municipio', 'municipio', 'cod_municipio', 'CodigoMunicipio', 'mun_codigo', 'ciudad', 'cod_ciudad']:
            val = emp.get(campo, '')
            if val:
                info['mp'] = pad_mpio(str(val).strip())
                break
        return info if (info.get('razon_social') or info.get('dir')) else None

    # FLUJO PRINCIPAL
    log("📡 **Paso 1:** Consultando datos.gov.co (lote completo)...")
    try:
        lote_result, lote_error = buscar_datos_gov_lote(nits_list)
        if lote_result:
            encontrados.update(lote_result)
            log(f"  ✅ datos.gov.co: {len(lote_result)} terceros encontrados")
        else:
            log(f"  ❌ datos.gov.co: {lote_error}")
    except Exception as e:
        log(f"  ❌ datos.gov.co: Error — {str(e)[:80]}")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    rues_funciona = False
    if nits_faltantes:
        log(f"📡 **Paso 2:** Probando RUES ({len(nits_faltantes)} NITs pendientes)...")
        test_nit = nits_faltantes[0]
        try:
            resultado, error = buscar_rues(test_nit)
            if resultado:
                resultado['_fuente'] = 'RUES'
                encontrados[test_nit] = resultado
                rues_funciona = True
                log(f"  ✅ RUES funciona (NIT {test_nit} encontrado)")
            else:
                log(f"  ❌ RUES: {error}")
        except Exception as e:
            log(f"  ❌ RUES no disponible: {str(e)[:80]}")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    buscador_web = None
    buscadores = [('DuckDuckGo', buscar_web_ddg), ('Bing', buscar_web_bing), ('Google', buscar_web_google)]

    if nits_faltantes:
        log(f"📡 **Paso 3:** Probando buscadores web...")
        test_nit = nits_faltantes[0]
        for nombre_b, fn_b in buscadores:
            try:
                resultado, error = fn_b(test_nit)
                if resultado:
                    resultado['_fuente'] = nombre_b
                    encontrados[test_nit] = resultado
                    buscador_web = (nombre_b, fn_b)
                    log(f"  ✅ {nombre_b} funciona")
                    break
                else:
                    log(f"  ⚠️ {nombre_b}: {error}")
            except Exception as e:
                log(f"  ❌ {nombre_b}: {str(e)[:60]}")
        if not buscador_web:
            log("  ❌ Ningún buscador web funcionó")

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    if nits_faltantes and (rues_funciona or buscador_web):
        fuentes_activas = []
        if rues_funciona: fuentes_activas.append(('RUES', buscar_rues))
        if buscador_web: fuentes_activas.append(buscador_web)
        nombres = " → ".join(f[0] for f in fuentes_activas)
        log(f"🔍 **Paso 4:** Buscando {len(nits_faltantes)} NITs restantes [{nombres}]...")

        for i, nit in enumerate(nits_faltantes):
            if progress_bar:
                progress_bar.progress((i + 1) / len(nits_faltantes),
                    text=f"🔍 {nit} ({i+1}/{len(nits_faltantes)}) — Encontrados: {len(encontrados)}")
            if errores_seguidos >= 15:
                log(f"  ⛔ Detenido tras {errores_seguidos} errores seguidos")
                break
            encontro = False
            for nombre_f, fn_f in fuentes_activas:
                try:
                    resultado, error = fn_f(nit)
                    if resultado:
                        resultado['_fuente'] = nombre_f
                        encontrados[nit] = resultado
                        encontro = True
                        errores_seguidos = 0
                        break
                except Exception:
                    continue
            if not encontro:
                errores_seguidos += 1
            sleep(0.8)

    nits_faltantes = [n for n in nits_list if n not in encontrados]
    if nits_faltantes and len(nits_faltantes) < 30:
        log(f"📡 **Paso 5:** Probando einforma.co ({len(nits_faltantes)} NITs pendientes)...")
        errores_ein = 0
        for nit in nits_faltantes:
            if errores_ein >= 5: break
            try:
                resultado, error = buscar_einforma(nit)
                if resultado:
                    resultado['_fuente'] = 'einforma.co'
                    encontrados[nit] = resultado
                    errores_ein = 0
                else:
                    errores_ein += 1
            except Exception:
                errores_ein += 1
            sleep(1.0)

    n_dir = sum(1 for d in encontrados.values() if d.get('dir'))
    n_rs = sum(1 for d in encontrados.values() if d.get('razon_social'))
    log(f"\n📊 **Resumen:** {len(encontrados)}/{total} terceros — {n_dir} con dirección, {n_rs} con razón social")
    return encontrados, len(encontrados) == 0 and total > 0


def normalizar_texto(t):
    import unicodedata
    if not t: return ""
    t = str(t).upper().strip()
    t = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    for ch in '.,;:-_/\\()[]{}#"\'':
        t = t.replace(ch, ' ')
    t = ' '.join(t.split())
    reemplazos = [
        ('S A S', 'SAS'), ('S A  S', 'SAS'), ('S A', 'SA'), ('S  A', 'SA'),
        ('E S P', 'ESP'), ('E  S  P', 'ESP'), ('E U', 'EU'), ('E  U', 'EU'),
        ('C I', 'CI'), ('C  I', 'CI'), ('N I T', 'NIT'),
        ('SOCIEDAD ANONIMA', 'SA'), ('SOCIEDAD POR ACCIONES SIMPLIFICADA', 'SAS'),
        ('EMPRESA DE SERVICIOS PUBLICOS', 'ESP'), ('LIMITADA', 'LTDA'),
    ]
    for viejo, nuevo in reemplazos:
        t = t.replace(viejo, nuevo)
    return ' '.join(t.split())


def similitud_textos(a, b):
    a = normalizar_texto(a)
    b = normalizar_texto(b)
    if not a or not b: return 0.0
    if a == b: return 1.0
    pa = set(a.split())
    pb = set(b.split())
    if not pa or not pb: return 0.0
    comunes = pa & pb
    return len(comunes) / max(len(pa), len(pb))

# === PARAMETRIZACION ===
PARAM_1001_NOMINA_SUB = {
    "5001": ["06", "07", "08", "09", "10", "15", "27", "30", "33", "36", "39", "42", "45"],
    "5002": ["01", "03", "05"],
    "5024": ["02"],
    "5025": ["03"],
    "5027": ["04"],
    "5023": ["68", "72", "75"],
}

PARAM_1001_RANGOS = [
    ("5005", "5120", "5120", True), ("5005", "5220", "5220", True),
    ("5011", "5230", "5230", True), ("5011", "5130", "5130", True),
    ("5055", "5115", "5115", True), ("5004", "5110", "5110", True),
    ("5016", "5125", "5125", True), ("5016", "5135", "5139", True),
    ("5016", "5140", "5199", True), ("5016", "5210", "5219", True),
    ("5016", "5235", "5249", True), ("5016", "5295", "5299", True),
    ("5055", "5115", "5115", True), ("5006", "5305", "5305", True),
    ("5101", "530515", "530515", True), ("5007", "1435", "1499", True),
    ("5010", "1504", "1699", True), ("5010", "1520", "1540", True),
]

PARAM_1003 = [
    # Retenciones en la fuente que le practicaron (135515xx)
    ("1301", "13551505", "13551509"),  # Salarios y pagos laborales
    ("1301", "13551510", "13551514"),  # Honorarios
    ("1302", "13551515", "13551519"),  # Comisiones
    ("1303", "13551520", "13551524"),  # Servicios
    ("1305", "13551525", "13551529"),  # Rendimientos financieros
    ("1304", "13551530", "13551534"),  # Arrendamientos
    ("1306", "13551535", "13551539"),  # Compras
    ("1308", "13551540", "13551594"),  # Otros conceptos
    ("1308", "13551595", "13551599"),  # Otros
    # Retención ICA que le practicaron (135518xx)
    ("1307", "135518", "135518"),      # ICA retenido
    # Autorretenciones (135599xx con nombre)
    ("1311", "135599", "135599"),      # Otros anticipos / autorretenciones
    # Rangos amplios por si usan estructura simplificada
    ("1303", "135515", "135515"),      # Retención en la fuente (genérico)
]

PARAM_1007 = [
    ("4001", "4101", "4199"), ("4001", "4135", "4135"),
    ("4002", "4201", "4299"), ("4003", "4210", "4210"),
]

PARAM_1008 = [("1315", "1305", "1305"), ("1316", "1380", "1399"), ("1317", "1330", "1365")]
PARAM_1009 = [("2201", "2205", "2209"), ("2202", "2105", "2199"), ("2207", "2380", "2389"), ("2210", "2335", "2369")]

MAPEO_1012 = [
    ("8302", "1105", "1105"), ("8301", "1110", "1110"),
    ("8305", "1201", "1204"), ("8303", "1205", "1205"),
    ("8304", "1210", "1210"), ("8306", "1225", "1225"),
    ("8309", "1265", "1265"),
]

# === CLASIFICADOR INTELIGENTE POR NOMBRE DE CUENTA ===
KEYWORDS_1001 = [
    ("5001", True, ["sueldo", "salario", "basico", "jornal", "horas extra", "recargo",
                     "auxilio transporte", "auxilio de transporte", "rodamiento"]),
    ("5002", True, ["honorario", "honorarios"]),
    ("5002", True, ["comision", "comisiones"]),
    ("5024", True, ["aporte salud", "aporte eps", "aportes a eps", "aporte a eps",
                     "aporte a salud", "cotizacion salud", "aportes eps"]),
    ("5025", True, ["aporte pension", "aporte a pension", "aportes a pension",
                     "pension obligatoria", "cotizacion pension", "fondo pension", "aportes pension"]),
    ("5027", True, ["arl", "riesgo laboral", "riesgos laborales", "riesgos profesionales",
                     "aporte arl", "aporte riesgo", "aportes arl"]),
    ("5023", True, ["parafiscal", "parafiscales", "icbf", "sena", "caja de compensacion",
                     "compensacion familiar", "comfama", "compensar", "cafam", "colsubsidio", "comfenalco"]),
    ("5001", True, ["vacacion", "vacaciones"]),
    ("5001", True, ["cesantia", "cesantias", "interes sobre cesantia", "intereses cesantia",
                     "intereses sobre cesantias"]),
    ("5001", True, ["prima de servicio", "prima servicio", "prima legal"]),
    ("5001", True, ["dotacion", "suministro a trabajador"]),
    ("5001", True, ["incapacidad", "incapacidades"]),
    ("5001", True, ["bonificacion", "bonificaciones"]),
    ("5011", True, ["seguro", "poliza", "prima de seguro", "todo riesgo", "cumplimiento"]),
    ("5005", True, ["arriendo", "arrendamiento", "arrendamientos", "canon", "alquiler"]),
    ("5004", True, ["acueducto", "alcantarillado", "energia", "electrica", "telefono",
                     "telecomunicacion", "internet", "gas", "servicio publico", "servicios publicos",
                     "vigilancia", "correo", "portes"]),
    ("5004", True, ["transporte", "flete", "acarreo", "taxi", "taxis", "buses", "envio", "mensajeria"]),
    ("5055", True, ["impuesto", "ica", "industria y comercio", "predial", "vehiculo",
                     "timbre", "estampilla", "estampillas"]),
    ("5006", True, ["interes bancario", "intereses bancarios", "interes mora", "gmf", "4x1000", "4 x 1000",
                     "comision bancaria", "comisiones bancarias", "gasto financiero", "gastos financieros",
                     "diferencia en cambio", "gravamen", "rendimiento financiero"]),
    ("5010", True, ["gastos de personal", "personal admn", "bienestar", "medicina prepagada",
                     "auxilio funerario", "auxilio educativo", "capacitacion empleado"]),
    ("5016", True, ["viaje", "viatico", "pasaje", "tiquete", "hospedaje", "hotel"]),
    ("5016", True, ["mantenimiento", "reparacion", "adecuacion", "instalacion electrica"]),
    ("5016", True, ["legal", "notarial", "registro", "licencia"]),
    ("5016", True, ["depreciacion", "amortizacion", "agotamiento", "provision"]),
    ("5016", True, ["aseo y cafeteria", "cafeteria", "papeleria", "utiles", "fotocopia", "parqueadero",
                     "casino", "restaurante", "representacion", "suscripcion", "afiliacion",
                     "publicidad", "propaganda", "seminario", "elemento de aseo", "diversos"]),
    ("5007", True, ["inventario", "compra de", "mercancia", "materia prima", "material", "insumo", "repuesto"]),
]

KEYWORDS_1007 = [
    ("4001", ["ingreso operacional", "consultoria", "asesoria", "soporte tecnico",
              "capacitacion", "outsourcing", "desarrollo software", "servicio",
              "honorario recibido", "ingreso actividad"]),
    ("4001", ["venta", "comercio", "producto", "mercancia"]),
    ("4002", ["ingreso no operacional", "extraordinario", "recuperacion"]),
    ("4003", ["arrendamiento recibido", "arriendo recibido", "canon recibido"]),
]

KEYWORDS_1003 = [
    ("1301", ["retencion honorario", "retfte honorario", "retefuente honorario", "rete fuente honorario"]),
    ("1302", ["retencion comision", "retfte comision", "retefuente comision"]),
    ("1303", ["retencion servicio", "retfte servicio", "retefuente servicio"]),
    ("1304", ["retencion arriendo", "retfte arriendo", "retefuente arriendo", "retencion arrendamiento"]),
    ("1305", ["retencion rendimiento", "retfte rendimiento", "retefuente rendimiento", 
              "retencion financiero", "rendimientos financieros", "rendimiento financiero"]),
    ("1306", ["retencion compra", "retfte compra", "retefuente compra", "retencion enajenacion"]),
    ("1307", ["retencion ica", "rete ica", "reteica", "industria y comercio retenido",
              "ica retenido", "impuesto de industria y comercio"]),
    ("1308", ["otras retencion", "otra retencion", "retencion otro", "retencion otros",
              "retenciones por cobrar"]),
    ("1311", ["autorretencion", "auto retencion", "autoretefte", "autoretfte",
              "anticipo autorretencion"]),
]

def normalizar_nombre(nom):
    import unicodedata
    if not nom: return ""
    nom = str(nom).lower().strip()
    nom = ''.join(c for c in unicodedata.normalize('NFD', nom) if unicodedata.category(c) != 'Mn')
    for ch in '.,;:-_/\\()[]{}#"\'&$%@!¿?':
        nom = nom.replace(ch, ' ')
    return ' '.join(nom.split())


def stem_es(palabra):
    if not palabra or len(palabra) < 4: return palabra
    if palabra.endswith('es') and len(palabra) > 5:
        base = palabra[:-2]
        if base.endswith('ion') or base.endswith('dad') or base.endswith('idad'):
            return base
        return base
    if palabra.endswith('s') and not palabra.endswith('ss'):
        return palabra[:-1]
    return palabra


def clasificar_por_nombre(nom, tabla_keywords):
    nom_n = normalizar_nombre(nom)
    if not nom_n: return None
    palabras_nom = set(nom_n.split())
    stems_nom = set(stem_es(p) for p in palabras_nom)
    nom_stemmed = ' '.join(stem_es(p) for p in nom_n.split())

    for item in tabla_keywords:
        if len(item) == 3:
            conc, ded, keywords = item
        else:
            conc, keywords = item
            ded = True
        for kw in keywords:
            if ' ' not in kw: continue
            kw_stemmed = ' '.join(stem_es(p) for p in kw.split())
            if kw_stemmed in nom_stemmed or kw in nom_n:
                return (conc, ded) if len(item) == 3 else conc

    for item in tabla_keywords:
        if len(item) == 3:
            conc, ded, keywords = item
        else:
            conc, keywords = item
            ded = True
        for kw in keywords:
            if ' ' in kw: continue
            kw_stem = stem_es(kw)
            if kw in palabras_nom or kw_stem in stems_nom:
                return (conc, ded) if len(item) == 3 else conc
    return None


def concepto_1001(cta, nom_cta=""):
    if en_rango(cta, "5105", "5105"):
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado
        sc = cta[4:6] if len(cta) >= 6 else cta[4:] if len(cta) > 4 else ""
        for conc, subs in PARAM_1001_NOMINA_SUB.items():
            if sc in subs: return conc, True
        return "5001", True

    if cta[:2] in ("51", "52", "53"):
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado
        for conc, d, h, ded in PARAM_1001_RANGOS:
            if en_rango(cta, d, h): return conc, ded
        return "5016", True

    if cta[:2] == "14":
        for conc, d, h, ded in PARAM_1001_RANGOS:
            if en_rango(cta, d, h): return conc, ded
        if nom_cta:
            resultado = clasificar_por_nombre(nom_cta, KEYWORDS_1001)
            if resultado: return resultado

    return None, True


def buscar_concepto(cta, params, nom_cta="", tabla_keywords=None):
    for c, d, h in params:
        if en_rango(cta, d, h): return c
    if tabla_keywords and nom_cta:
        resultado = clasificar_por_nombre(nom_cta, tabla_keywords)
        if resultado:
            return resultado if isinstance(resultado, str) else resultado[0]
    return ""

# === PROCESAMIENTO PRINCIPAL ===
def procesar_balance(df_balance, df_directorio=None, datos_rues=None, col_map=None, cierra_impuestos=True, dir_central=None):
    """Procesa el balance y retorna un workbook con todos los formatos.
    
    col_map: dict con los índices de columna detectados automáticamente.
    cierra_impuestos: True = usar saldo final para 1355/2365/2408.
                      False = usar débitos - créditos (movimiento del periodo).
    dir_central: dict con directorio centralizado {nit: {dir, depto, mpio, pais, td, dv}}.
    """
    
    # Si no se pasó mapeo, detectar automáticamente
    if col_map is None:
        col_map = detectar_columnas(df_balance)
    
    # Índices de columna (con defaults seguros)
    CI = col_map.get('cuenta', 0)
    NI = col_map.get('nombre', 1)
    TI = col_map.get('nit', 2)
    RI = col_map.get('razon_social', 3)
    DI = col_map.get('debito', 4)
    KI = col_map.get('credito', 5)
    SI = col_map.get('saldo', 6)

    # Cargar directorio externo de direcciones
    dir_externo = {}
    if df_directorio is not None:
        for _, row in df_directorio.iterrows():
            nit_d = safe_str(row.iloc[0])
            if not nit_d: continue
            if '.' in nit_d:
                try: nit_d = str(int(float(nit_d)))
                except: pass
            dir_externo[nit_d] = {
                'dir': safe_str(row.iloc[1]) if len(row) > 1 else "",
                'dp': pad_dpto(safe_str(row.iloc[2])) if len(row) > 2 else "",
                'mp': pad_mpio(safe_str(row.iloc[3])) if len(row) > 3 else "",
                'pais': safe_str(row.iloc[4]) if len(row) > 4 else "169",
            }

    if datos_rues:
        for nit_r, info_r in datos_rues.items():
            if nit_r not in dir_externo:
                dir_externo[nit_r] = {
                    'dir': info_r.get('dir', ''),
                    'dp': info_r.get('dp', ''),
                    'mp': info_r.get('mp', ''),
                    'pais': info_r.get('pais', '169'),
                }

    # === LEER BALANCE (columnas detectadas automáticamente) ===
    bal = []
    for _, row in df_balance.iterrows():
        cta = safe_str(row.iloc[CI])       # Cuenta
        nit = safe_str(row.iloc[TI]) if TI < len(row) else ""  # NIT/Tercero
        if not cta or not nit:
            continue
        # Limpiar NIT
        if '.' in nit:
            try: nit = str(int(float(nit)))
            except: pass
        bal.append({
            'cta': cta,
            'nom_cta': safe_str(row.iloc[NI]) if NI < len(row) else "",          # Nombre cuenta
            'td': detectar_tipo_doc(nit),                                          # Auto-detectado
            'nit': nit,
            'razon': safe_str(row.iloc[RI]) if RI < len(row) else "",             # Razón Social
            'deb': safe_num(row.iloc[DI]) if DI < len(row) else 0,               # Débito
            'cred': safe_num(row.iloc[KI]) if KI < len(row) else 0,              # Crédito
            'saldo': safe_num(row.iloc[SI]) if SI is not None and SI < len(row) else 0,  # Saldo
        })

    # === Helper para valor de cuentas de impuestos ===
    def valor_impuesto(f, tipo='activo'):
        """Calcula el valor a reportar para cuentas de impuestos (1355, 2365, 2408).
        
        Si cierra_impuestos=True: usa saldo final (ya está limpio).
        Si cierra_impuestos=False: usa débitos - créditos (movimiento del periodo).
        
        tipo='activo' (1355, 2408 descontable): débitos = nuevos, créditos = cierre
        tipo='pasivo' (2365, 2408 generado): créditos = nuevos, débitos = cierre
        """
        if cierra_impuestos:
            return abs(f['saldo'])
        else:
            # Movimiento del periodo
            if tipo == 'activo':
                return max(f['deb'] - f['cred'], 0)  # Débitos netos
            else:
                return max(f['cred'] - f['deb'], 0)  # Créditos netos

    # Directorio — Capas de búsqueda:
    # 1. Directorio centralizado (Google Sheet con empresas comunes)
    # 2. Directorio del cliente (archivo subido)
    # 3. Datos del balance (nombre, tipo doc auto-detectado)
    if dir_central is None:
        dir_central = {}
    
    direc = {}
    nits_nuevos = {}  # NITs con dirección encontrada que NO están en el directorio central
    
    for f in bal:
        if f['nit'] not in direc:
            td = f['td'] if f['td'] else detectar_tipo_doc(f['nit'])
            r = f['razon']
            dv = calc_dv(f['nit'])
            d = {'td': td, 'dv': dv,
                 'a1': '', 'a2': '', 'n1': '', 'n2': '',
                 'rs': '', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}
            if td == "13":
                p = r.split()
                if len(p) >= 1: d['a1'] = p[0]
                if len(p) >= 2: d['a2'] = p[1]
                if len(p) >= 3: d['n1'] = p[2]
                if len(p) >= 4: d['n2'] = ' '.join(p[3:])
            else:
                d['rs'] = r
            
            # CAPA 1: Directorio centralizado (Google Sheet)
            if f['nit'] in dir_central:
                dc = dir_central[f['nit']]
                if dc.get('dir'): d['dir'] = dc['dir']
                if dc.get('depto'): d['dp'] = pad_dpto(dc['depto'])
                if dc.get('mpio'): d['mp'] = pad_mpio(dc['mpio'])
                if dc.get('pais'): d['pais'] = dc['pais']
                if dc.get('td'): d['td'] = dc['td']
                if dc.get('dv'): d['dv'] = dc['dv']
                if dc.get('razon') and not d['rs'] and td != "13": d['rs'] = dc['razon']
            
            # CAPA 2: Directorio del cliente (sobreescribe si tiene datos)
            if f['nit'] in dir_externo:
                ext = dir_externo[f['nit']]
                if ext['dir']: d['dir'] = ext['dir']
                if ext['dp']: d['dp'] = pad_dpto(ext['dp'])
                if ext['mp']: d['mp'] = pad_mpio(ext['mp'])
                if ext.get('pais'): d['pais'] = ext['pais']
            
            # Registrar NITs con dirección que NO están en el directorio central
            if d['dir'] and f['nit'] not in dir_central and f['nit'] != NM:
                nits_nuevos[f['nit']] = {
                    'razon': d['rs'] or r,
                    'dir': d['dir'],
                    'dp': d.get('dp', ''),
                    'mp': d.get('mp', ''),
                    'pais': d.get('pais', '169'),
                    'td': d['td'],
                    'dv': d['dv'],
                }
            
            direc[f['nit']] = d
    direc[NM] = {'td': TDM, 'dv': '', 'a1': '', 'a2': '', 'n1': '', 'n2': '',
                 'rs': 'CUANTIAS MENORES', 'dir': '', 'dp': '', 'mp': '', 'pais': '169'}

    # === VALIDACIÓN DE RAZÓN SOCIAL vs DIRECTORIO CENTRAL ===
    alertas_rs = []  # [(nit, rs_balance, rs_central, similitud)]
    for nit, d in direc.items():
        if nit == NM or nit not in dir_central:
            continue
        rs_central = dir_central[nit].get('razon', '').strip().upper()
        if not rs_central:
            continue
        # Razón social del balance
        rs_balance = (d.get('rs', '') or '').strip().upper()
        if not rs_balance:
            # Para personas naturales, reconstruir nombre
            partes = [d.get('a1',''), d.get('a2',''), d.get('n1',''), d.get('n2','')]
            rs_balance = ' '.join(p for p in partes if p).strip().upper()
        if not rs_balance:
            continue
        # Comparar
        if rs_balance == rs_central:
            continue
        sim = difflib.SequenceMatcher(None, rs_balance, rs_central).ratio()
        if sim < 0.95:  # Solo alertar si hay diferencia significativa
            alertas_rs.append((nit, rs_balance, rs_central, sim))

    def t(nit):
        return direc.get(nit, {'td': detectar_tipo_doc(nit), 'dv': calc_dv(nit),
                                'a1': '', 'a2': '', 'n1': '', 'n2': '',
                                'rs': nit, 'dir': '', 'dp': '', 'mp': '', 'pais': '169'})

    # === CREAR WORKBOOK ===
    wb = openpyxl.Workbook()
    hf = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    thin = Side(style='thin', color='808080')

    def nueva_hoja(nombre, headers):
        ws = wb.create_sheet(nombre)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = hfont
            cell.fill = hf
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        return ws

    def escribir_tercero(ws, fila, col, nit, con_pais=False):
        d = t(nit)
        c = col
        # Asegurar que td y dv siempre tengan valor
        td_val = d.get('td', '') or detectar_tipo_doc(nit)
        dv_val = d.get('dv', '') or calc_dv(nit)
        valores = [td_val, nit, dv_val, d['a1'], d['a2'], d['n1'], d['n2'], d['rs'], d['dir'], d['dp'], d['mp']]
        for v in valores:
            cell = ws.cell(fila, c)
            cell.value = str(v) if v else ""
            cell.number_format = '@'
            c += 1
        if con_pais:
            cell = ws.cell(fila, c)
            cell.value = str(d.get('pais', '169') or "169")
            cell.number_format = '@'
            c += 1
        return c

    def fmt(ws, fila, cols):
        for c in cols:
            ws.cell(fila, c).number_format = '#,##0'

    def zebra(ws, fila):
        for c in range(1, ws.max_column + 1):
            ws.cell(fila, c).font = Font(size=10, name='Arial')
            ws.cell(fila, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        if (fila - 2) % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(fila, c).fill = PatternFill('solid', fgColor='F2F7FB')

    resultados = {}

    # ========== PRE-CALCULAR RETENCIONES Y BASES POR NIT ==========
    ret_fte_por_nit = defaultdict(float)
    ret_iva_por_nit = defaultdict(float)
    gastos_por_nit = defaultdict(float)

    for f in bal:
        cta = f['cta']
        # Retenciones que practiqué (pasivo 2365): usar valor_impuesto pasivo
        if en_rango(cta, "236505", "236530"):
            ret_fte_por_nit[f['nit']] += valor_impuesto(f, 'pasivo')
        elif en_rango(cta, "2367", "2367"):
            ret_iva_por_nit[f['nit']] += valor_impuesto(f, 'pasivo')
        if cta[:2] in ("51", "52", "53") and abs(f['saldo']) > 0:
            gastos_por_nit[f['nit']] += abs(f['saldo'])

    # ========== F1001 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Pago Deducible", "Pago No Deducible",
         "IVA Ded", "IVA No Ded", "Ret Fte Renta", "Ret Fte Asumida", "Ret IVA R.Comun", "Ret IVA No Dom"]
    ws = nueva_hoja("F1001 Pagos", h)

    # === REGLAS DE DEDUCIBILIDAD ESPECIALES ===
    # Retorna: 'ded' (100% deducible), 'no_ded' (100% no deducible), 'gmf' (50/50)
    def clasificar_deducibilidad(cta, nom_cta=""):
        nom = normalizar_nombre(nom_cta)
        
        # GMF / 4x1000 → 50% deducible, 50% no deducible
        if 'gmf' in nom or '4x1000' in nom or '4 x 1000' in nom or 'gravamen movimiento' in nom:
            return 'gmf'
        
        # Intereses de mora → 100% no deducible
        if 'interes moratorio' in nom or 'interes de mora' in nom or 'intereses mora' in nom:
            return 'no_ded'
        if cta.startswith('53050504'):  # Cuenta específica intereses mora
            return 'no_ded'
        
        # Gastos no deducibles (cuenta 5395xx o nombre)
        if 'no deducible' in nom or 'no deduci' in nom:
            return 'no_ded'
        if cta.startswith('53950520'):  # Gasto no deducible
            return 'no_ded'
        
        # Multas, sanciones, litigios → no deducible
        if 'multa' in nom or 'sancion' in nom or 'litigio' in nom:
            return 'no_ded'
        if cta.startswith('539520'):  # Multas y sanciones
            return 'no_ded'
        
        # Donaciones → normalmente no deducible (salvo entidades art.257 ET)
        if 'donacion' in nom or 'donaciones' in nom:
            return 'no_ded'
        if cta.startswith('539525'):
            return 'no_ded'
        
        return 'ded'  # Default: deducible

    dic = defaultdict(lambda: [0.0] * 5)
    nits_en_1001 = set()
    # Conceptos de nómina que SOLO van a F2276, NO a F1001
    CONCEPTOS_NOMINA = {"5001", "5024", "5025", "5027", "5023"}
    for f in bal:
        # Usar SALDO FINAL (acumulado anual), no débitos del periodo
        valor = abs(f['saldo'])
        if valor == 0:
            continue  # No incluir filas sin saldo
        conc, ded = concepto_1001(f['cta'], f.get('nom_cta', ''))
        if not conc or conc in CONCEPTOS_NOMINA:
            continue  # Nómina y aportes van SOLO a F2276
        k = (conc, f['nit'])
        
        # Aplicar reglas de deducibilidad
        tipo_ded = clasificar_deducibilidad(f['cta'], f.get('nom_cta', ''))
        
        if tipo_ded == 'gmf':
            # GMF: 50% deducible, 50% no deducible
            dic[k][0] += valor * 0.5   # Pago deducible
            dic[k][1] += valor * 0.5   # Pago no deducible
        elif tipo_ded == 'no_ded':
            # 100% no deducible
            dic[k][1] += valor
        else:
            # 100% deducible (default)
            if ded: dic[k][0] += valor
            else: dic[k][1] += valor
        
        nits_en_1001.add(f['nit'])

    nit_conceptos = defaultdict(list)
    for (conc, nit), v in dic.items():
        nit_conceptos[nit].append((conc, v[0] + v[1]))

    for nit in nit_conceptos:
        ret_total = ret_fte_por_nit.get(nit, 0)
        ret_iva = ret_iva_por_nit.get(nit, 0)
        if ret_total == 0 and ret_iva == 0: continue
        total_pagado = sum(v for _, v in nit_conceptos[nit])
        if total_pagado == 0: continue
        for conc, pago in nit_conceptos[nit]:
            proporcion = pago / total_pagado if total_pagado > 0 else 0
            dic[(conc, nit)][2] += ret_total * proporcion
            dic[(conc, nit)][4] += ret_iva * proporcion

    final = {}
    menores = defaultdict(lambda: [0.0] * 5)
    for (c, n), v in dic.items():
        total_pago = v[0] + v[1]
        if total_pago == 0:
            continue  # No incluir registros sin movimiento
        tiene_retencion = v[2] > 0 or v[4] > 0
        if total_pago < C3UVT and not tiene_retencion:
            for i in range(5): menores[(c, NM)][i] += v[i]
        else:
            final[(c, n)] = v
    for k, v in menores.items():
        if v[0] + v[1] > 0:  # Solo cuantías menores con valor
            if k not in final: final[k] = v

    fila = 2
    for (conc, nit), v in sorted(final.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit, True)
        ws.cell(fila, 14).value = int(v[0])
        ws.cell(fila, 15).value = int(v[1])
        ws.cell(fila, 16).value = 0
        ws.cell(fila, 17).value = 0
        ws.cell(fila, 18).value = int(v[2])
        ws.cell(fila, 19).value = 0
        ws.cell(fila, 20).value = int(v[4])
        ws.cell(fila, 21).value = 0
        fmt(ws, fila, range(14, 22))
        zebra(ws, fila)
        fila += 1
    resultados['F1001 Pagos'] = len(final)

    # ========== F1003 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Base Retencion", "Retencion Acumulada"]
    ws = nueva_hoja("F1003 Retenciones", h)

    dic3 = defaultdict(lambda: [0.0, 0.0])
    for f in bal:
        conc = buscar_concepto(f['cta'], PARAM_1003, f.get('nom_cta', ''), KEYWORDS_1003)
        if not conc: continue
        # 1355 = activo (retenciones que me practicaron)
        val = valor_impuesto(f, 'activo')
        if val > 0:
            dic3[(conc, f['nit'])][1] += val

    # Calcular base gravable: ingresos recibidos del mismo NIT (cuentas 4xxx)
    ingresos_por_nit = defaultdict(float)
    for f in bal:
        if f['cta'][:1] == "4" and abs(f['saldo']) > 0:
            ingresos_por_nit[f['nit']] += abs(f['saldo'])
    
    for (conc, nit), v in dic3.items():
        if v[1] > 0:
            # Base = ingresos del mismo NIT, o estimar desde retención
            v[0] = ingresos_por_nit.get(nit, 0)

    # Filtrar registros con retención en cero
    dic3 = {k: v for k, v in dic3.items() if v[1] > 0}

    fila = 2
    for (conc, nit), v in sorted(dic3.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(v[0])
        ws.cell(fila, 14).value = int(v[1])
        fmt(ws, fila, [13, 14])
        zebra(ws, fila)
        fila += 1
    resultados['F1003 Retenciones'] = len(dic3)

    # ========== F1005 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "IVA Descontable", "IVA Devol Ventas"]
    ws = nueva_hoja("F1005 IVA Descontable", h)

    dic5 = defaultdict(float)
    for f in bal:
        # IVA Descontable: cuentas 240810+ o nombre contiene "descontable"
        if en_rango(f['cta'], "2408", "2408"):
            nom = normalizar_nombre(f.get('nom_cta', ''))
            es_descontable = 'descontable' in nom or f['cta'][:6] >= '240810'
            if es_descontable:
                val = valor_impuesto(f, 'activo')  # IVA descontable es como un activo
                if val > 0:
                    dic5[f['nit']] += val

    fila = 2
    for nit, val in sorted(dic5.items()):
        escribir_tercero(ws, fila, 1, nit)
        ws.cell(fila, 12).value = int(val)
        ws.cell(fila, 13).value = 0
        fmt(ws, fila, [12, 13])
        zebra(ws, fila)
        fila += 1
    resultados['F1005 IVA Descontable'] = len(dic5)

    # ========== F1006 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "IVA Generado", "IVA Devol Compras", "Imp Consumo"]
    ws = nueva_hoja("F1006 IVA Generado", h)

    dic6 = defaultdict(float)
    for f in bal:
        # IVA Generado: cuentas 2408 que NO son descontable
        # 240801-240809 = IVA generado, 240810+ = IVA descontable
        if en_rango(f['cta'], "2408", "2408"):
            nom = normalizar_nombre(f.get('nom_cta', ''))
            es_descontable = 'descontable' in nom or f['cta'][:6] >= '240810'
            if not es_descontable:
                val = valor_impuesto(f, 'pasivo')  # IVA generado es pasivo
                if val > 0:
                    dic6[f['nit']] += val

    fila = 2
    for nit, val in sorted(dic6.items()):
        escribir_tercero(ws, fila, 1, nit)
        ws.cell(fila, 12).value = int(val)
        ws.cell(fila, 13).value = 0
        ws.cell(fila, 14).value = 0
        fmt(ws, fila, [12, 13, 14])
        zebra(ws, fila)
        fila += 1
    resultados['F1006 IVA Generado'] = len(dic6)

    # ========== F1007 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Ingresos Brutos", "Devoluciones"]
    ws = nueva_hoja("F1007 Ingresos", h)

    dic7 = defaultdict(float)
    for f in bal:
        conc = buscar_concepto(f['cta'], PARAM_1007, f.get('nom_cta', ''), KEYWORDS_1007)
        if not conc: continue
        # Usar saldo final (acumulado anual) para ingresos
        valor = abs(f['saldo'])
        if valor > 0:
            dic7[(conc, f['nit'])] += valor

    final7 = {}
    men7 = defaultdict(float)
    for (c, n), v in dic7.items():
        if v < C3UVT: men7[(c, NM)] += v
        else: final7[(c, n)] = v
    for k, v in men7.items():
        if k not in final7: final7[k] = v

    fila = 2
    for (conc, nit), val in sorted(final7.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit, True)
        ws.cell(fila, 14).value = int(val)
        ws.cell(fila, 15).value = 0
        fmt(ws, fila, [14, 15])
        zebra(ws, fila)
        fila += 1
    resultados['F1007 Ingresos'] = len(final7)

    # ========== F1008 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Saldo CxC Dic31"]
    ws = nueva_hoja("F1008 CxC", h)

    dic8 = defaultdict(float)
    for f in bal:
        conc = buscar_concepto(f['cta'], PARAM_1008, f.get('nom_cta', ''))
        if not conc: continue
        s = abs(f['saldo'])
        if s == 0: continue
        dic8[(conc, f['nit'])] += s

    final8 = {}
    men8 = defaultdict(float)
    for (c, n), v in dic8.items():
        if v < C12UVT: men8[(c, NM)] += v
        else: final8[(c, n)] = v
    for k, v in men8.items():
        if k not in final8: final8[k] = v

    fila = 2
    for (conc, nit), val in sorted(final8.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(val)
        fmt(ws, fila, [13])
        zebra(ws, fila)
        fila += 1
    resultados['F1008 CxC'] = len(final8)

    # ========== F1009 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Saldo CxP Dic31"]
    ws = nueva_hoja("F1009 CxP", h)

    dic9 = defaultdict(float)
    for f in bal:
        conc = buscar_concepto(f['cta'], PARAM_1009, f.get('nom_cta', ''))
        if not conc: continue
        s = abs(f['saldo'])
        if s == 0: continue
        dic9[(conc, f['nit'])] += s

    final9 = {}
    men9 = defaultdict(float)
    for (c, n), v in dic9.items():
        if v < C12UVT: men9[(c, NM)] += v
        else: final9[(c, n)] = v
    for k, v in men9.items():
        if k not in final9: final9[k] = v

    fila = 2
    for (conc, nit), val in sorted(final9.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 13).value = int(val)
        fmt(ws, fila, [13])
        zebra(ws, fila)
        fila += 1
    resultados['F1009 CxP'] = len(final9)

    # ========== F1010 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Direccion", "Dpto", "Mpio", "Pais", "Valor Patrimonial", "% Participacion", "Valor Porcentual"]
    ws = nueva_hoja("F1010 Socios", h)

    dic10 = defaultdict(float)
    for f in bal:
        if en_rango(f['cta'], "3105", "3115") or en_rango(f['cta'], "3110", "3110"):
            dic10[f['nit']] += abs(f['saldo'])

    capital_total = sum(dic10.values())
    fila = 2
    for nit, val in sorted(dic10.items()):
        escribir_tercero(ws, fila, 1, nit, True)
        ws.cell(fila, 13).value = int(val)
        pct = round(val / capital_total * 100, 2) if capital_total > 0 else 0
        ws.cell(fila, 14).value = pct / 100
        ws.cell(fila, 14).number_format = '0.00%'
        ws.cell(fila, 15).value = int(val)
        fmt(ws, fila, [13, 15])
        zebra(ws, fila)
        fila += 1
    resultados['F1010 Socios'] = len(dic10)

    # ========== F1012 ==========
    h = ["Concepto", "Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Razon Social", "Saldo Dic31", "Valor Patrimonial"]
    ws = nueva_hoja("F1012 Inversiones", h)

    dic12 = defaultdict(float)
    for f in bal:
        for conc, d, h2 in MAPEO_1012:
            if en_rango(f['cta'], d, h2):
                dic12[(conc, f['nit'])] += abs(f['saldo'])
                break

    fila = 2
    for (conc, nit), val in sorted(dic12.items()):
        ws.cell(fila, 1).value = conc
        escribir_tercero(ws, fila, 2, nit)
        ws.cell(fila, 10).value = int(val)
        ws.cell(fila, 11).value = int(val)
        fmt(ws, fila, [10, 11])
        zebra(ws, fila)
        fila += 1
    resultados['F1012 Inversiones'] = len(dic12)

    # ========== F2276 ==========
    h = ["Tipo Doc", "No ID", "DV", "Apellido1", "Apellido2", "Nombre1", "Nombre2",
         "Direccion", "Dpto", "Mpio", "Pais", "Salarios", "Emol Ecles", "Honor 383",
         "Serv 383", "Comis 383", "Pensiones", "Vacaciones", "Cesantias e Int",
         "Incapacidades", "Otros Pag Lab", "Total Bruto", "Aporte Salud", "Aporte Pension",
         "Sol Pensional", "Vol Empleador", "Vol Trabajador", "AFC", "Ret Fte", "Total Pagos"]
    ws = nueva_hoja("F2276 Rentas Trabajo", h)

    dic26 = defaultdict(lambda: [0.0] * 19)
    for f in bal:
        if not en_rango(f['cta'], "5105", "5105"): continue
        # Usar saldo final (acumulado anual)
        valor = abs(f['saldo'])
        if valor == 0: continue
        nit = f['nit']
        nom = normalizar_nombre(f.get('nom_cta', ''))
        sc = f['cta'][4:6] if len(f['cta']) >= 6 else ""
        clasificado = False

        if sc in ("06", "07", "08", "09", "10", "15"):
            dic26[nit][0] += valor; clasificado = True
        elif sc in ("27",):
            dic26[nit][10] += valor; clasificado = True
        elif sc in ("30", "33"):
            dic26[nit][8] += valor; clasificado = True
        elif sc in ("36",):
            dic26[nit][10] += valor; clasificado = True
        elif sc in ("39",):
            dic26[nit][7] += valor; clasificado = True
        elif sc in ("42", "45"):
            dic26[nit][10] += valor; clasificado = True
        elif sc in ("01", "03", "05"):
            d2 = t(nit)
            if d2['td'] == "13":
                dic26[nit][3] += valor; clasificado = True
        elif sc in ("02",):
            dic26[nit][12] += valor; clasificado = True
        elif sc in ("04",):
            dic26[nit][13] += valor; clasificado = True
        elif sc in ("68", "72", "75"):
            clasificado = True

        if not clasificado and nom:
            palabras = set(nom.split())
            if any(kw in palabras for kw in ["sueldo", "salario", "basico", "jornal"]) or \
               any(kw in nom for kw in ["hora extra", "horas extra", "recargo"]):
                dic26[nit][0] += valor
            elif any(kw in nom for kw in ["cesantia", "interes sobre cesantia", "intereses cesantia"]):
                dic26[nit][8] += valor
            elif any(kw in palabras for kw in ["vacacion", "vacaciones"]):
                dic26[nit][7] += valor
            elif any(kw in nom for kw in ["prima de servicio", "prima servicio"]):
                dic26[nit][10] += valor
            elif any(kw in palabras for kw in ["incapacidad", "incapacidades"]):
                dic26[nit][9] += valor
            elif any(kw in nom for kw in ["aporte salud", "aporte eps", "aportes eps", "aportes a eps"]):
                dic26[nit][12] += valor
            elif any(kw in nom for kw in ["aporte pension", "aportes pension", "aportes a pension"]):
                dic26[nit][13] += valor
            elif any(kw in palabras for kw in ["dotacion", "bonificacion", "auxilio"]):
                dic26[nit][10] += valor
            elif any(kw in palabras for kw in ["honorario", "honorarios"]):
                d2 = t(nit)
                if d2['td'] == "13": dic26[nit][3] += valor
                else: dic26[nit][10] += valor
            elif any(kw in palabras for kw in ["parafiscal", "parafiscales", "icbf", "sena",
                                                "compensar", "comfama", "cafam"]):
                pass
            else:
                dic26[nit][10] += valor

    for f in bal:
        if en_rango(f['cta'], "2365", "2365") and f['nit'] in dic26:
            dic26[f['nit']][17] += valor_impuesto(f, 'pasivo')

    for nit in dic26:
        dic26[nit][11] = sum(dic26[nit][:11])
        dic26[nit][18] = dic26[nit][11]

    fila = 2
    for nit, v in sorted(dic26.items()):
        d = t(nit)
        for ci, val in [(1, d['td']), (2, nit), (3, d['dv']),
                        (4, d['a1']), (5, d['a2']), (6, d['n1']), (7, d['n2']),
                        (8, d['dir']), (9, d['dp']), (10, d['mp']),
                        (11, d.get('pais', '169') or '169')]:
            cell = ws.cell(fila, ci)
            cell.value = val
            cell.number_format = '@'
        for i in range(19):
            ws.cell(fila, 12 + i).value = int(v[i])
            ws.cell(fila, 12 + i).number_format = '#,##0'
        zebra(ws, fila)
        fila += 1
    resultados['F2276 Rentas Trabajo'] = len(dic26)

    # ========== VALIDACION TERCEROS (vs INTERNET) ==========
    validaciones = []
    if datos_rues:
        warn_fill = PatternFill('solid', fgColor='FFF3CD')
        err_fill = PatternFill('solid', fgColor='F8D7DA')
        ok_fill = PatternFill('solid', fgColor='D4EDDA')

        h_val = ["NIT", "Campo", "Valor en Balance", "Valor en Internet", "Estado", "Observación"]
        ws_val = nueva_hoja("Validacion Terceros", h_val)

        fila_val = 2
        for nit_v, info_rues in sorted(datos_rues.items()):
            if nit_v not in direc or nit_v == NM: continue
            d_bal = direc[nit_v]

            dv_calculado = calc_dv(nit_v)
            dv_rues = str(info_rues.get('dv', '')).strip()
            if dv_rues and dv_calculado:
                if dv_calculado == dv_rues:
                    estado_dv = "✅ OK"; fill_dv = ok_fill
                    obs_dv = "DV coincide con internet y cálculo"
                else:
                    estado_dv = "❌ ERROR"; fill_dv = err_fill
                    obs_dv = f"DV calculado={dv_calculado}, Internet={dv_rues}. REVISAR"
                    validaciones.append(('dv_error', nit_v))

                ws_val.cell(fila_val, 1).value = nit_v
                ws_val.cell(fila_val, 1).number_format = '@'
                ws_val.cell(fila_val, 2).value = "Dígito Verificación"
                ws_val.cell(fila_val, 3).value = dv_calculado
                ws_val.cell(fila_val, 4).value = dv_rues
                ws_val.cell(fila_val, 5).value = estado_dv
                ws_val.cell(fila_val, 6).value = obs_dv
                for c in range(1, 7):
                    ws_val.cell(fila_val, c).fill = fill_dv
                    ws_val.cell(fila_val, c).font = Font(size=10, name='Arial')
                    ws_val.cell(fila_val, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                fila_val += 1

            rs_balance = d_bal.get('rs', '').strip()
            if d_bal.get('td') == '13':
                partes = [d_bal.get('a1',''), d_bal.get('a2',''), d_bal.get('n1',''), d_bal.get('n2','')]
                rs_balance = ' '.join(p for p in partes if p).strip()
            rs_rues = info_rues.get('razon_social', '').strip()

            if rs_rues and rs_balance:
                sim = similitud_textos(rs_balance, rs_rues)
                rs_norm_bal = normalizar_texto(rs_balance)
                rs_norm_rues = normalizar_texto(rs_rues)

                if rs_norm_bal == rs_norm_rues:
                    estado_rs = "✅ OK"; fill_rs = ok_fill
                    obs_rs = "Razón social coincide exactamente"
                elif sim >= 0.7:
                    estado_rs = "⚠️ SIMILAR"; fill_rs = warn_fill
                    obs_rs = f"Similitud {sim:.0%}. Verificar ortografía"
                    validaciones.append(('rs_warn', nit_v))
                else:
                    estado_rs = "❌ DIFERENTE"; fill_rs = err_fill
                    obs_rs = f"Similitud {sim:.0%}. Razón social NO coincide con fuente pública"
                    validaciones.append(('rs_error', nit_v))

                ws_val.cell(fila_val, 1).value = nit_v
                ws_val.cell(fila_val, 1).number_format = '@'
                ws_val.cell(fila_val, 2).value = "Razón Social"
                ws_val.cell(fila_val, 3).value = rs_balance
                ws_val.cell(fila_val, 4).value = rs_rues
                ws_val.cell(fila_val, 5).value = estado_rs
                ws_val.cell(fila_val, 6).value = obs_rs
                for c in range(1, 7):
                    ws_val.cell(fila_val, c).fill = fill_rs
                    ws_val.cell(fila_val, c).font = Font(size=10, name='Arial')
                    ws_val.cell(fila_val, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                fila_val += 1

        anchos_val = [18, 22, 40, 40, 16, 50]
        for i, ancho in enumerate(anchos_val, 1):
            ws_val.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        ws_val.freeze_panes = 'A2'

    # ========== ALERTAS RAZÓN SOCIAL ==========
    if alertas_rs:
        h_alert = ["NIT", "Razón Social en Balance", "Razón Social Registrada (DIAN/RUES)", "Similitud %", "Acción Requerida"]
        ws_alert = nueva_hoja("Alertas Razón Social", h_alert)
        
        # Ordenar por similitud (más diferentes primero)
        alertas_rs.sort(key=lambda x: x[3])
        
        fill_rojo = PatternFill('solid', fgColor='FADBD8')
        fill_naranja = PatternFill('solid', fgColor='FDF2E9')
        fill_amarillo = PatternFill('solid', fgColor='FEF9E7')
        
        for idx, (nit, rs_bal, rs_cen, sim) in enumerate(alertas_rs):
            r = idx + 2
            ws_alert.cell(r, 1).value = nit
            ws_alert.cell(r, 2).value = rs_bal
            ws_alert.cell(r, 3).value = rs_cen
            ws_alert.cell(r, 4).value = round(sim * 100, 1)
            ws_alert.cell(r, 4).number_format = '0.0"%"'
            
            # Clasificar severidad
            if sim < 0.5:
                accion = "⛔ CRÍTICO — Razón social completamente diferente. Verificar NIT."
                fill = fill_rojo
            elif sim < 0.75:
                accion = "⚠️ DIFERENTE — Verificar con RUT del tercero."
                fill = fill_naranja
            else:
                accion = "🔍 SIMILAR — Posible abreviación o error de digitación. Confirmar."
                fill = fill_amarillo
            
            ws_alert.cell(r, 5).value = accion
            for c in range(1, 6):
                ws_alert.cell(r, c).fill = fill
                ws_alert.cell(r, c).font = Font(size=9, name='Calibri')
                ws_alert.cell(r, c).alignment = Alignment(wrap_text=True, vertical='center')
        
        # Nota al final
        r_nota = len(alertas_rs) + 3
        ws_alert.cell(r_nota, 1).value = (
            "IMPORTANTE: La razón social en la exógena DEBE coincidir exactamente con el RUT del tercero. "
            "Las diferencias pueden generar rechazo de la información por parte de la DIAN y requerimientos "
            "bajo el Art. 651 del Estatuto Tributario. Verifique cada caso con el RUT actualizado del tercero."
        )
        ws_alert.cell(r_nota, 1).font = Font(size=10, name='Calibri', color='C0392B', bold=True)
        ws_alert.cell(r_nota, 1).alignment = Alignment(wrap_text=True)
        ws_alert.merge_cells(start_row=r_nota, start_column=1, end_row=r_nota, end_column=5)
        
        anchos_alert = [15, 45, 45, 12, 50]
        for i, ancho in enumerate(anchos_alert, 1):
            ws_alert.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        ws_alert.freeze_panes = 'A2'

    # ========== RESUMEN ==========
    n_con_dir = sum(1 for d in direc.values() if d.get('dir', ''))
    n_de_central = sum(1 for nit in direc if nit in dir_central and dir_central.get(nit, {}).get('dir', ''))
    n_de_cliente = sum(1 for nit in direc if nit in dir_externo and dir_externo.get(nit, {}).get('dir', ''))

    wsr = wb.active
    wsr.title = "Resumen"
    wsr['A1'] = "RESUMEN PROCESAMIENTO EXOGENA AG 2025"
    wsr['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4E79')
    wsr['A3'] = "Fecha:"
    wsr['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    wsr['A4'] = "Filas del balance:"
    wsr['B4'] = len(bal)
    wsr['A5'] = "Terceros:"
    wsr['B5'] = len(direc)
    wsr['A6'] = "Con dirección:"
    wsr['B6'] = n_con_dir
    wsr['A7'] = "  → Del directorio centralizado:"
    wsr['B7'] = n_de_central
    wsr['A7'].font = Font(size=9, name='Arial', color='666666')
    wsr['A8'] = "  → Del directorio del cliente:"
    wsr['B8'] = n_de_cliente
    wsr['A8'].font = Font(size=9, name='Arial', color='666666')
    
    rr = 9  # Fila dinámica
    wsr.cell(rr, 1).value = "Direcciones nuevas para agregar:"
    wsr.cell(rr, 2).value = len(nits_nuevos)
    wsr.cell(rr, 1).font = Font(size=9, name='Arial', color='0066CC')
    rr += 1
    
    if alertas_rs:
        wsr.cell(rr, 1).value = "⚠️ Alertas razón social:"
        wsr.cell(rr, 2).value = len(alertas_rs)
        wsr.cell(rr, 1).font = Font(size=9, name='Arial', color='CC6600', bold=True)
        wsr.cell(rr, 2).font = Font(bold=True, color='CC6600')
        rr += 1

    if datos_rues:
        n_dv_err = sum(1 for tipo, _ in validaciones if tipo == 'dv_error')
        n_rs_err = sum(1 for tipo, _ in validaciones if tipo == 'rs_error')
        n_rs_warn = sum(1 for tipo, _ in validaciones if tipo == 'rs_warn')
        wsr.cell(rr, 1).value = "Validados vs Internet:"
        wsr.cell(rr, 2).value = len(datos_rues)
        rr += 1
        wsr.cell(rr, 1).value = "  ❌ DV con error:"
        wsr.cell(rr, 2).value = n_dv_err
        if n_dv_err > 0: wsr.cell(rr, 2).font = Font(bold=True, color='CC0000')
        rr += 1
        wsr.cell(rr, 1).value = "  ❌ Razón social diferente:"
        wsr.cell(rr, 2).value = n_rs_err
        if n_rs_err > 0: wsr.cell(rr, 2).font = Font(bold=True, color='CC0000')
        rr += 1
        wsr.cell(rr, 1).value = "  ⚠️ Razón social similar:"
        wsr.cell(rr, 2).value = n_rs_warn
        rr += 1

    fila_inicio_formatos = rr + 1

    wsr.cell(fila_inicio_formatos, 1).value = "Formato"
    wsr.cell(fila_inicio_formatos, 2).value = "Registros"
    wsr.cell(fila_inicio_formatos, 1).font = Font(bold=True)
    wsr.cell(fila_inicio_formatos, 2).font = Font(bold=True)
    row = fila_inicio_formatos + 1
    for nombre, n in resultados.items():
        wsr.cell(row, 1).value = nombre
        wsr.cell(row, 2).value = n
        row += 1
    wsr.cell(row + 1, 1).value = "TOTAL"
    wsr.cell(row + 1, 1).font = Font(bold=True)
    wsr.cell(row + 1, 2).value = sum(resultados.values())
    wsr.cell(row + 1, 2).font = Font(bold=True)
    wsr.cell(row + 3, 1).value = "F1004, F1011, F1647: requieren datos manuales"

    # === ESTILOS PARA SECCIONES DE ADVERTENCIAS ===
    # Colores
    ROJO_BANNER  = 'C0392B'
    ROJO_CLARO   = 'FADBD8'
    NARANJA_BANNER = 'E67E22'
    NARANJA_CLARO  = 'FDF2E9'
    AZUL_BANNER  = '2471A3'
    AZUL_CLARO   = 'D6EAF8'
    VERDE_BANNER = '1D8348'
    VERDE_CLARO  = 'D5F5E3'
    GRIS_TEXTO   = '2C3E50'
    GRIS_CLARO   = 'F8F9F9'
    BLANCO       = 'FFFFFF'
    
    border_thin = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )
    border_bottom_thick = Border(bottom=Side(style='medium', color='ABB2B9'))
    
    def banner_row(ws, r, texto, color_fondo, cols=3):
        """Fila tipo banner con fondo de color y texto blanco."""
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=color_fondo)
            cell.font = Font(bold=True, size=12, name='Calibri', color=BLANCO)
            cell.alignment = Alignment(vertical='center')
            cell.border = border_thin
        ws.cell(r, 1).value = texto
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        ws.row_dimensions[r].height = 30
    
    def sub_banner(ws, r, texto, color_fondo, color_texto='FFFFFF', cols=3):
        """Sub-encabezado más sutil."""
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=color_fondo)
            cell.font = Font(bold=True, size=10, name='Calibri', color=color_texto)
            cell.alignment = Alignment(vertical='center')
            cell.border = border_thin
        ws.cell(r, 1).value = texto
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        ws.row_dimensions[r].height = 24
    
    def item_row(ws, r, titulo, detalle, color_alt, cols=3):
        """Fila de item con título en col A y detalle en col A de la fila siguiente."""
        # Título
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=color_alt)
            cell.border = border_thin
        ws.cell(r, 1).value = titulo
        ws.cell(r, 1).font = Font(bold=True, size=10, name='Calibri', color=GRIS_TEXTO)
        ws.cell(r, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        # Detalle
        r2 = r + 1
        for c in range(1, cols + 1):
            cell = ws.cell(r2, c)
            cell.fill = PatternFill('solid', fgColor=BLANCO)
            cell.border = border_thin
        ws.cell(r2, 1).value = detalle
        ws.cell(r2, 1).font = Font(size=9, name='Calibri', color='566573')
        ws.cell(r2, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=cols)
        ws.row_dimensions[r2].height = 45
        return r2 + 1  # Retorna la siguiente fila disponible
    
    def valor_row(ws, r, titulo, detalle, color_fondo, cols=3):
        """Fila para validaciones automáticas con valor numérico destacado."""
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=color_fondo)
            cell.border = border_thin
        ws.cell(r, 1).value = titulo
        ws.cell(r, 1).font = Font(bold=True, size=10, name='Calibri', color=GRIS_TEXTO)
        ws.cell(r, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        r2 = r + 1
        for c in range(1, cols + 1):
            cell = ws.cell(r2, c)
            cell.fill = PatternFill('solid', fgColor=BLANCO)
            cell.border = border_thin
        ws.cell(r2, 1).value = detalle
        ws.cell(r2, 1).font = Font(size=9, name='Calibri', color='566573')
        ws.cell(r2, 1).alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=cols)
        ws.row_dimensions[r2].height = 40
        return r2 + 1

    NUM_COLS = 3  # Ancho de las secciones (columnas A-C)

    # ================================================================
    # SECCIÓN 0: COMPROMISO DE CONFIDENCIALIDAD
    # ================================================================
    r = row + 6
    banner_row(wsr, r, "   COMPROMISO DE CONFIDENCIALIDAD Y PROTECCIÓN DE DATOS", VERDE_BANNER, NUM_COLS)
    r += 1
    
    confidencialidad = [
        ("🛡️  USO EXCLUSIVO DE LA INFORMACIÓN",
         "Los archivos suministrados por el cliente (balance de prueba, directorio de terceros "
         "y cualquier otro documento) son utilizados ÚNICA Y EXCLUSIVAMENTE para la generación "
         "de los formatos de información exógena solicitados. No se utilizan para ningún otro fin."),
        ("🗑️  ELIMINACIÓN INMEDIATA",
         "Una vez generada y entregada la información exógena al cliente, TODOS los archivos "
         "recibidos son ELIMINADOS de forma definitiva e irreversible de nuestros sistemas. "
         "No se conservan copias, respaldos ni registros del contenido de los archivos del cliente."),
        ("🚫  NO DIVULGACIÓN A TERCEROS",
         "La información contable, tributaria y de terceros del cliente NO se comparte, vende, "
         "cede ni transfiere a ninguna persona natural o jurídica, bajo ninguna circunstancia. "
         "Esto incluye datos de proveedores, clientes, empleados y cualquier tercero del balance."),
        ("📋  MARCO LEGAL APLICABLE",
         "Este compromiso se rige por: Ley 1581 de 2012 (Protección de Datos Personales y Habeas Data), "
         "Ley 43 de 1990, Art. 63 (Secreto Profesional del Contador Público), "
         "Estatuto Tributario Art. 583 (Reserva de las declaraciones tributarias), "
         "y demás normas concordantes sobre confidencialidad de la información financiera."),
    ]
    
    for i, (titulo, detalle) in enumerate(confidencialidad):
        alt_color = VERDE_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # ================================================================
    # SECCIÓN 1: DESLINDE DE RESPONSABILIDAD
    # ================================================================
    r += 2
    banner_row(wsr, r, "   DESLINDE DE RESPONSABILIDAD", ROJO_BANNER, NUM_COLS)
    r += 1
    
    disclaimer_lines = [
        ("Esta herramienta es un ASISTENTE de generación de información exógena. "
         "Los resultados generados son un BORRADOR que requiere revisión profesional."),
        ("Es responsabilidad exclusiva del CONTADOR PÚBLICO y/o del CONTRIBUYENTE "
         "verificar, ajustar y validar toda la información antes de presentarla a la DIAN."),
        ("Ni el desarrollador ni la herramienta se hacen responsables por errores, "
         "omisiones, sanciones, intereses o multas derivados de información incorrecta, "
         "ni por clasificaciones de deducibilidad que no correspondan al caso particular."),
        ("El uso de esta herramienta implica la aceptación total de estos términos. "
         "Art. 631 y ss. del Estatuto Tributario — Resolución DIAN 000227 de 2025."),
    ]
    
    for i, texto in enumerate(disclaimer_lines):
        bg = ROJO_CLARO if i % 2 == 0 else BLANCO
        for c in range(1, NUM_COLS + 1):
            cell = wsr.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = border_thin
        wsr.cell(r, 1).value = texto
        wsr.cell(r, 1).font = Font(size=9, name='Calibri', color=GRIS_TEXTO)
        wsr.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='center')
        wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
        wsr.row_dimensions[r].height = 35
        r += 1

    # ================================================================
    # SECCIÓN 2: ADVERTENCIAS SOBRE DEDUCIBILIDAD
    # ================================================================
    r += 2
    banner_row(wsr, r, "   ADVERTENCIAS — DEDUCIBILIDAD DE GASTOS", NARANJA_BANNER, NUM_COLS)
    r += 1
    sub_banner(wsr, r, "  El contador DEBE verificar los siguientes aspectos que pueden afectar la deducibilidad fiscal.", NARANJA_CLARO, GRIS_TEXTO, NUM_COLS)
    r += 1
    
    advertencias = [
        ("PAGOS EN EFECTIVO (Art. 771-5 ET)",
         "No son deducibles los pagos en efectivo que superen el MENOR valor entre: "
         "el 40% del total pagado, o el 35% de los costos y deducciones totales. "
         "El contribuyente debe verificar qué pagos se hicieron en efectivo y reclasificarlos."),
        ("SEGURIDAD SOCIAL NO PAGADA (Art. 664, 108 ET)",
         "Los salarios y prestaciones solo son deducibles si el empleador está al día "
         "en aportes a salud, pensión, ARL y parafiscales. Periodos sin pago = gasto NO deducible."),
        ("PRESTACIONES SOCIALES NO PAGADAS (Art. 108 ET)",
         "Cesantías, intereses, primas y vacaciones solo son deducibles cuando se pagan. "
         "Las provisiones contables no pagadas al 31 de diciembre NO son deducibles."),
        ("INDEPENDIENTES SIN SEGURIDAD SOCIAL (Art. 108 par. 2 ET)",
         "Pagos por honorarios/servicios a personas naturales independientes NO son deducibles si no "
         "se verificó el pago de aportes a seguridad social (salud y pensión sobre el 40% del contrato)."),
        ("FACTURA ELECTRÓNICA / RADIAN (Art. 771-2, 616-1 ET)",
         "Los costos y gastos deben estar soportados con factura electrónica o DSNO. "
         "Gastos sin soporte electrónico válido registrado en RADIAN pueden ser rechazados."),
        ("GASTOS SIN SOPORTE DOCUMENTAL (Art. 771-2 ET)",
         "Todo gasto requiere soporte idóneo: factura electrónica, documento equivalente, "
         "comprobante de nómina electrónica o DSNO. Sin soporte = NO deducible."),
        ("GMF / 4x1000 (Art. 115 ET)",
         "Deducible SOLO en un 50%. Esta herramienta ya clasifica el GMF automáticamente "
         "como 50% deducible y 50% no deducible en el formato 1001."),
        ("INTERESES DE MORA Y SANCIONES (Art. 107, 107-1 ET)",
         "Intereses moratorios, multas, sanciones y litigios NO son deducibles. "
         "Esta herramienta los clasifica automáticamente como no deducibles."),
        ("GASTOS EN EL EXTERIOR (Art. 121, 122, 124 ET)",
         "Limitados al 15% de la renta líquida, salvo excepciones. Verificar retención en la fuente "
         "cuando corresponda (Art. 406 ET)."),
        ("DONACIONES (Art. 257 ET)",
         "Solo deducibles como descuento tributario (25%) a entidades del régimen especial "
         "(Art. 19 ET) con certificación. Donaciones a otras entidades NO son deducibles."),
        ("CAUSALIDAD, NECESIDAD Y PROPORCIONALIDAD (Art. 107 ET)",
         "Todo gasto debe cumplir: (1) causalidad con la actividad, (2) necesidad, (3) proporcionalidad. "
         "La herramienta NO evalúa estos criterios — es responsabilidad del contador."),
        ("LÍMITE COSTOS Y DEDUCCIONES (Art. 177-1 ET)",
         "No son aceptables costos/deducciones imputables a ingresos no constitutivos de renta. "
         "Si hay ingresos mixtos, aplicar proporcionalidad en gastos comunes."),
        ("DEPRECIACIONES Y AMORTIZACIONES (Art. 128-141 ET)",
         "Solo deducibles sobre el costo fiscal y dentro de la vida útil fiscal. "
         "Depreciaciones aceleradas o sobre avalúos NO deducibles sin autorización DIAN."),
    ]
    
    for i, (titulo, detalle) in enumerate(advertencias):
        alt_color = NARANJA_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {i+1}. {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # Nota final advertencias
    for c in range(1, NUM_COLS + 1):
        cell = wsr.cell(r, c)
        cell.fill = PatternFill('solid', fgColor=NARANJA_CLARO)
        cell.border = border_thin
    wsr.cell(r, 1).value = ("NOTA: Esta lista NO es exhaustiva. Existen otras limitaciones según tipo de contribuyente, "
                             "régimen fiscal y actividad económica. Consulte siempre con su asesor tributario.")
    wsr.cell(r, 1).font = Font(size=9, name='Calibri', italic=True, color='7D6608')
    wsr.cell(r, 1).alignment = Alignment(wrap_text=True)
    wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    r += 1

    # ================================================================
    # SECCIÓN 3: CRUCES DIAN
    # ================================================================
    r += 2
    banner_row(wsr, r, "   CRUCES DE INFORMACIÓN — POSIBLES REQUERIMIENTOS DIAN", AZUL_BANNER, NUM_COLS)
    r += 1
    sub_banner(wsr, r, "  La DIAN cruza la información exógena con otras fuentes. Verifique estos puntos para evitar requerimientos.", AZUL_CLARO, GRIS_TEXTO, NUM_COLS)
    r += 2

    # --- 3A: Validaciones automáticas ---
    sub_banner(wsr, r, "  A. VALIDACIONES AUTOMÁTICAS (calculadas de su balance)", VERDE_BANNER, BLANCO, NUM_COLS)
    r += 1

    total_ingresos_4 = sum(abs(f['saldo']) for f in bal if f['cta'][:1] == '4' and abs(f['saldo']) > 0)
    total_gastos_5 = sum(abs(f['saldo']) for f in bal if f['cta'][:2] in ('51','52','53') and abs(f['saldo']) > 0)
    total_costos_6 = sum(abs(f['saldo']) for f in bal if f['cta'][:1] == '6' and abs(f['saldo']) > 0)
    total_iva_desc = sum(v for v in dic5.values()) if dic5 else 0
    total_iva_gen = sum(v for v in dic6.values()) if dic6 else 0
    total_ret_fte_2365 = sum(v for v in ret_fte_por_nit.values())
    total_ret_iva_2367 = sum(v for v in ret_iva_por_nit.values())
    total_ret_1355 = sum(v[1] for v in dic3.values()) if dic3 else 0
    total_nomina = sum(abs(f['saldo']) for f in bal if en_rango(f['cta'], '5105', '5105') and abs(f['saldo']) > 0)
    total_f1001 = sum(v[0] + v[1] for v in dic.values())

    validaciones_auto = []

    if total_gastos_5 + total_costos_6 > 0 and total_iva_desc > 0:
        pct_iva = (total_iva_desc / (total_gastos_5 + total_costos_6)) * 100
        validaciones_auto.append((
            f"IVA Descontable / (Gastos + Costos) = {pct_iva:.1f}%",
            f"IVA Descontable: ${total_iva_desc:,.0f}  |  Gastos+Costos: ${total_gastos_5 + total_costos_6:,.0f}\n"
            f"Si la tarifa general es 19%, este % debería estar cercano al 19% (ajustado por exentos/excluidos). "
            f"Diferencias grandes → requerimiento al cruzar con declaraciones de IVA."
        ))

    if total_f1001 > 0 and total_ret_fte_2365 > 0:
        pct_ret = (total_ret_fte_2365 / total_f1001) * 100
        validaciones_auto.append((
            f"Retención Fte Practicada / Pagos F1001 = {pct_ret:.1f}%",
            f"Ret. Fte (2365): ${total_ret_fte_2365:,.0f}  |  Pagos F1001: ${total_f1001:,.0f}\n"
            f"La DIAN cruza retenciones practicadas vs certificados de terceros. Diferencias = requerimiento."
        ))

    if total_ingresos_4 > 0 and total_ret_1355 > 0:
        pct_ret_ing = (total_ret_1355 / total_ingresos_4) * 100
        validaciones_auto.append((
            f"Retenciones Recibidas (1355) / Ingresos = {pct_ret_ing:.1f}%",
            f"Ret. recibidas: ${total_ret_1355:,.0f}  |  Ingresos: ${total_ingresos_4:,.0f}\n"
            f"Cruza F1007 vs retenciones certificadas por clientes. Diferencias activan fiscalización."
        ))

    if total_ingresos_4 > 0:
        validaciones_auto.append((
            f"Total Ingresos (4xxx): ${total_ingresos_4:,.0f}",
            f"Debe coincidir con ingresos brutos de la declaración de renta (renglón 33 del F110). "
            f"Diferencias entre exógena F1007 y renta son la causa #1 de requerimientos."
        ))

    if total_gastos_5 + total_costos_6 > 0:
        validaciones_auto.append((
            f"Total Gastos: ${total_gastos_5:,.0f}  |  Total Costos: ${total_costos_6:,.0f}",
            f"Suma: ${total_gastos_5 + total_costos_6:,.0f}. Debe coincidir con costos y deducciones de renta. "
            f"La DIAN cruza F1001 + F2276 vs declaración de renta."
        ))

    if total_nomina > 0:
        validaciones_auto.append((
            f"Total Nómina (5105): ${total_nomina:,.0f}",
            f"Cruza F2276 vs: (a) PILA, (b) Nómina electrónica mensual, "
            f"(c) Certificados de ingresos y retenciones (F220). Los tres deben ser consistentes."
        ))

    for i, (titulo, detalle) in enumerate(validaciones_auto):
        alt_color = VERDE_CLARO if i % 2 == 0 else GRIS_CLARO
        r = valor_row(wsr, r, f"  ✓ {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # --- 3B: Cruces manuales ---
    r += 1
    sub_banner(wsr, r, "  B. CRUCES QUE REALIZA LA DIAN (verificación manual del contador)", AZUL_BANNER, BLANCO, NUM_COLS)
    r += 1

    cruces_dian = [
        ("EXÓGENA vs DECLARACIÓN DE IVA (Form. 300)",
         "IVA generado (F1006) vs IVA declarado bimestralmente  |  IVA descontable (F1005) vs IVA solicitado  |  "
         "Bases de ingresos F1007 vs bases de IVA. Diferencias → requerimiento ordinario (Art. 686 ET)."),
        ("EXÓGENA vs DECLARACIÓN DE RENTA (Form. 110/210)",
         "Ingresos F1007 = Ingresos brutos renta  |  Pagos F1001 + Nómina F2276 = Costos y deducciones  |  "
         "Retenciones F1003 = Anticipos y retenciones  |  Patrimonio exógena = Patrimonio bruto renta."),
        ("EXÓGENA vs RETENCIÓN EN LA FUENTE (Form. 350)",
         "Retenciones 2365 reportadas en F1001 deben coincidir mes a mes con el Form. 350. "
         "El sistema de la DIAN detecta diferencias mayores a $1 automáticamente."),
        ("EXÓGENA PROPIA vs EXÓGENA DE TERCEROS",
         "Su F1001 (pagos) debe coincidir con el F1007 (ingresos) del tercero. "
         "Sus retenciones practicadas deben coincidir con las retenciones que el tercero reporta en F1003."),
        ("EXÓGENA vs NÓMINA ELECTRÓNICA Y PILA",
         "F2276 vs nómina electrónica mensual, PILA (seguridad social) y certificados F220. "
         "Diferencias en salarios, prestaciones o aportes generan requerimiento."),
        ("EXÓGENA vs FACTURACIÓN ELECTRÓNICA (RADIAN)",
         "La DIAN cruza valores de exógena contra facturas en RADIAN. Gastos sin factura electrónica "
         "o DSNO pueden ser rechazados. Verificar todos los pagos del F1001."),
        ("EXÓGENA vs ENTIDADES FINANCIERAS",
         "Bancos reportan: saldos en cuentas, CDTs, inversiones, préstamos, intereses, GMF pagado. "
         "Cruzan con exógena y renta. Cuentas no reportadas = hallazgo frecuente."),
        ("EXÓGENA vs REGISTROS PÚBLICOS (SNR, RUNT, CCB)",
         "Activos vs: inmuebles en Superintendencia de Notariado, vehículos en RUNT, "
         "inversiones societarias en Cámaras de Comercio. Activos omitidos → requerimiento especial (Art. 685 ET)."),
        ("CUANTÍAS MENORES vs MATERIALIDAD",
         "Si el acumulado con NIT genérico 222222222 es muy alto vs el total de pagos, la DIAN "
         "puede exigir identificación individual. Recomendación: cuantías menores < 5-10% del total."),
        ("PRECIOS DE TRANSFERENCIA (Art. 260-1 ET)",
         "Operaciones con vinculados del exterior deben ser consistentes con la declaración informativa "
         "y documentación comprobatoria. Aplica si patrimonio > 100.000 UVT o ingresos > 61.000 UVT."),
        ("BENEFICIARIO EFECTIVO / RUB (Ley 2155/2021)",
         "Las sociedades deben reportar el Registro Único de Beneficiarios Finales. "
         "La DIAN cruza RUB con pagos al exterior y operaciones con vinculados."),
    ]

    for i, (titulo, detalle) in enumerate(cruces_dian):
        alt_color = AZUL_CLARO if i % 2 == 0 else GRIS_CLARO
        r = item_row(wsr, r, f"  {i+1}. {titulo}", f"       {detalle}", alt_color, NUM_COLS)

    # Nota final — sanción
    r += 1
    banner_row(wsr, r, "   RECUERDE: Sanción por errores en exógena — Art. 651 ET", ROJO_BANNER, NUM_COLS)
    r += 1
    nota_sancion = (
        "La sanción por no enviar información, enviarla con errores o enviarla extemporáneamente "
        "puede ser hasta del 5% de los valores no reportados o reportados incorrectamente, "
        "sin exceder de 15.000 UVT. La DIAN utiliza sistemas de inteligencia artificial y cruces "
        "masivos automatizados. Verifique TODA la información antes de presentarla."
    )
    for c in range(1, NUM_COLS + 1):
        cell = wsr.cell(r, c)
        cell.fill = PatternFill('solid', fgColor=ROJO_CLARO)
        cell.border = border_thin
    wsr.cell(r, 1).value = nota_sancion
    wsr.cell(r, 1).font = Font(size=10, name='Calibri', color=ROJO_BANNER, bold=True)
    wsr.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='center')
    wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    wsr.row_dimensions[r].height = 50
    wsr.column_dimensions['A'].width = 40
    wsr.column_dimensions['B'].width = 30
    wsr.column_dimensions['C'].width = 25

    for ws_name in wb.sheetnames:
        ws2 = wb[ws_name]
        if ws_name.startswith("F"):
            for col in range(1, ws2.max_column + 1):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
            ws2.freeze_panes = 'A2'

    return wb, resultados, len(bal), len(direc), n_con_dir, validaciones, nits_nuevos, alertas_rs


# === INTERFAZ ===
st.markdown('<p class="main-header">📊 Exógena DIAN — AG 2025</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Genera automáticamente los formatos de información exógena a partir del balance de prueba por tercero</p>', unsafe_allow_html=True)

st.markdown("""
<div class="privacy-box">
    <h4><span class="privacy-icon">🔒</span> Compromiso de Confidencialidad y Protección de Datos</h4>
    <p>Su información contable es tratada con la <strong>máxima confidencialidad</strong>. Al utilizar este servicio, usted cuenta con las siguientes garantías:</p>
    <p>🛡️ <strong>Uso exclusivo:</strong> Los archivos que nos suministre serán utilizados <strong>única y exclusivamente</strong> para la generación de los formatos de información exógena solicitados. No se utilizarán para ningún otro fin.</p>
    <p>🗑️ <strong>Eliminación inmediata:</strong> Una vez generada y entregada la información exógena, todos sus archivos (balance de prueba, directorio de terceros y cualquier otro documento) serán <strong>eliminados de forma definitiva e irreversible</strong> de nuestros sistemas.</p>
    <p>🚫 <strong>No divulgación:</strong> No compartimos, vendemos, cedemos ni transferimos su información contable, tributaria o de terceros a ninguna persona natural o jurídica, bajo ninguna circunstancia.</p>
    <p>📋 <strong>Marco legal:</strong> Este compromiso se rige por la Ley 1581 de 2012 (Protección de Datos Personales), el Código de Ética del Contador Público (Ley 43 de 1990, Art. 63 — Secreto Profesional) y el Estatuto Tributario (Art. 583 — Reserva de las declaraciones tributarias).</p>
</div>
""", unsafe_allow_html=True)

st.divider()

col1, col2 = st.columns([2, 1])

with col1:
    st.markdown("### 📁 Sube el Balance de Prueba por Tercero")
    st.markdown("""
    Sube el archivo Excel de tu balance de prueba por tercero.  
    La app **detecta automáticamente** las columnas — no importa el orden ni formato.
    
    Columnas que busca: **Cuenta, Nombre/Descripción, NIT/Tercero, Razón Social, Débitos, Créditos, Saldo Final**
    
    💡 *El tipo de documento (NIT/Cédula) se detecta automáticamente.*
    """)

    uploaded_file = st.file_uploader("Selecciona el archivo Excel", type=['xlsx', 'xls', 'csv'])

with col2:
    st.markdown("### ⚙️ Configuración")
    fila_encabezado = st.number_input("Fila del encabezado", min_value=1, max_value=20, value=1,
                                       help="Fila donde están los títulos de columna (ej: 1, 3)")
    nombre_hoja = st.text_input("Nombre de la hoja (dejar vacío = primera hoja)", value="",
                                 help="Si tu archivo tiene varias hojas, escribe el nombre de la del balance")
    cierra_impuestos = st.toggle("¿La empresa cierra cuentas de impuestos?", value=True,
                                  help="**SÍ:** Las cuentas 1355, 2365, 2408 se cierran cada año → se usa el Saldo Final.\n\n"
                                       "**NO:** Esas cuentas acumulan de años anteriores → se usa Débitos - Créditos del periodo.")

st.divider()

# === DIRECTORIO DE TERCEROS (DIRECCIONES) ===
st.markdown("### 📋 Directorio de Terceros (opcional)")
st.markdown("""
Sube un archivo con las **direcciones** de tus terceros para que aparezcan en los formatos.
Si no lo subes, el programa puede **buscar las direcciones automáticamente** en internet.
""")

col_dir1, col_dir2 = st.columns([2, 1])

with col_dir1:
    st.markdown("""
    | Col A | Col B | Col C | Col D | Col E |
    |-------|-------|-------|-------|-------|
    | NIT | Dirección | Cod Depto | Cod Municipio | Cod País |

    Ejemplo: `890904997 | CL 30 65-15 | 05 | 001 | 169`

    *Los códigos deben ser formato DIAN (ej: Antioquia = 05, Medellín = 001, Colombia = 169)*
    """)
    uploaded_dir = st.file_uploader("Selecciona el archivo de directorio", type=['xlsx', 'xls', 'csv'],
                                     key="directorio")

with col_dir2:
    st.markdown("#### 📥 Plantilla")
    st.markdown("Descarga la plantilla y llénala con los datos de tus terceros:")

    from io import BytesIO as _BytesIO
    wb_plantilla = openpyxl.Workbook()
    ws_p = wb_plantilla.active
    ws_p.title = "Directorio"
    headers_p = ["NIT", "Direccion", "Cod Departamento", "Cod Municipio", "Cod Pais"]
    hf_p = PatternFill('solid', fgColor='1F4E79')
    hfont_p = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    for c, h in enumerate(headers_p, 1):
        cell = ws_p.cell(1, c, h)
        cell.font = hfont_p
        cell.fill = hf_p
        cell.alignment = Alignment(horizontal='center')
    ejemplos = [
        ("890904997", "CL 30 65-15", "05", "001", "169"),
        ("800123456", "KR 7 32-16 OF 801", "11", "001", "169"),
        ("900555111", "AV 6 NORTE 23-45", "76", "001", "169"),
    ]
    for r, ej in enumerate(ejemplos, 2):
        for c, v in enumerate(ej, 1):
            cell = ws_p.cell(r, c)
            cell.value = v
            cell.number_format = '@'

    ws_cod = wb_plantilla.create_sheet("Codigos DIAN")
    ws_cod['A1'] = "DEPARTAMENTOS DIAN"
    ws_cod['A1'].font = Font(bold=True, size=11)
    ws_cod['D1'] = "PAISES"
    ws_cod['D1'].font = Font(bold=True, size=11)
    dptos = [
        ("05","Antioquia"),("08","Atlántico"),("11","Bogotá DC"),("13","Bolívar"),
        ("15","Boyacá"),("17","Caldas"),("18","Caquetá"),("19","Cauca"),("20","Cesar"),
        ("23","Córdoba"),("25","Cundinamarca"),("27","Chocó"),("41","Huila"),
        ("44","La Guajira"),("47","Magdalena"),("50","Meta"),("52","Nariño"),
        ("54","Norte Santander"),("63","Quindío"),("66","Risaralda"),("68","Santander"),
        ("70","Sucre"),("73","Tolima"),("76","Valle del Cauca"),("81","Arauca"),
        ("85","Casanare"),("86","Putumayo"),("88","San Andrés"),("91","Amazonas"),
        ("94","Guainía"),("95","Guaviare"),("97","Vaupés"),("99","Vichada"),
    ]
    ws_cod['A2'] = "Código"; ws_cod['B2'] = "Departamento"
    ws_cod['A2'].font = Font(bold=True); ws_cod['B2'].font = Font(bold=True)
    for i, (cod, nom) in enumerate(dptos, 3):
        ws_cod.cell(i, 1).value = cod
        ws_cod.cell(i, 1).number_format = '@'
        ws_cod.cell(i, 2).value = nom
    ws_cod['D2'] = "Código"; ws_cod['E2'] = "País"
    ws_cod['D2'].font = Font(bold=True); ws_cod['E2'].font = Font(bold=True)
    paises_ej = [("169","Colombia"),("249","Estados Unidos"),("200","México"),
                 ("210","Panamá"),("234","España"),("240","Venezuela")]
    for i, (cod, nom) in enumerate(paises_ej, 3):
        ws_cod.cell(i, 4).value = cod
        ws_cod.cell(i, 5).value = nom

    for col in range(1, 6):
        ws_p.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        ws_cod.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    buf_p = _BytesIO()
    wb_plantilla.save(buf_p)
    buf_p.seek(0)
    st.download_button("📥 Descargar plantilla directorio",
                       data=buf_p.getvalue(),
                       file_name="plantilla_directorio_terceros.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.divider()

if uploaded_file:
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, header=fila_encabezado - 1)
        else:
            sheet = nombre_hoja if nombre_hoja.strip() else 0
            df = pd.read_excel(uploaded_file, header=fila_encabezado - 1, sheet_name=sheet)

        st.success(f"✅ Archivo cargado: **{len(df)} filas** x {len(df.columns)} columnas")

        # === DETECCIÓN AUTOMÁTICA DE COLUMNAS ===
        col_map = detectar_columnas(df)
        cols_ok, cols_faltantes = validar_columnas(col_map)
        
        if cols_ok:
            # Mostrar columnas detectadas
            nombres_campos = {
                'cuenta': '📋 Cuenta', 'nombre': '📝 Nombre', 'nit': '🔢 NIT/Tercero',
                'razon_social': '🏢 Razón Social', 'debito': '➕ Débitos',
                'credito': '➖ Créditos', 'saldo': '💰 Saldo Final'
            }
            detectadas_texto = " | ".join(
                f"{nombres_campos.get(campo, campo)}: **Col {chr(65+idx)}** ({df.columns[idx]})"
                for campo, idx in sorted(col_map.items(), key=lambda x: x[1])
            )
            st.info(f"🔍 **Columnas detectadas:** {detectadas_texto}")
        else:
            st.error(f"❌ No se encontraron las columnas: **{', '.join(cols_faltantes)}**. "
                     f"Verifica que el encabezado esté en la fila correcta. "
                     f"Columnas encontradas: {list(df.columns)}")
            st.stop()

        with st.expander("👁 Vista previa del balance", expanded=False):
            st.dataframe(df.head(20), use_container_width=True)

        df_dir = None
        if uploaded_dir:
            try:
                if uploaded_dir.name.endswith('.csv'):
                    df_dir = pd.read_csv(uploaded_dir, dtype=str)
                else:
                    df_dir = pd.read_excel(uploaded_dir, dtype=str)
                st.success(f"✅ Directorio cargado: **{len(df_dir)} terceros** con direcciones")
            except Exception as e:
                st.warning(f"⚠️ No se pudo leer el directorio: {str(e)}. Se procesará sin direcciones.")

        st.divider()

        col_btn1, col_btn2 = st.columns(2)
        btn_normal = col_btn1.button("🚀 PROCESAR EXÓGENA", type="primary", use_container_width=True)
        btn_internet = col_btn2.button("🌐 PROCESAR + BUSCAR INTERNET", use_container_width=True,
                                        help="Busca direcciones y razón social en RUES, Datos Abiertos y web")

        if btn_normal or btn_internet:
            datos_rues = None
            df_dir_auto = None
            if btn_internet:
                import requests as _req

                st.markdown("---")
                st.markdown("### 🌐 Búsqueda de terceros en internet")

                st.write("**Probando conexión a internet...**")
                _test_urls = [
                    ("Google", "https://www.google.com", 5),
                    ("datos.gov.co", "https://www.datos.gov.co", 5),
                    ("RUES", "https://www.rues.org.co", 5),
                    ("DuckDuckGo", "https://html.duckduckgo.com", 5),
                    ("Bing", "https://www.bing.com", 5),
                ]
                _test_results = []
                for nombre, url, timeout in _test_urls:
                    try:
                        r = _req.get(url, timeout=timeout, headers={
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
                        _test_results.append(f"✅ {nombre}: HTTP {r.status_code}")
                    except _req.exceptions.Timeout:
                        _test_results.append(f"⏱️ {nombre}: Timeout ({timeout}s)")
                    except _req.exceptions.ConnectionError as e:
                        _test_results.append(f"❌ {nombre}: Conexión fallida — {str(e)[:60]}")
                    except Exception as e:
                        _test_results.append(f"❌ {nombre}: {type(e).__name__}: {str(e)[:60]}")

                for tr in _test_results:
                    st.write(tr)

                hay_internet = any("✅" in tr for tr in _test_results)

                if not hay_internet:
                    st.error("⛔ **No hay acceso a internet desde esta aplicación.** "
                             "Streamlit Cloud puede estar bloqueando las conexiones salientes. "
                             "Usa la plantilla de directorio para agregar direcciones manualmente.")
                else:
                    st.write("---")
                    # Extraer NITs únicos del balance
                    col_map_temp = detectar_columnas(df)
                    nit_col_idx = col_map_temp.get('nit', 2)
                    nits_unicos = set()
                    for _, row in df.iterrows():
                        nit_val = safe_str(row.iloc[nit_col_idx])
                        if nit_val and nit_val != NM:
                            if '.' in nit_val:
                                try: nit_val = str(int(float(nit_val)))
                                except: pass
                            nits_unicos.add(nit_val)

                    nits_list = sorted(list(nits_unicos))
                    st.write(f"**{len(nits_list)}** NITs únicos para consultar")

                    progress_bar = st.progress(0, text="🔍 Buscando...")
                    _log_msgs = []
                    def _log_fn(msg): _log_msgs.append(msg)

                    try:
                        datos_rues, api_fallida = buscar_info_terceros(nits_list, progress_bar, log_fn=_log_fn)
                    except Exception as e:
                        _log_msgs.append(f"💥 **ERROR INESPERADO:** {type(e).__name__}: {str(e)}")
                        datos_rues = {}
                        api_fallida = True

                    progress_bar.empty()
                    st.write("**📋 Log de búsqueda:**")
                    for msg in _log_msgs:
                        st.markdown(msg)

                    if datos_rues:
                        n_con_dir_rues = sum(1 for d in datos_rues.values() if d.get('dir'))
                        n_con_rs = sum(1 for d in datos_rues.values() if d.get('razon_social'))
                        st.success(f"✅ **{len(datos_rues)}** terceros encontrados — "
                                   f"{n_con_dir_rues} con dirección, {n_con_rs} con razón social")

                        fuentes_usadas = {}
                        for d in datos_rues.values():
                            f_nombre = d.get('_fuente', 'Desconocida')
                            fuentes_usadas[f_nombre] = fuentes_usadas.get(f_nombre, 0) + 1
                        if fuentes_usadas:
                            desglose = " | ".join(f"{f}: {n}" for f, n in fuentes_usadas.items())
                            st.info(f"🔗 Fuentes: {desglose}")

                        if not uploaded_dir:
                            rows_auto = []
                            for nit_e, datos_e in datos_rues.items():
                                if datos_e.get('dir'):
                                    rows_auto.append({
                                        'NIT': nit_e, 'Direccion': datos_e['dir'],
                                        'Cod_Depto': datos_e['dp'], 'Cod_Mpio': datos_e['mp'],
                                        'Cod_Pais': datos_e.get('pais', '169'),
                                    })
                            if rows_auto:
                                df_dir_auto = pd.DataFrame(rows_auto)
                    elif api_fallida:
                        st.error("❌ No se encontraron datos. Revisa el log arriba.")
                    else:
                        st.warning("⚠️ Las fuentes respondieron pero sin datos para estos NITs.")

                st.markdown("---")

            dir_final = df_dir if df_dir is not None else df_dir_auto

            # Cargar directorio centralizado
            dir_central, dir_central_err = cargar_directorio_central()
            if dir_central:
                st.info(f"📂 Directorio centralizado: **{len(dir_central)} empresas** cargadas")
            
            with st.spinner("Procesando formatos..."):
                wb, resultados, n_filas, n_terceros, n_con_dir, validaciones, nits_nuevos, alertas_rs = procesar_balance(
                    df, dir_final, datos_rues, col_map, cierra_impuestos, dir_central)

            st.markdown('<div class="success-box">', unsafe_allow_html=True)
            st.markdown("### ✅ Procesamiento completado")
            st.markdown('</div>', unsafe_allow_html=True)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Filas procesadas", f"{n_filas:,}")
            c2.metric("Terceros identificados", f"{n_terceros:,}")
            c3.metric("Con dirección", f"{n_con_dir:,}")
            c4.metric("Registros generados", f"{sum(resultados.values()):,}")

            # === ALERTAS DE RAZÓN SOCIAL ===
            if alertas_rs:
                criticos = sum(1 for _, _, _, s in alertas_rs if s < 0.5)
                diferentes = sum(1 for _, _, _, s in alertas_rs if 0.5 <= s < 0.75)
                similares = sum(1 for _, _, _, s in alertas_rs if s >= 0.75)
                
                st.markdown("### ⚠️ Alertas de Razón Social")
                st.warning(f"Se encontraron **{len(alertas_rs)} terceros** cuya razón social en el balance "
                           f"**difiere** de la registrada en la base de datos DIAN/RUES. "
                           f"Revise la hoja **'Alertas Razón Social'** en el Excel.")
                
                ca1, ca2, ca3 = st.columns(3)
                if criticos > 0:
                    ca1.error(f"⛔ **{criticos}** Críticos (muy diferentes)")
                else:
                    ca1.success("✅ Sin críticos")
                if diferentes > 0:
                    ca2.warning(f"⚠️ **{diferentes}** Diferentes")
                else:
                    ca2.success("✅ Sin diferencias")
                if similares > 0:
                    ca3.info(f"🔍 **{similares}** Similares (posible abreviación)")
                else:
                    ca3.success("✅ Sin similares")
                
                with st.expander("Ver detalle de alertas", expanded=False):
                    df_alertas = pd.DataFrame([
                        {
                            'NIT': nit,
                            'En Balance': rs_bal,
                            'Registrada DIAN/RUES': rs_cen,
                            'Similitud': f"{sim*100:.0f}%",
                            'Severidad': '⛔ CRÍTICO' if sim < 0.5 else ('⚠️ DIFERENTE' if sim < 0.75 else '🔍 SIMILAR')
                        }
                        for nit, rs_bal, rs_cen, sim in sorted(alertas_rs, key=lambda x: x[3])
                    ])
                    st.dataframe(df_alertas, use_container_width=True, hide_index=True)

            # Mostrar direcciones nuevas para agregar al directorio central
            if nits_nuevos:
                can_write = sheets_escritura_disponible()
                
                if can_write:
                    # === SINCRONIZACIÓN AUTOMÁTICA ===
                    with st.expander(f"📋 **{len(nits_nuevos)} direcciones nuevas** encontradas", expanded=True):
                        st.caption("Estas direcciones vienen del directorio del cliente y no estaban en tu base centralizada.")
                        df_nuevos = pd.DataFrame([
                            {
                                'NIT': nit,
                                'Razón Social': info['razon'],
                                'Dirección': info['dir'],
                                'Cod Depto': info['dp'],
                                'Cod Municipio': info['mp'],
                            }
                            for nit, info in sorted(nits_nuevos.items())
                        ])
                        st.dataframe(df_nuevos, use_container_width=True, hide_index=True)
                        
                        if st.button("🔄 Agregar automáticamente al directorio centralizado", type="primary", use_container_width=True):
                            with st.spinner("Sincronizando con Google Sheets..."):
                                n_agregados, err_sync = agregar_direcciones_a_sheet(nits_nuevos)
                            if err_sync:
                                st.error(f"❌ Error al sincronizar: {err_sync}")
                                # Fallback: ofrecer CSV
                                csv_nuevos = pd.DataFrame([
                                    {'NIT': nit, 'Razón Social': info['razon'], 'Dirección': info['dir'],
                                     'Cod Depto': info['dp'], 'Cod Municipio': info['mp'],
                                     'Cod País': info['pais'], 'Tipo Doc': info['td'], 'DV': info['dv']}
                                    for nit, info in sorted(nits_nuevos.items())
                                ]).to_csv(index=False)
                                st.download_button("⬇️ Descargar CSV como respaldo", csv_nuevos,
                                                   file_name="direcciones_nuevas.csv", mime="text/csv")
                            else:
                                if n_agregados > 0:
                                    st.success(f"✅ **{n_agregados} direcciones** agregadas al directorio centralizado.")
                                    st.cache_data.clear()  # Limpiar cache para que la próxima vez lea las nuevas
                                else:
                                    st.info("ℹ️ Todas las direcciones ya existían en el directorio. No se agregaron duplicados.")
                else:
                    # === MODO MANUAL (sin gspread configurado) ===
                    with st.expander(f"📋 **{len(nits_nuevos)} direcciones nuevas** para agregar al directorio centralizado", expanded=False):
                        st.caption("Estas direcciones vienen del directorio del cliente pero NO están en tu Google Sheet centralizado. "
                                   "Cópialas para que estén disponibles para futuros clientes.")
                        st.info("💡 **Tip:** Configura la cuenta de servicio de Google para que se agreguen automáticamente. "
                                "Ver instrucciones en las líneas 25-43 del código.")
                        df_nuevos = pd.DataFrame([
                            {
                                'NIT': nit,
                                'Razón Social': info['razon'],
                                'Dirección': info['dir'],
                                'Cod Depto': info['dp'],
                                'Cod Municipio': info['mp'],
                                'Cod País': info['pais'],
                                'Tipo Doc': info['td'],
                                'DV': info['dv'],
                            }
                            for nit, info in sorted(nits_nuevos.items())
                        ])
                        st.dataframe(df_nuevos, use_container_width=True, hide_index=True)
                        csv_nuevos = df_nuevos.to_csv(index=False)
                        st.download_button(
                            "⬇️ Descargar CSV para pegar en Google Sheets",
                            csv_nuevos,
                            file_name="direcciones_nuevas.csv",
                            mime="text/csv"
                        )

            if validaciones:
                n_dv_err = sum(1 for tipo, _ in validaciones if tipo == 'dv_error')
                n_rs_err = sum(1 for tipo, _ in validaciones if tipo == 'rs_error')
                n_rs_warn = sum(1 for tipo, _ in validaciones if tipo == 'rs_warn')

                st.markdown("### 🔍 Validación vs Internet")
                cv1, cv2, cv3 = st.columns(3)
                if n_dv_err > 0: cv1.error(f"❌ **{n_dv_err}** DV con error")
                else: cv1.success("✅ Todos los DV correctos")
                if n_rs_err > 0: cv2.error(f"❌ **{n_rs_err}** razones sociales diferentes")
                else: cv2.success("✅ Razones sociales correctas")
                if n_rs_warn > 0: cv3.warning(f"⚠️ **{n_rs_warn}** razones sociales similares (verificar)")
                else: cv3.success("✅ Sin advertencias")
                st.info("📋 Revisa la hoja **'Validacion Terceros'** en el archivo descargado para ver el detalle completo.")
            elif btn_internet and datos_rues:
                st.success("✅ Validación completada: no se encontraron diferencias.")

            st.markdown("### 📋 Resultados por formato")
            cols = st.columns(5)
            for i, (nombre, n) in enumerate(resultados.items()):
                with cols[i % 5]:
                    st.metric(nombre.split(" ", 1)[0], f"{n} reg")

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.divider()
            st.download_button(
                label="📥 DESCARGAR ARCHIVO PROCESADO",
                data=output.getvalue(),
                file_name=f"exogena_AG2025_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            st.info("💡 **Formatos manuales:** F1004 (Descuentos), F1011 (Declaraciones) y F1647 (Ingresos para terceros) requieren datos que no están en el balance.")
            
            st.markdown("""
            <div style="background: #eafaf1; border-radius: 8px; padding: 0.8rem 1rem; border-left: 4px solid #27ae60; margin-top: 1rem;">
                <span style="font-size: 0.9rem; color: #1e8449;">
                    🔒 <strong>Recordatorio:</strong> Sus archivos serán eliminados de nuestros sistemas una vez entregada la información. 
                    Su información contable no será utilizada para ningún otro fin ni compartida con terceros.
                </span>
            </div>
            """, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ Error al leer el archivo: {str(e)}")
        st.info("Verifica que la fila del encabezado sea correcta y que el archivo tenga la estructura esperada.")

st.divider()
st.markdown("""
<div style="text-align: center; color: #999; font-size: 0.85rem;">
    Exógena DIAN AG 2025 | Resolución 000227/2025 | UVT $49.799<br>
    ⚠️ Esta herramienta es un asistente. El contador debe validar los resultados antes de presentar a la DIAN.<br>
    🔒 Su información es tratada con total confidencialidad — Ley 1581/2012, Ley 43/1990 Art. 63.
</div>
""", unsafe_allow_html=True)


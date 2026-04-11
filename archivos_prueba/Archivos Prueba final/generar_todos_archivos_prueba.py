"""
GENERADOR COMPLETO DE ARCHIVOS DE PRUEBA - EXÓGENA DIAN
========================================================
Genera archivos Excel para probar TODAS las herramientas de exogenadian.com
"""
import openpyxl
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os, sys

OUT = os.path.dirname(os.path.abspath(__file__))

def money_fmt(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.number_format = '#,##0'

def header_style(ws, row, ncols):
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 4, 40)

# ============================================================
# 1. EXOGENA - Balance con Terceros
# ============================================================
def gen_exogena():
    print(">>> Generando TEST_Exogena_Balance_Terceros.xlsx ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Terceros"

    headers = ["Cuenta", "Nombre Cuenta", "NIT", "Razon Social", "Debito", "Credito"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    header_style(ws, 1, len(headers))

    # Data: (cuenta, nombre, nit, razon_social, debito, credito)
    data = [
        # === GASTOS (51xx, 52xx) - F1001 Pagos a terceros ===
        ("510506","Sueldos","1001234567","Juan Carlos Rodriguez Perez",185000000,0),
        ("510506","Sueldos","1019876543","Maria Fernanda Lopez Garcia",142000000,0),
        ("510506","Sueldos","80234567","Carlos Alberto Martinez Ruiz",168000000,0),
        ("510506","Sueldos","52198765","Ana Maria Gomez Torres",96000000,0),
        ("510506","Sueldos","1098765432","Pedro Antonio Sanchez",78000000,0),
        ("510530","Cesantias","1001234567","Juan Carlos Rodriguez Perez",15416667,0),
        ("510530","Cesantias","1019876543","Maria Fernanda Lopez Garcia",11833333,0),
        ("510530","Cesantias","80234567","Carlos Alberto Martinez Ruiz",14000000,0),
        ("510530","Cesantias","52198765","Ana Maria Gomez Torres",8000000,0),
        ("510530","Cesantias","1098765432","Pedro Antonio Sanchez",6500000,0),
        ("510533","Intereses sobre cesantias","1001234567","Juan Carlos Rodriguez Perez",1850000,0),
        ("510533","Intereses sobre cesantias","1019876543","Maria Fernanda Lopez Garcia",1420000,0),
        ("510536","Prima de servicios","1001234567","Juan Carlos Rodriguez Perez",15416667,0),
        ("510536","Prima de servicios","1019876543","Maria Fernanda Lopez Garcia",11833333,0),
        ("510536","Prima de servicios","80234567","Carlos Alberto Martinez Ruiz",14000000,0),
        ("510539","Vacaciones","1001234567","Juan Carlos Rodriguez Perez",7708333,0),
        ("510539","Vacaciones","1019876543","Maria Fernanda Lopez Garcia",5916667,0),
        ("510568","Aportes ARL","800123456","ARL Sura",8400000,0),
        ("510569","Aportes EPS","800234567","EPS Sura",28000000,0),
        ("510570","Aportes AFP","800345678","AFP Proteccion",22400000,0),
        ("510572","Aportes ICBF SENA CCF","800456789","Caja Compensacion Colsubsidio",9000000,0),
        ("5110","Honorarios","900123456","Consultoria ABC S.A.S.",45000000,0),
        ("5110","Honorarios","860012345","Asesoria Legal Ltda",22000000,0),
        ("5110","Honorarios","1098765432","Pedro Antonio Sanchez",11000000,0),
        ("5115","Impuestos","800567890","Secretaria Distrital Hacienda",22000000,0),
        ("5120","Arrendamientos","900678901","Inmobiliaria Centro S.A.",36000000,0),
        ("5120","Arrendamientos","19876543","Roberto Diaz Mejia",24000000,0),
        ("5125","Seguros","800678901","Seguros Bolivar S.A.",12000000,0),
        ("5130","Servicios","900345678","Servicios Integrados S.A.S.",32000000,0),
        ("5130","Servicios","900456789","TechSolutions Colombia S.A.S.",18000000,0),
        ("5130","Servicios","811234567","Aseo y Cafeteria Plus S.A.S.",6400000,0),
        ("5135","Gastos legales","860012345","Asesoria Legal Ltda",9000000,0),
        ("5140","Servicios publicos","899999999","EPM Empresas Publicas Medellin",8500000,0),
        ("5140","Servicios publicos","830037248","Codensa S.A. ESP",5500000,0),
        ("5145","Gastos de viaje","900901234","Agencia Viajes Aviatur",18000000,0),
        ("5150","Publicidad y propaganda","900234567","Publicidad Digital SAS",10000000,0),
        ("5155","Mantenimiento","900567890","Mantenimiento Express Ltda",7000000,0),
        ("5160","Depreciacion","","",15000000,0),
        ("5195","Gastos diversos","","",6000000,0),
        ("5305","Gastos financieros","860002964","Banco de Bogota S.A.",12000000,0),
        ("5305","Gastos financieros","890903938","Bancolombia S.A.",6000000,0),
        ("530525","Multas y sanciones","800123456","DIAN",4800000,0),
        # === GASTOS VENTAS (52xx) ===
        ("5205","Gastos personal ventas","1045678901","Laura Cristina Vargas",72000000,0),
        ("5210","Honorarios ventas","900111222","Marketing Pro S.A.S.",15000000,0),
        ("5220","Comisiones ventas","900234567","Agencia Comercial Norte Ltda",25000000,0),
        # === COSTOS (61xx) ===
        ("6135","Costo mercancias vendidas","900789012","Suministros Industriales S.A.",275000000,0),
        ("6135","Costo mercancias vendidas","900890123","Materiales del Caribe Ltda",150000000,0),
        ("6135","Costo mercancias vendidas","800456789","Ferreteria Nacional S.A.S.",50000000,0),
        ("6135","Costo mercancias vendidas","900901234","Importadora Global S.A.",180000000,0),
        ("6140","Costo servicios prestados","900345678","Servicios Integrados S.A.S.",85000000,0),
        # === RETENCIONES PRACTICADAS (2365xx, 2367xx) - F1003 ===
        ("236505","Retencion salarios","1001234567","Juan Carlos Rodriguez Perez",0,2850000),
        ("236505","Retencion salarios","1019876543","Maria Fernanda Lopez Garcia",0,1920000),
        ("236505","Retencion salarios","80234567","Carlos Alberto Martinez Ruiz",0,3150000),
        ("236505","Retencion salarios","52198765","Ana Maria Gomez Torres",0,980000),
        ("236515","Retencion honorarios","900123456","Consultoria ABC S.A.S.",0,4950000),
        ("236515","Retencion honorarios","860012345","Asesoria Legal Ltda",0,2420000),
        ("236515","Retencion honorarios","1098765432","Pedro Antonio Sanchez",0,1210000),
        ("236525","Retencion servicios","900345678","Servicios Integrados S.A.S.",0,1280000),
        ("236525","Retencion servicios","900456789","TechSolutions Colombia S.A.S.",0,720000),
        ("236525","Retencion servicios","900567890","Mantenimiento Express Ltda",0,280000),
        ("236525","Retencion servicios","811234567","Aseo y Cafeteria Plus S.A.S.",0,256000),
        ("236530","Retencion arrendamientos","900678901","Inmobiliaria Centro S.A.",0,1260000),
        ("236530","Retencion arrendamientos","19876543","Roberto Diaz Mejia",0,840000),
        ("236540","Retencion compras","900789012","Suministros Industriales S.A.",0,6875000),
        ("236540","Retencion compras","900890123","Materiales del Caribe Ltda",0,3750000),
        ("236540","Retencion compras","800456789","Ferreteria Nacional S.A.S.",0,1250000),
        ("236540","Retencion compras","900901234","Importadora Global S.A.",0,4500000),
        ("236545","Retencion rendimientos financieros","860002964","Banco de Bogota S.A.",0,595000),
        ("236545","Retencion rendimientos financieros","890903938","Bancolombia S.A.",0,420000),
        ("236570","Retencion ICA","900123456","Consultoria ABC S.A.S.",0,450000),
        ("236570","Retencion ICA","900345678","Servicios Integrados S.A.S.",0,320000),
        ("236570","Retencion ICA","900789012","Suministros Industriales S.A.",0,275000),
        # === RETENCION IVA (2367xx) ===
        ("236701","Retencion IVA regimen comun","900123456","Consultoria ABC S.A.S.",0,1282500),
        ("236701","Retencion IVA regimen comun","900345678","Servicios Integrados S.A.S.",0,912000),
        ("236701","Retencion IVA regimen comun","900456789","TechSolutions Colombia S.A.S.",0,513000),
        ("236701","Retencion IVA regimen comun","900678901","Inmobiliaria Centro S.A.",0,360000),
        ("236701","Retencion IVA regimen comun","900789012","Suministros Industriales S.A.",0,1959375),
        # === IVA DESCONTABLE (2408xx) - F1005 ===
        ("240805","IVA descontable compras 19%","900789012","Suministros Industriales S.A.",52250000,0),
        ("240805","IVA descontable compras 19%","900890123","Materiales del Caribe Ltda",28500000,0),
        ("240805","IVA descontable compras 19%","800456789","Ferreteria Nacional S.A.S.",9500000,0),
        ("240810","IVA descontable servicios","900345678","Servicios Integrados S.A.S.",6080000,0),
        ("240810","IVA descontable servicios","900456789","TechSolutions Colombia S.A.S.",3420000,0),
        ("240810","IVA descontable servicios","900678901","Inmobiliaria Centro S.A.",2400000,0),
        # === IVA GENERADO (2408xx credito) - F1006 ===
        ("240801","IVA generado ventas 19%","900345678","Cliente Grande S.A.",0,38000000),
        ("240801","IVA generado ventas 19%","900111222","Distribuidora Norte S.A.",0,28500000),
        ("240801","IVA generado ventas 19%","900222333","Comercial Sur Ltda",0,19000000),
        ("240801","IVA generado ventas 19%","901012345","Gobierno Municipal",0,9500000),
        # === INGRESOS (41xx, 42xx) - F1007 ===
        ("4135","Comercio al por mayor y menor","900345678","Cliente Grande S.A.",0,200000000),
        ("4135","Comercio al por mayor y menor","900111222","Distribuidora Norte S.A.",0,150000000),
        ("4135","Comercio al por mayor y menor","900222333","Comercial Sur Ltda",0,100000000),
        ("4140","Servicios","901012345","Gobierno Municipal",0,50000000),
        ("4140","Servicios","900456789","TechSolutions Colombia S.A.S.",0,120000000),
        ("4140","Servicios","900345678","Cliente Grande S.A.",0,150000000),
        ("4210","Intereses financieros","860002964","Banco de Bogota S.A.",0,5500000),
        ("4210","Intereses financieros","890903938","Bancolombia S.A.",0,3000000),
        ("4220","Comisiones","900111222","Distribuidora Norte S.A.",0,12000000),
        ("4250","Recuperaciones","800567890","Recuperacion cartera antigua",0,5500000),
        # === CUENTAS POR COBRAR (13xx) - F1008 ===
        ("1305","Clientes nacionales","900345678","Cliente Grande S.A.",185000000,150000000),
        ("1305","Clientes nacionales","900111222","Distribuidora Norte S.A.",120000000,100000000),
        ("1305","Clientes nacionales","900222333","Comercial Sur Ltda",85000000,70000000),
        ("1305","Clientes nacionales","901012345","Gobierno Municipal",50000000,35000000),
        ("1310","Cuentas por cobrar socios","31234567","Carlos Eduardo Pinzon Socio",15000000,0),
        ("1330","Anticipos y avances","900567890","Mantenimiento Express Ltda",8000000,5000000),
        ("1355","Anticipo impuestos renta","800123456","DIAN",25000000,0),
        ("1355","Retencion en la fuente","800123456","DIAN",10000000,0),
        # === CUENTAS POR PAGAR (22xx, 23xx) - F1009 ===
        ("2205","Proveedores nacionales","900789012","Suministros Industriales S.A.",180000000,215000000),
        ("2205","Proveedores nacionales","900890123","Materiales del Caribe Ltda",80000000,98000000),
        ("2205","Proveedores nacionales","800456789","Ferreteria Nacional S.A.S.",45000000,52000000),
        ("2335","Costos por pagar","900345678","Servicios Integrados S.A.S.",5000000,18000000),
        # === SOCIOS (31xx) - F1010 ===
        ("3115","Aportes sociales","31234567","Carlos Eduardo Pinzon Socio",0,120000000),
        ("3115","Aportes sociales","41567890","Maria Isabel Vargas Socia",0,80000000),
        # === INVENTARIOS (14xx) - F1012 ===
        ("1435","Mercancias no fabricadas","900789012","Suministros Industriales S.A.",75000000,60000000),
        ("1435","Mercancias no fabricadas","900901234","Importadora Global S.A.",45000000,35000000),
        # === OTROS ACTIVOS ===
        ("1105","Caja general","","",5000000,3500000),
        ("1110","Bancos nacionales","860002964","Banco de Bogota S.A.",450000000,360000000),
        ("1110","Bancos nacionales","890903938","Bancolombia S.A.",400000000,360000000),
        ("1120","Cuentas de ahorro","890903938","Bancolombia S.A.",45000000,0),
        ("1205","Inversiones CDT","860002964","Banco de Bogota S.A.",25000000,0),
        ("1504","Terrenos","","",80000000,0),
        ("1516","Construcciones","","",150000000,0),
        ("1524","Equipo de oficina","","",35000000,0),
        ("1528","Equipo computacion","","",22000000,0),
        ("1592","Depreciacion acumulada","","",0,45000000),
        ("1610","Marcas","","",12000000,0),
        # === OTROS PASIVOS/PATRIMONIO ===
        ("2105","Obligaciones financieras","860002964","Banco de Bogota S.A.",30000000,80000000),
        ("2404","IVA por pagar","800123456","DIAN",25000000,38000000),
        ("2505","Salarios por pagar","","",8000000,12000000),
        ("2510","Cesantias consolidadas","","",0,6500000),
        ("2515","Intereses sobre cesantias","","",0,780000),
        ("2520","Prima de servicios","","",0,5400000),
        ("2525","Vacaciones consolidadas","","",0,3200000),
        ("2610","Provision obligaciones fiscales","","",0,4000000),
        ("3205","Reservas obligatorias","","",0,35000000),
        ("3605","Utilidad del ejercicio","","",0,75000000),
        ("3705","Utilidades acumuladas","","",0,75000000),
    ]

    total_deb = sum(r[4] for r in data)
    total_cre = sum(r[5] for r in data)

    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        money_fmt(ws, i, [5, 6])

    r = len(data) + 2
    ws.cell(row=r, column=4, value="TOTALES")
    ws.cell(row=r, column=4).font = Font(bold=True)
    ws.cell(row=r, column=5, value=total_deb)
    ws.cell(row=r, column=6, value=total_cre)
    money_fmt(ws, r, [5, 6])

    auto_width(ws)
    fp = os.path.join(OUT, "TEST_Exogena_Balance_Terceros.xlsx")
    wb.save(fp)
    print(f"    OK: {len(data)} filas, Deb={total_deb:,.0f}, Cre={total_cre:,.0f}")
    print(f"    Diferencia Deb-Cre = {total_deb - total_cre:,.0f}")

    # --- Direcciones ---
    print(">>> Generando TEST_Exogena_Direcciones.xlsx ...")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Directorio Terceros"
    headers2 = ["NIT", "Razon Social", "Direccion", "Ciudad", "Departamento", "Pais", "Telefono", "Email"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    header_style(ws2, 1, len(headers2))

    dirs = [
        ("1001234567","Juan Carlos Rodriguez Perez","Calle 100 # 19-61 Oficina 1201","Bogota D.C.","Cundinamarca","Colombia","3001234567","jcrodriguez@email.com"),
        ("1019876543","Maria Fernanda Lopez Garcia","Carrera 7 No. 71-21 Torre B Piso 12","Bogota D.C.","Cundinamarca","Colombia","3109876543","mflopez@email.com"),
        ("80234567","Carlos Alberto Martinez Ruiz","Diagonal 45A Bis # 23-15","Medellin","Antioquia","Colombia","3148765432","camartinez@email.com"),
        ("52198765","Ana Maria Gomez Torres","Transversal 93 # 53-44 Local 205","Cali","Valle del Cauca","Colombia","3175432198","amgomez@email.com"),
        ("1098765432","Pedro Antonio Sanchez","Avenida El Dorado # 68C-61","Bogota D.C.","Cundinamarca","Colombia","3204567890","pasanchez@email.com"),
        ("1045678901","Laura Cristina Vargas","Calle 80 # 10-43 Apto 501","Barranquilla","Atlantico","Colombia","3156789012","lcvargas@email.com"),
        ("31234567","Carlos Eduardo Pinzon Socio","Carrera 15 # 93-75 Oficina 604","Bogota D.C.","Cundinamarca","Colombia","3012345678","cepinzon@email.com"),
        ("41567890","Maria Isabel Vargas Socia","Calle 72 # 10-07 Piso 8","Bogota D.C.","Cundinamarca","Colombia","3013456789","mivargas@email.com"),
        ("19876543","Roberto Diaz Mejia","Km 5 Via Bogota-Tunja Vereda El Roble","Tunja","Boyaca","Colombia","3167890123","rdiaz@email.com"),
        ("900123456","Consultoria ABC S.A.S.","Calle 93 # 11A-28 Oficina 301","Bogota D.C.","Cundinamarca","Colombia","6012345678","info@consultoriaabc.com"),
        ("860012345","Asesoria Legal Ltda","Carrera 9 # 74-08 Oficina 503","Bogota D.C.","Cundinamarca","Colombia","6019876543","contacto@asesorialegal.com"),
        ("900345678","Servicios Integrados S.A.S.","Calle 26 # 69D-91 Torre 1 Piso 9","Bogota D.C.","Cundinamarca","Colombia","6014567890","admin@serviciosintegrados.com"),
        ("900456789","TechSolutions Colombia S.A.S.","Carrera 43A # 1 Sur-200 Piso 14","Medellin","Antioquia","Colombia","6044567890","info@techsolutions.co"),
        ("900567890","Mantenimiento Express Ltda","Avenida 6 Norte # 25N-60","Cali","Valle del Cauca","Colombia","6023456789","servicio@mantenimientoexp.com"),
        ("900678901","Inmobiliaria Centro S.A.","Calle 52 # 47-28 Edificio Coltejer P4","Medellin","Antioquia","Colombia","6045678901","arriendos@inmobcentro.com"),
        ("900789012","Suministros Industriales S.A.","Zona Franca Km 17 Via 40 Bod. 12","Barranquilla","Atlantico","Colombia","6053456789","ventas@suministrosind.com"),
        ("900890123","Materiales del Caribe Ltda","Carrera 54 # 72-80 Bodega 5","Cartagena","Bolivar","Colombia","6056789012","pedidos@materialescaribe.com"),
        ("800456789","Ferreteria Nacional S.A.S.","Calle 13 # 65-10 Zona Industrial","Bogota D.C.","Cundinamarca","Colombia","6017890123","ventas@ferreterianacional.com"),
        ("900901234","Importadora Global S.A.","Transversal 25 # 45-120 Bod. Internacional","Bucaramanga","Santander","Colombia","6078901234","importaciones@globalsa.com"),
        ("900234567","Publicidad Digital SAS","Carrera 35 # 48-38 Oficina 201","Bucaramanga","Santander","Colombia","6079012345","hola@publidigital.com"),
        ("811234567","Aseo y Cafeteria Plus S.A.S.","Calle 30 # 15-22","Pereira","Risaralda","Colombia","6063456789","admin@aseoplus.com"),
        ("800123456","DIAN","Carrera 8 # 6C-38 Edificio San Agustin","Bogota D.C.","Cundinamarca","Colombia","6017428100","contactenos@dian.gov.co"),
        ("800234567","EPS Sura","Carrera 64B # 49A-30","Medellin","Antioquia","Colombia","6043658585","afiliaciones@epssura.com"),
        ("800345678","AFP Proteccion","Carrera 43A # 1-50 San Fernando Plaza","Medellin","Antioquia","Colombia","6043106060","info@proteccion.com"),
        ("800456789","Caja Compensacion Colsubsidio","Calle 26 # 25-50","Bogota D.C.","Cundinamarca","Colombia","6017450600","contacto@colsubsidio.com"),
        ("800567890","Secretaria Distrital Hacienda","Carrera 30 # 25-90 Piso 1 CAD","Bogota D.C.","Cundinamarca","Colombia","6013385000","impuestos@shd.gov.co"),
        ("800678901","Seguros Bolivar S.A.","Calle 72 # 10-71 Piso 5","Bogota D.C.","Cundinamarca","Colombia","6013124600","seguros@bolivar.com"),
        ("860002964","Banco de Bogota S.A.","Calle 36 # 7-47 Piso 15","Bogota D.C.","Cundinamarca","Colombia","6013382000","empresarial@bancodebogota.com"),
        ("890903938","Bancolombia S.A.","Carrera 48 # 26-85 Avenida Los Industriales","Medellin","Antioquia","Colombia","6045105610","contacto@bancolombia.com"),
        ("830037248","Codensa S.A. ESP","Carrera 13A # 93-66","Bogota D.C.","Cundinamarca","Colombia","6016019115","servicio@enel.com"),
        ("899999999","EPM Empresas Publicas Medellin","Carrera 58 # 42-125","Medellin","Antioquia","Colombia","6044444115","linea@epm.com.co"),
        ("900901234","Agencia Viajes Aviatur","Avenida 19 # 4-62","Bogota D.C.","Cundinamarca","Colombia","6015879710","reservas@aviatur.com"),
        ("901012345","Gobierno Municipal Bogota","Calle 11 # 8-17 Palacio Lievano","Bogota D.C.","Cundinamarca","Colombia","6013813000","alcaldia@bogota.gov.co"),
        ("900111222","Distribuidora Norte S.A.","Diagonal 78B # 45-10 Zona Franca","Barranquilla","Atlantico","Colombia","6053786543","ventas@distnorte.com"),
        ("900222333","Comercial Sur Ltda","Carrera 100 # 5-169 CC Unicentro","Cali","Valle del Cauca","Colombia","6023981234","info@comercialsur.com"),
        # Internacionales
        ("EIN123456","Tech Supplies Inc.","1200 NW 78th Ave Suite 200","Miami, FL 33126","Florida","Estados Unidos","+13057654321","sales@techsupplies.com"),
        ("RUC12345678","Importaciones Panama S.A.","Calle 50 Edificio Global Tower Piso 8","Ciudad de Panama","Panama","Panama","+5072654321","ventas@imppanama.com"),
        ("CIF98765432","Consultores Madrid S.L.","Gran Via 28 Planta 5","Madrid","Madrid","Espana","+34915678901","info@consultoresmadrid.es"),
        # Mas ciudades Colombia
        ("900333444","Distribuciones Eje Cafetero","Carrera 23 # 64-20","Manizales","Caldas","Colombia","6068812345","ventas@disteje.com"),
        ("900444555","Logistica Santandereana","Calle 36 # 19-22 Zona Industrial Chimita","Giron","Santander","Colombia","6076389012","operaciones@logisanti.com"),
        ("900555666","Agroindustria del Huila","Km 3 Via Neiva-Rivera","Neiva","Huila","Colombia","6088756789","contacto@agrohuila.com"),
        ("900666777","Pesquera del Pacifico","Avenida del Rio # 12-45 Puerto","Tumaco","Narino","Colombia","6027234567","ventas@pesquerapacifico.com"),
        ("900777888","Minera del Cesar","Carrera 19 # 15-40 Centro","Valledupar","Cesar","Colombia","6055876543","admin@mineracesar.com"),
        ("900888999","Textiles del Tolima","Calle 42 # 4-80 Zona Industrial","Ibague","Tolima","Colombia","6082765432","pedidos@textolima.com"),
    ]

    for i, row in enumerate(dirs, 2):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)

    auto_width(ws2)
    fp2 = os.path.join(OUT, "TEST_Exogena_Direcciones.xlsx")
    wb2.save(fp2)
    print(f"    OK: {len(dirs)} direcciones")


# ============================================================
# 2. RENTA 110 - Balance con gastos no deducibles
# ============================================================
def gen_renta110():
    print(">>> Generando TEST_Renta110_Balance.xlsx ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance de Prueba"

    headers = ["Cuenta", "Nombre Cuenta", "Debito", "Credito"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    header_style(ws, 1, len(headers))

    # OPEN trial balance (before closing entries): NO 3605 account.
    # P&L accounts (4,5,6) carry the utility decomposition.
    # Rule: Total Deb = Total Cre when A+G+C = ContraA+P+Pt_base+I
    # No sub-account duplication: if parent 5115 exists, no 511570 etc.
    # For non-deductibles we use sub-accounts WITHOUT their parent.

    # 1) Define balance sheet saldos (4-digit only)
    activos = [
        ("1105","Caja",1500000),
        ("1110","Bancos",130000000),
        ("1120","Cuentas de ahorro",45000000),
        ("1205","Acciones",25000000),
        ("1305","Clientes",40000000),
        ("1310","Cuentas por cobrar a vinculados",10000000),
        ("1355","Anticipo de impuestos y contribuciones",18000000),
        ("1380","Deudores varios",5000000),
        ("1399","Provision deudores",-2500000),
        ("1435","Mercancias no fabricadas por la empresa",25000000),
        ("1504","Terrenos",80000000),
        ("1516","Construcciones y edificaciones",150000000),
        ("1524","Equipo de oficina",35000000),
        ("1528","Equipo de computacion y comunicacion",22000000),
        ("1540","Flota y equipo de transporte",45000000),
        ("1592","Depreciacion acumulada",-52000000),
        ("1610","Marcas",12000000),
        ("1698","Amortizacion acumulada",-3000000),
        ("1705","Gastos pagados por anticipado",4000000),
        ("1805","Bienes de arte y cultura",8000000),
    ]
    pasivos = [
        ("2105","Bancos nacionales",50000000),
        ("2205","Proveedores nacionales",35000000),
        ("2210","Proveedores del exterior",8000000),
        ("2335","Costos y gastos por pagar",13000000),
        ("2360","Dividendos por pagar",12000000),
        ("2365","Retencion en la fuente",3000000),
        ("2367","Impuesto a las ventas retenido",2000000),
        ("2368","Impuesto de industria y comercio retenido",1300000),
        ("2404","De renta y complementarios",7000000),
        ("2408","Impuesto sobre las ventas por pagar",13000000),
        ("2505","Salarios por pagar",4000000),
        ("2510","Cesantias consolidadas",8500000),
        ("2515","Intereses sobre cesantias",1020000),
        ("2520","Prima de servicios",7000000),
        ("2525","Vacaciones consolidadas",4200000),
        ("2610","Para obligaciones fiscales",5000000),
        ("2615","Para obligaciones laborales",3500000),
        ("2805","Anticipos y avances recibidos",6000000),
    ]
    patrimonio_base = [
        ("3105","Capital suscrito y pagado",200000000),
        ("3115","Aportes sociales",50000000),
        ("3205","Reserva legal",35000000),
        ("3305","Reservas estatutarias",15000000),
        ("3705","Utilidades acumuladas",75000000),
    ]

    # 2) Expenses: use sub-accounts ONLY where non-deducibles exist (no parent)
    #    use parent accounts where no non-deducible sub-account
    gastos_list = [
        # 5105 - Personal (no non-ded subs, use parent)
        ("5105","Gastos de personal - Salarios y prestaciones",185000000),
        # 5110 - Honorarios (use parent)
        ("5110","Honorarios",45000000),
        # 5115 - Impuestos: has non-ded sub 511570, 511595 -> use subs only
        ("511505","Impuesto de industria y comercio",12000000),
        ("511510","Impuesto predial",6000000),
        ("511515","Impuesto de vehiculos",4000000),
        ("511570","Donaciones sin requisitos legales",8000000),      # NO DEDUCIBLE
        ("511595","Aportes parafiscales sin pago efectivo",2500000), # NO DEDUCIBLE
        # 5120 - Arrendamientos (use parent)
        ("5120","Arrendamientos",36000000),
        # 5125 - Seguros (use parent)
        ("5125","Seguros",12000000),
        # 5130 - Servicios (use parent)
        ("5130","Servicios",28000000),
        # 5135 - Legales: has non-ded sub 513535 -> use subs
        ("513505","Gastos notariales",3000000),
        ("513510","Tramites y licencias",6000000),
        ("513535","Impuesto al patrimonio",15000000),               # NO DEDUCIBLE
        # 5140 - Servicios publicos (use parent)
        ("5140","Gastos de servicios publicos",14000000),
        # 5145 - Viaje (use parent)
        ("5145","Gastos de viaje",18000000),
        # 5150 - Publicidad (use parent)
        ("5150","Publicidad y propaganda",10000000),
        # 5155 - Mantenimiento (use parent)
        ("5155","Mantenimiento y reparaciones",7000000),
        # 5160 - Depreciacion: has non-ded sub 516099 -> use subs
        ("516005","Depreciacion edificaciones",5000000),
        ("516010","Depreciacion maquinaria",4000000),
        ("516015","Depreciacion vehiculos",3000000),
        ("516020","Depreciacion equipo computo",3000000),
        ("516099","Depreciacion en exceso fiscal",5000000),         # NO DEDUCIBLE
        # 5195 - Diversos: has non-ded subs -> use subs
        ("519505","Gastos de representacion",2000000),
        ("519530","Regalos y atenciones a clientes",4000000),       # NO DEDUCIBLE
        ("519595","Otros gastos diversos sin soporte",3000000),     # NO DEDUCIBLE
        # 5199 - Provisiones (is itself non-deductible, use as-is)
        ("5199","Provisiones del ejercicio",7000000),               # NO DEDUCIBLE
        # 5295 - Gastos no operacionales: non-ded subs -> use subs
        ("529505","Impuestos asumidos no deducibles",1800000),      # NO DEDUCIBLE
        ("529510","Intereses presuntos",1200000),                   # NO DEDUCIBLE
        ("529595","Gastos de ejercicios anteriores",2500000),       # NO DEDUCIBLE
        # 5305 - Financieros: has non-ded subs -> use subs
        ("530505","Intereses de mora DIAN",3200000),                # NO DEDUCIBLE
        ("530510","Intereses bancarios deducibles",14000000),
        ("530520","Comisiones bancarias",4000000),
        ("530525","Multas y sanciones",4800000),                    # NO DEDUCIBLE
        ("530535","Sanciones tributarias",2000000),                  # NO DEDUCIBLE
        # 5310 - Perdida venta bienes (non-deductible, use parent)
        ("5310","Perdida en venta y retiro de bienes",3500000),     # NO DEDUCIBLE
        # 5315 - Extraordinarios (non-deductible, use parent)
        ("5315","Gastos extraordinarios",2000000),                  # NO DEDUCIBLE
    ]
    costos_list = [
        ("6135","Costo de mercancias vendidas",420000000),
        ("6140","Costo de servicios",85000000),
    ]

    # 3) Calculate totals
    total_activos_pos = sum(s for _,_,s in activos if s > 0)
    total_activos_neg = sum(abs(s) for _,_,s in activos if s < 0)
    total_activos_net = total_activos_pos - total_activos_neg
    total_pasivos = sum(s for _,_,s in pasivos)
    total_pt_base = sum(s for _,_,s in patrimonio_base)
    total_gastos = sum(s for _,_,s in gastos_list)
    total_costos = sum(s for _,_,s in costos_list)

    # 4) Calculate required income so Deb = Cre
    # Deb = activos_pos + gastos + costos
    # Cre = activos_neg + pasivos + pt_base + ingresos
    # For Deb = Cre: ingresos = activos_pos + gastos + costos - activos_neg - pasivos - pt_base
    required_income = total_activos_pos + total_gastos + total_costos - total_activos_neg - total_pasivos - total_pt_base

    # Distribute income across accounts
    ingresos_list = [
        ("4135","Comercio al por mayor y menor",int(required_income * 0.55)),
        ("4140","Servicios",int(required_income * 0.30)),
        ("4145","Actividades inmobiliarias",int(required_income * 0.04)),
        ("4210","Intereses",int(required_income * 0.02)),
        ("4220","Comisiones",int(required_income * 0.03)),
        ("4230","Dividendos y participaciones",int(required_income * 0.01)),
        ("4250","Recuperaciones",int(required_income * 0.015)),
        ("4295","Ingresos diversos",int(required_income * 0.005)),
        ("429595","Ingresos no constitutivos de renta",3000000),
    ]
    # Adjust first account to absorb rounding
    total_ingresos_raw = sum(s for _,_,s in ingresos_list)
    diff = required_income - total_ingresos_raw
    ingresos_list[0] = (ingresos_list[0][0], ingresos_list[0][1], ingresos_list[0][2] + diff)
    total_ingresos = sum(s for _,_,s in ingresos_list)
    utilidad = total_ingresos - total_gastos - total_costos

    # 5) Build data rows: (cuenta, nombre, deb, cre)
    data = []
    for acct, name, saldo in activos:
        if saldo >= 0:
            data.append((acct, name, saldo, 0))
        else:
            data.append((acct, name, 0, abs(saldo)))
    for acct, name, saldo in pasivos:
        data.append((acct, name, 0, saldo))
    for acct, name, saldo in patrimonio_base:
        data.append((acct, name, 0, saldo))
    # NO 3605 in open trial balance - utility is implicit in P&L accounts
    for acct, name, saldo in ingresos_list:
        data.append((acct, name, 0, saldo))
    for acct, name, saldo in gastos_list:
        data.append((acct, name, saldo, 0))
    for acct, name, saldo in costos_list:
        data.append((acct, name, saldo, 0))

    total_deb = sum(r[2] for r in data)
    total_cre = sum(r[3] for r in data)

    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        money_fmt(ws, i, [3, 4])

    r = len(data) + 2
    ws.cell(row=r, column=2, value="TOTALES")
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=3, value=total_deb)
    ws.cell(row=r, column=4, value=total_cre)
    money_fmt(ws, r, [3, 4])

    no_deducibles = [
        ("511570",8000000),("513535",15000000),("511595",2500000),
        ("516099",5000000),("519530",4000000),("519595",3000000),
        ("5199",7000000),("530505",3200000),("530525",4800000),
        ("530535",2000000),("5310",3500000),("5315",2000000),
        ("529505",1800000),("529510",1200000),("529595",2500000),
    ]
    total_no_ded = sum(x[1] for x in no_deducibles)

    auto_width(ws)
    fp = os.path.join(OUT, "TEST_Renta110_Balance.xlsx")
    wb.save(fp)
    print(f"    OK: {len(data)} cuentas")
    print(f"    Ingresos: {total_ingresos:,.0f}")
    print(f"    Gastos: {total_gastos:,.0f}")
    print(f"    Costos: {total_costos:,.0f}")
    print(f"    Utilidad: {utilidad:,.0f}")
    print(f"    Total Activos neto: {total_activos_net:,.0f}")
    print(f"    Total Pasivos: {total_pasivos:,.0f}")
    print(f"    Patrimonio base + Utilidad: {total_pt_base+utilidad:,.0f}")
    print(f"    A=P+Pt? {total_activos_net:,.0f} vs {total_pasivos+total_pt_base+utilidad:,.0f} -> {'OK' if total_activos_net==total_pasivos+total_pt_base+utilidad else 'DESCUADRADO!'}")
    print(f"    Total Deb: {total_deb:,.0f} | Total Cre: {total_cre:,.0f} | Dif: {total_deb-total_cre:,.0f}")
    print(f"    Gastos NO deducibles ({len(no_deducibles)} partidas): {total_no_ded:,.0f}")


# ============================================================
# 3. IVA 300 - Balance + Listado DIAN
# ============================================================
def gen_iva300():
    print(">>> Generando TEST_IVA300_Balance.xlsx ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance de Prueba"

    headers = ["Cuenta", "Nombre Cuenta", "Saldo Inicial", "Debito", "Credito", "Saldo Final"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    header_style(ws, 1, len(headers))

    # (cuenta, nombre, saldo_ini, debito, credito, saldo_final)
    data = [
        # INGRESOS GRAVADOS
        ("413505","Venta productos gravados 19%",0,3000000,285000000,282000000),
        ("413510","Venta productos gravados 5%",0,500000,48000000,47500000),
        ("414005","Servicios gravados 19%",0,2000000,185000000,183000000),
        ("414010","Servicios gravados 5%",0,0,38000000,38000000),
        # INGRESOS EXCLUIDOS/EXENTOS
        ("4145","Ingresos excluidos de IVA",0,0,25000000,25000000),
        ("4150","Ingresos exentos (tarifa 0%)",0,0,18000000,18000000),
        # DEVOLUCIONES
        ("4175","Devoluciones en ventas gravadas",0,12000000,0,-12000000),
        # IVA GENERADO
        ("240801","IVA generado ventas tarifa general 19%",0,8500000,89060000,80560000),
        ("240802","IVA generado ventas tarifa 5%",0,600000,4300000,3700000),
        # IVA DESCONTABLE
        ("240805","IVA pagado compras gravadas 19%",0,49400000,6000000,43400000),
        ("240810","IVA pagado servicios 19%",0,16150000,2500000,13650000),
        ("240815","IVA pagado compras 5%",0,3650000,400000,3250000),
        # IVA RETENIDO QUE PRACTICAMOS
        ("2367","IVA retenido por pagar",0,2200000,8900000,6700000),
        # RETENCION IVA QUE NOS PRACTICARON
        ("135517","Retencion de IVA a favor",0,6800000,0,6800000),
        # COMPRAS
        ("6135","Compras mercancias gravadas 19%",0,260000000,4000000,256000000),
        ("613510","Compras mercancias gravadas 5%",0,73000000,1500000,71500000),
        ("6210","Compras no gravadas / excluidas",0,42000000,0,42000000),
        # GASTOS CON IVA
        ("5130","Servicios contratados gravados",0,85000000,0,85000000),
        ("5120","Arrendamientos gravados",0,38000000,0,38000000),
        ("5155","Mantenimiento y reparaciones",0,12000000,0,12000000),
        # BALANCE - ACTIVOS
        ("1105","Caja",2500000,18000000,15000000,5500000),
        ("1110","Bancos",95000000,680000000,650000000,125000000),
        ("1305","Clientes",42000000,395000000,365000000,72000000),
        ("1435","Inventarios",38000000,260000000,245000000,53000000),
        ("1524","Equipo de oficina",15000000,0,0,15000000),
        ("1528","Equipo de computacion",22000000,0,0,22000000),
        ("1592","Depreciacion acumulada",-12000000,0,3500000,-15500000),
        # BALANCE - PASIVOS
        ("2105","Obligaciones financieras",0,15000000,45000000,30000000),
        ("2205","Proveedores nacionales",32000000,285000000,310000000,57000000),
        ("2365","Retencion en la fuente",0,9000000,12500000,3500000),
        ("2505","Salarios por pagar",0,14000000,18000000,4000000),
        ("2510","Cesantias consolidadas",5000000,0,3500000,8500000),
        # BALANCE - PATRIMONIO
        ("3105","Capital suscrito y pagado",100000000,0,0,100000000),
        ("3205","Reserva legal",15000000,0,0,15000000),
        ("3705","Utilidades acumuladas",45000000,0,0,45000000),
        # GASTOS ADICIONALES
        ("5105","Gastos de personal",0,145000000,0,145000000),
        ("5110","Honorarios",0,32000000,0,32000000),
        ("5305","Gastos financieros",0,8000000,0,8000000),
    ]

    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        money_fmt(ws, i, [3,4,5,6])

    total_deb = sum(r[3] for r in data)
    total_cre = sum(r[4] for r in data)
    r = len(data) + 2
    ws.cell(row=r, column=2, value="TOTALES")
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=4, value=total_deb)
    ws.cell(row=r, column=5, value=total_cre)
    money_fmt(ws, r, [4,5])

    auto_width(ws)
    fp = os.path.join(OUT, "TEST_IVA300_Balance.xlsx")
    wb.save(fp)
    print(f"    OK: {len(data)} cuentas, Deb={total_deb:,.0f}, Cre={total_cre:,.0f}")

    # --- Listado DIAN (formato factura electronica) ---
    print(">>> Generando TEST_IVA300_Listado_DIAN.xlsx ...")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Listado FE DIAN"

    headers2 = ["Tipo de Documento","Folio","Fecha Emision","NIT Emisor","Nombre Emisor",
                 "NIT Receptor","Nombre Receptor","IVA","Rete IVA","Rete Renta","Total","Grupo"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    header_style(ws2, 1, len(headers2))

    # Facturas emitidas (ventas) - IVA generado. Valores con leves diferencias vs balance.
    nit_empresa = "901234567"
    emitidas = [
        ("Factura","FE-001","15/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","900345678","Cliente Grande S.A.",19000000,0,0,119000000,"Emitido"),
        ("Factura","FE-002","18/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","900111222","Distribuidora Norte S.A.",9500000,0,0,59500000,"Emitido"),
        ("Factura","FE-003","22/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","900222333","Comercial Sur Ltda",7600000,0,0,47600000,"Emitido"),
        ("Factura","FE-004","25/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","901012345","Gobierno Municipal",3800000,0,0,23800000,"Emitido"),
        ("Factura","FE-005","28/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","900456789","Tech Solutions S.A.S.",11400000,0,0,71400000,"Emitido"),
        ("Factura","FE-006","02/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900345678","Cliente Grande S.A.",15200000,0,0,95200000,"Emitido"),
        ("Factura","FE-007","05/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900678901","Distribuidora Centro",5700000,0,0,35700000,"Emitido"),
        ("Factura","FE-008","10/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900111222","Distribuidora Norte S.A.",4750000,0,0,29750000,"Emitido"),
        ("Factura","FE-009","15/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900890123","Comercial Sur Ltda",3800000,0,0,23800000,"Emitido"),
        ("Factura","FE-010","18/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900333444","Distribuciones Eje Cafetero",2850000,0,0,17850000,"Emitido"),
        # Facturas con tarifa 5%
        ("Factura","FE-011","20/01/2026",nit_empresa,"TechServicios Colombia S.A.S.","900444555","Agroindustria Huila",1200000,0,0,25200000,"Emitido"),
        ("Factura","FE-012","12/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900555666","Pesquera Pacifico",800000,0,0,16800000,"Emitido"),
        ("Factura","FE-013","25/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900666777","Minera del Cesar",2100000,0,0,44100000,"Emitido"),
        # Nota credito emitida (reduce ventas)
        ("Nota Credito","NC-001","28/02/2026",nit_empresa,"TechServicios Colombia S.A.S.","900345678","Cliente Grande S.A.",2280000,0,0,14280000,"Emitido"),
    ]
    # Facturas recibidas (compras) - IVA descontable
    recibidas = [
        ("Factura","FV-2345","10/01/2026","900789012","Suministros Industriales S.A.",nit_empresa,"TechServicios Colombia S.A.S.",9500000,1425000,1375000,50000000,"Recibido"),
        ("Factura","FV-2400","15/01/2026","900890123","Materiales del Caribe Ltda",nit_empresa,"TechServicios Colombia S.A.S.",5415000,812250,787500,28500000,"Recibido"),
        ("Factura","FV-3456","20/01/2026","800456789","Ferreteria Nacional S.A.S.",nit_empresa,"TechServicios Colombia S.A.S.",1805000,270750,262500,9500000,"Recibido"),
        ("Factura","FV-4567","25/01/2026","900901234","Importadora Global S.A.",nit_empresa,"TechServicios Colombia S.A.S.",8550000,1282500,1250000,45000000,"Recibido"),
        ("Factura","FV-5001","28/01/2026","900345678","Servicios Integrados S.A.S.",nit_empresa,"TechServicios Colombia S.A.S.",6080000,912000,640000,32000000,"Recibido"),
        ("Factura","FV-5002","30/01/2026","900456789","TechSolutions Colombia",nit_empresa,"TechServicios Colombia S.A.S.",3420000,513000,360000,18000000,"Recibido"),
        ("Factura","FV-5003","05/02/2026","900678901","Inmobiliaria Centro S.A.",nit_empresa,"TechServicios Colombia S.A.S.",1140000,0,630000,6000000,"Recibido"),
        ("Factura","FV-5004","10/02/2026","900789012","Suministros Industriales S.A.",nit_empresa,"TechServicios Colombia S.A.S.",9500000,1425000,1375000,50000000,"Recibido"),
        ("Factura","FV-5005","15/02/2026","900890123","Materiales del Caribe Ltda",nit_empresa,"TechServicios Colombia S.A.S.",5415000,812250,787500,28500000,"Recibido"),
        ("Factura","FV-5006","18/02/2026","900567890","Mantenimiento Express Ltda",nit_empresa,"TechServicios Colombia S.A.S.",1330000,0,140000,7000000,"Recibido"),
        # Compras gravadas 5%
        ("Factura","FV-6001","08/01/2026","900444555","Agroindustria Huila",nit_empresa,"TechServicios Colombia S.A.S.",1825000,0,0,38325000,"Recibido"),
        ("Factura","FV-6002","20/02/2026","900555666","Pesquera Pacifico",nit_empresa,"TechServicios Colombia S.A.S.",1750000,0,0,36750000,"Recibido"),
        # Compras excluidas
        ("Factura","FV-7001","12/01/2026","899999999","EPM",nit_empresa,"TechServicios Colombia S.A.S.",0,0,0,8500000,"Recibido"),
        ("Factura","FV-7002","25/01/2026","830037248","Codensa S.A. ESP",nit_empresa,"TechServicios Colombia S.A.S.",0,0,0,5500000,"Recibido"),
        ("Factura","FV-7003","05/02/2026","899999999","EPM",nit_empresa,"TechServicios Colombia S.A.S.",0,0,0,8500000,"Recibido"),
        ("Factura","FV-7004","20/02/2026","830037248","Codensa S.A. ESP",nit_empresa,"TechServicios Colombia S.A.S.",0,0,0,5500000,"Recibido"),
    ]

    all_fe = emitidas + recibidas
    for i, row in enumerate(all_fe, 2):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
        money_fmt(ws2, i, [8,9,10,11])

    iva_gen = sum(r[7] for r in emitidas)
    iva_desc = sum(r[7] for r in recibidas)

    auto_width(ws2)
    fp2 = os.path.join(OUT, "TEST_IVA300_Listado_DIAN.xlsx")
    wb2.save(fp2)
    print(f"    OK: {len(all_fe)} facturas ({len(emitidas)} emitidas, {len(recibidas)} recibidas)")
    print(f"    IVA generado FE: {iva_gen:,.0f} | IVA descontable FE: {iva_desc:,.0f}")


# ============================================================
# 4. RETENCION 350 - Libro auxiliar por terceros
# ============================================================
def gen_retencion350():
    print(">>> Generando TEST_Retencion350_Auxiliar.xlsx ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auxiliar Retenciones"

    # Titulo
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value="LIBRO AUXILIAR POR TERCEROS - CUENTAS DE RETENCION")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value="Periodo: Enero a Marzo 2026 | NIT: 901234567-1 TechServicios Colombia S.A.S.")
    ws.cell(row=2, column=1).font = Font(italic=True, size=10)

    headers = ["Cuenta", "Nombre Cuenta", "NIT", "Nombre Tercero", "Debito", "Credito"]
    h_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=h_row, column=i, value=h)
    header_style(ws, h_row, len(headers))

    data = [
        # RETENCION SALARIOS 236505
        ("236505","Retencion salarios (art 383 ET)","1001234567","Juan Carlos Rodriguez Perez",0,2850000),
        ("236505","Retencion salarios (art 383 ET)","1019876543","Maria Fernanda Lopez Garcia",0,1920000),
        ("236505","Retencion salarios (art 383 ET)","80234567","Carlos Alberto Martinez Ruiz",0,3150000),
        ("236505","Retencion salarios (art 383 ET)","52198765","Ana Maria Gomez Torres",0,980000),
        ("236505","Retencion salarios (art 383 ET)","1098765432","Pedro Antonio Sanchez Luna",0,1450000),
        ("236505","Retencion salarios (art 383 ET)","1045678901","Laura Cristina Vargas Diaz",0,720000),
        # Pago DIAN mes anterior
        ("236505","Retencion salarios (art 383 ET)","800123456","DIAN - Pago declaracion enero",8200000,0),
        # RETENCION HONORARIOS 236515
        ("236515","Retencion honorarios 11%","900123456","Consultoria ABC S.A.S.",0,4950000),
        ("236515","Retencion honorarios 11%","860012345","Asesoria Legal Ltda",0,2200000),
        ("236515","Retencion honorarios 11%","1098765432","Pedro Antonio Sanchez Luna",0,1210000),
        ("236515","Retencion honorarios 11%","900111222","Marketing Pro S.A.S.",0,1650000),
        ("236515","Retencion honorarios 11%","800123456","DIAN - Pago declaracion enero",7500000,0),
        # RETENCION COMISIONES 236520
        ("236520","Retencion comisiones 11%","900234567","Distribuidora XYZ S.A.",0,1650000),
        ("236520","Retencion comisiones 11%","800345678","Agencia Comercial Norte Ltda",0,880000),
        ("236520","Retencion comisiones 11%","900333444","Broker Inmobiliario S.A.S.",0,550000),
        # RETENCION SERVICIOS 236525
        ("236525","Retencion servicios 4%","900345678","Servicios Integrados S.A.S.",0,3200000),
        ("236525","Retencion servicios 4%","900456789","TechSolutions Colombia S.A.S.",0,1800000),
        ("236525","Retencion servicios 6%","900567890","Transporte Nacional S.A.",0,2400000),
        ("236525","Retencion servicios 4%","811234567","Aseo y Cafeteria Plus S.A.S.",0,640000),
        ("236525","Retencion servicios 4%","1045678901","Freelancer Digital (PN declarante)",0,960000),
        ("236525","Retencion servicios 2%","900444555","Logistica Santandereana Ltda",0,480000),
        ("236525","Retencion servicios 4%","900555666","Soporte IT Express S.A.S.",0,1120000),
        ("236525","Retencion servicios 4%","800123456","DIAN - Pago declaracion enero",6800000,0),
        # RETENCION ARRENDAMIENTOS 236530
        ("236530","Retencion arrendamientos 3.5%","900678901","Inmobiliaria Centro S.A.",0,1260000),
        ("236530","Retencion arrendamientos 3.5%","19876543","Roberto Diaz Mejia (PN)",0,840000),
        ("236530","Retencion arrendamientos 3.5%","900666777","Edificio Empresarial Torre 26 PH",0,525000),
        # RETENCION COMPRAS 236540
        ("236540","Retencion compras 2.5%","900789012","Suministros Industriales S.A.",0,6875000),
        ("236540","Retencion compras 2.5%","900890123","Materiales del Caribe Ltda",0,3750000),
        ("236540","Retencion compras 2.5%","800456789","Ferreteria Nacional S.A.S.",0,1250000),
        ("236540","Retencion compras 2.5%","900901234","Importadora Global S.A.",0,4500000),
        ("236540","Retencion compras 2.5%","900777888","Dotaciones Corporativas SAS",0,875000),
        ("236540","Retencion compras 2.5%","900888999","Papeleria y Suministros Ltda",0,625000),
        ("236540","Retencion compras 2.5%","800123456","DIAN - Pago declaracion enero",12500000,0),
        # RETENCION RENDIMIENTOS FINANCIEROS 236545
        ("236545","Retencion rendimientos financieros 7%","860002964","Banco de Bogota S.A.",0,595000),
        ("236545","Retencion rendimientos financieros 7%","890903938","Bancolombia S.A.",0,420000),
        ("236545","Retencion rendimientos financieros 7%","860007738","Banco Popular S.A.",0,280000),
        # RETENCION POR OTROS CONCEPTOS 236570 (ICA)
        ("236570","Retencion ICA Bogota 11.04x1000","900123456","Consultoria ABC S.A.S.",0,450000),
        ("236570","Retencion ICA Bogota 9.66x1000","900345678","Servicios Integrados S.A.S.",0,320000),
        ("236570","Retencion ICA Bogota 4.14x1000","900789012","Suministros Industriales S.A.",0,275000),
        ("236570","Retencion ICA Bogota 11.04x1000","900456789","TechSolutions Colombia S.A.S.",0,198000),
        ("236570","Retencion ICA Bogota","800123456","DIAN/Secretaria Hacienda - Pago",900000,0),
        # AUTORRETENCION RENTA 236575
        ("236575","Autorretencion especial renta 0.80%","901234567","NUESTRA EMPRESA S.A.S. (auto)",0,4120000),
        ("236575","Autorretencion especial renta 0.80%","901234567","NUESTRA EMPRESA S.A.S. (auto)",0,3850000),
        ("236575","Autorretencion especial renta 0.80%","901234567","NUESTRA EMPRESA S.A.S. (auto)",0,4280000),
        ("236575","Autorretencion especial renta","800123456","DIAN - Pago declaracion enero",3500000,0),
        # RETENCION IVA 236701
        ("236701","Retencion IVA regimen comun 15%","900123456","Consultoria ABC S.A.S.",0,1414500),
        ("236701","Retencion IVA regimen comun 15%","900345678","Servicios Integrados S.A.S.",0,912000),
        ("236701","Retencion IVA regimen comun 15%","900456789","TechSolutions Colombia S.A.S.",0,513000),
        ("236701","Retencion IVA regimen comun 15%","900678901","Inmobiliaria Centro S.A.",0,360000),
        ("236701","Retencion IVA regimen comun 15%","900789012","Suministros Industriales S.A.",0,1959375),
        ("236701","Retencion IVA regimen comun 15%","900890123","Materiales del Caribe Ltda",0,1068750),
        ("236701","Retencion IVA","800123456","DIAN - Pago declaracion enero",4200000,0),
        # RETENCION ICA 236805
        ("236805","Retencion ICA municipio Bogota","900234567","Distribuidora XYZ S.A.",0,150000),
        ("236805","Retencion ICA municipio Bogota","900567890","Transporte Nacional S.A.",0,120000),
        ("236805","Retencion ICA municipio Medellin","900456789","TechSolutions Colombia S.A.S.",0,95000),
        ("236805","Retencion ICA","800123456","Secretaria Hacienda - Pago",250000,0),
        # AUTORRETENCION A FAVOR 135515
        ("135515","Autorretencion renta que nos practicaron","900345678","Cliente Grande S.A.",3800000,0),
        ("135515","Autorretencion renta que nos practicaron","900111222","Distribuidora Norte S.A.",2200000,0),
        ("135515","Autorretencion renta que nos practicaron","901012345","Gobierno Municipal Bogota",1500000,0),
        ("135515","Autorretencion renta que nos practicaron","900222333","Comercial Sur Ltda",980000,0),
        ("135517","Retencion IVA que nos practicaron","900345678","Cliente Grande S.A.",1200000,0),
        ("135517","Retencion IVA que nos practicaron","901012345","Gobierno Municipal Bogota",850000,0),
    ]

    for i, row in enumerate(data, h_row + 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=h_row + 1 + (i - h_row - 1), column=j, value=val)
        money_fmt(ws, h_row + 1 + (i - h_row - 1), [5, 6])

    total_deb = sum(r[4] for r in data)
    total_cre = sum(r[5] for r in data)
    r_tot = h_row + len(data) + 2
    ws.cell(row=r_tot, column=4, value="TOTALES")
    ws.cell(row=r_tot, column=4).font = Font(bold=True)
    ws.cell(row=r_tot, column=5, value=total_deb)
    ws.cell(row=r_tot, column=6, value=total_cre)
    money_fmt(ws, r_tot, [5, 6])

    auto_width(ws)
    fp = os.path.join(OUT, "TEST_Retencion350_Auxiliar.xlsx")
    wb.save(fp)
    print(f"    OK: {len(data)} movimientos, Deb={total_deb:,.0f}, Cre={total_cre:,.0f}")


# ============================================================
# 5. ESTADOS FINANCIEROS - Dos balances comparativos
# ============================================================
def gen_eeff():
    # OPEN trial balance: NO 3605 (utility is in P&L accounts 4,5,6)
    # Only 4-digit accounts to avoid double counting
    # Income auto-calculated so Deb = Cre
    for year, label in [(2024, "2024"), (2025, "2025")]:
        print(f">>> Generando TEST_EEFF_Balance_{label}.xlsx ...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Balance {label}"

        headers = ["Cuenta", "Nombre Cuenta", "Saldo Inicial", "Debito", "Credito", "Saldo Final"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        header_style(ws, 1, len(headers))

        if year == 2024:
            activos = [
                ("1105","Caja general",4500000),
                ("1110","Bancos nacionales",185000000),
                ("1120","Cuentas de ahorro",45000000),
                ("1205","Inversiones CDT",30000000),
                ("1305","Clientes nacionales",280000000),
                ("1330","Anticipos y avances",8000000),
                ("1355","Anticipo de impuestos",22000000),
                ("1380","Deudores varios",5000000),
                ("1399","Provision deudores",-4000000),
                ("1504","Terrenos",120000000),
                ("1516","Construcciones y edificaciones",180000000),
                ("1520","Maquinaria y equipo",45000000),
                ("1524","Equipo de oficina",28000000),
                ("1528","Equipo de computacion",65000000),
                ("1540","Flota y equipo de transporte",85000000),
                ("1592","Depreciacion acumulada",-78000000),
                ("1610","Marcas",8000000),
                ("1698","Amortizacion acumulada",-2000000),
                ("1705","Gastos pagados por anticipado",6000000),
            ]
            pasivos = [
                ("2105","Obligaciones financieras corto plazo",45000000),
                ("2205","Proveedores nacionales",65000000),
                ("2210","Proveedores del exterior",12000000),
                ("2335","Costos y gastos por pagar",22000000),
                ("2365","Retencion en la fuente por pagar",8500000),
                ("2367","Impuesto IVA retenido",3200000),
                ("2368","Impuesto ICA retenido",1800000),
                ("2404","Impuesto de renta por pagar",15000000),
                ("2408","IVA por pagar",15000000),
                ("2505","Salarios por pagar",18000000),
                ("2510","Cesantias consolidadas",15000000),
                ("2515","Intereses sobre cesantias",1800000),
                ("2520","Prima de servicios",12000000),
                ("2525","Vacaciones consolidadas",8000000),
                ("2610","Provision obligaciones fiscales",12000000),
                ("2805","Anticipos y avances recibidos",5000000),
            ]
            patrimonio_base = [
                ("3105","Capital suscrito y pagado",200000000),
                ("3115","Aportes sociales",50000000),
                ("3205","Reserva legal",25000000),
                ("3210","Reserva estatutaria",10000000),
                ("3305","Resultados de ejercicios anteriores",85000000),
            ]
            gastos_list = [
                ("5105","Gastos de personal",520000000),
                ("5110","Honorarios",85000000),
                ("5115","Impuestos",35000000),
                ("5120","Arrendamientos",72000000),
                ("5125","Seguros",18000000),
                ("5130","Servicios",45000000),
                ("5135","Gastos legales",12000000),
                ("5140","Gastos de servicios publicos",24000000),
                ("5145","Gastos de viaje",32000000),
                ("5150","Publicidad y propaganda",18000000),
                ("5155","Mantenimiento y reparaciones",15000000),
                ("5160","Depreciaciones",28000000),
                ("5195","Gastos diversos",8000000),
                ("5199","Provisiones",4000000),
                ("5305","Gastos financieros",6800000),
                ("5315","Gastos extraordinarios",2000000),
            ]
            costos_list = [
                ("6135","Costo de mercancias vendidas",290000000),
                ("6140","Costo de servicios prestados",780000000),
            ]
        else:  # 2025
            activos = [
                ("1105","Caja general",4000000),
                ("1110","Bancos nacionales",245000000),
                ("1120","Cuentas de ahorro",68000000),
                ("1205","Inversiones CDT",50000000),
                ("1305","Clientes nacionales",365000000),
                ("1310","Cuentas por cobrar a socios",15000000),
                ("1330","Anticipos y avances",12000000),
                ("1355","Anticipo de impuestos",35000000),
                ("1380","Deudores varios",3500000),
                ("1399","Provision deudores",-6500000),
                ("1435","Mercancias no fabricadas",15000000),
                ("1504","Terrenos",120000000),
                ("1516","Construcciones y edificaciones",280000000),
                ("1520","Maquinaria y equipo",55000000),
                ("1524","Equipo de oficina",35000000),
                ("1528","Equipo de computacion",95000000),
                ("1540","Flota y equipo de transporte",85000000),
                ("1592","Depreciacion acumulada",-115000000),
                ("1610","Marcas",8000000),
                ("1615","Patentes",12000000),
                ("1698","Amortizacion acumulada",-4000000),
                ("1705","Gastos pagados por anticipado",9000000),
            ]
            pasivos = [
                ("2105","Obligaciones financieras corto plazo",35000000),
                ("2120","Obligaciones financieras largo plazo",180000000),
                ("2205","Proveedores nacionales",128000000),
                ("2210","Proveedores del exterior",18000000),
                ("2335","Costos y gastos por pagar",28000000),
                ("2360","Dividendos por pagar",25000000),
                ("2365","Retencion en la fuente por pagar",11200000),
                ("2367","Impuesto IVA retenido",4500000),
                ("2368","Impuesto ICA retenido",2400000),
                ("2404","Impuesto de renta por pagar",22000000),
                ("2408","IVA por pagar",22000000),
                ("2505","Salarios por pagar",24000000),
                ("2510","Cesantias consolidadas",22000000),
                ("2515","Intereses sobre cesantias",2640000),
                ("2520","Prima de servicios",18000000),
                ("2525","Vacaciones consolidadas",11500000),
                ("2610","Provision obligaciones fiscales",18000000),
                ("2805","Anticipos y avances recibidos",15000000),
            ]
            patrimonio_base = [
                ("3105","Capital suscrito y pagado",200000000),
                ("3115","Aportes sociales",100000000),
                ("3205","Reserva legal",42000000),
                ("3210","Reserva estatutaria",15000000),
                ("3305","Resultados de ejercicios anteriores",198000000),
            ]
            gastos_list = [
                ("5105","Gastos de personal",750000000),
                ("5110","Honorarios",110000000),
                ("5115","Impuestos",48000000),
                ("5120","Arrendamientos",48000000),
                ("5125","Seguros",25000000),
                ("5130","Servicios",58000000),
                ("5135","Gastos legales",15000000),
                ("5140","Gastos de servicios publicos",28000000),
                ("5145","Gastos de viaje",45000000),
                ("5150","Publicidad y propaganda",35000000),
                ("5155","Mantenimiento y reparaciones",18000000),
                ("5160","Depreciaciones",37000000),
                ("5195","Gastos diversos",10000000),
                ("5199","Provisiones",6500000),
                ("5305","Gastos financieros",12000000),
                ("5315","Gastos extraordinarios",3000000),
            ]
            costos_list = [
                ("6135","Costo de mercancias vendidas",380000000),
                ("6140","Costo de servicios prestados",950000000),
            ]

        # Calculate required income so trial balance cuadra
        total_activos_pos = sum(s for _,_,s in activos if s > 0)
        total_activos_neg = sum(abs(s) for _,_,s in activos if s < 0)
        total_activos_net = total_activos_pos - total_activos_neg
        total_pasivos = sum(s for _,_,s in pasivos)
        total_pt_base = sum(s for _,_,s in patrimonio_base)
        total_gastos = sum(s for _,_,s in gastos_list)
        total_costos = sum(s for _,_,s in costos_list)

        # Deb = activos_pos + gastos + costos
        # Cre = activos_neg + pasivos + pt_base + ingresos
        # For Deb = Cre: ingresos = activos_pos + gastos + costos - activos_neg - pasivos - pt_base
        required_income = total_activos_pos + total_gastos + total_costos - total_activos_neg - total_pasivos - total_pt_base

        # Distribute income across accounts
        ingresos_list = [
            ("4135","Comercio al por mayor y menor",int(required_income * 0.22)),
            ("4140","Servicios de consultoria IT",int(required_income * 0.62)),
            ("4145","Servicios desarrollo software",int(required_income * 0.08)),
            ("4210","Intereses financieros",int(required_income * 0.01)),
            ("4220","Comisiones",int(required_income * 0.02)),
            ("4250","Recuperaciones",int(required_income * 0.005)),
            ("4255","Indemnizaciones",int(required_income * 0.003)),
            ("4295","Ingresos diversos",0),
        ]
        total_ingresos_raw = sum(s for _,_,s in ingresos_list)
        diff = required_income - total_ingresos_raw
        ingresos_list[-1] = ("4295","Ingresos diversos", diff)
        total_ingresos = sum(s for _,_,s in ingresos_list)
        utilidad = total_ingresos - total_gastos - total_costos

        # Build rows: (cuenta, nombre, si, deb, cre, sf)
        data = []
        for acct, name, saldo in activos:
            if saldo >= 0:
                data.append((acct, name, 0, saldo, 0, saldo))
            else:
                data.append((acct, name, 0, 0, abs(saldo), saldo))
        for acct, name, saldo in pasivos:
            data.append((acct, name, 0, 0, saldo, saldo))
        for acct, name, saldo in patrimonio_base:
            data.append((acct, name, 0, 0, saldo, saldo))
        # NO 3605 - open trial balance
        for acct, name, saldo in ingresos_list:
            data.append((acct, name, 0, 0, saldo, saldo))
        for acct, name, saldo in gastos_list:
            data.append((acct, name, 0, saldo, 0, saldo))
        for acct, name, saldo in costos_list:
            data.append((acct, name, 0, saldo, 0, saldo))

        total_deb = sum(r[3] for r in data)
        total_cre = sum(r[4] for r in data)

        for i, row in enumerate(data, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)
            money_fmt(ws, i, [3,4,5,6])

        r = len(data) + 2
        ws.cell(row=r, column=2, value="TOTALES")
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=4, value=total_deb)
        ws.cell(row=r, column=5, value=total_cre)
        money_fmt(ws, r, [4,5])

        auto_width(ws)
        fp = os.path.join(OUT, f"TEST_EEFF_Balance_{label}.xlsx")
        wb.save(fp)
        print(f"    OK {label}: {len(data)} cuentas")
        print(f"    Ingresos: {total_ingresos:,.0f}")
        print(f"    Gastos: {total_gastos:,.0f}")
        print(f"    Costos: {total_costos:,.0f}")
        print(f"    Utilidad: {utilidad:,.0f}")
        print(f"    Activos neto: {total_activos_net:,.0f}")
        print(f"    Pasivos: {total_pasivos:,.0f}")
        print(f"    Patrimonio base+Util: {total_pt_base+utilidad:,.0f}")
        print(f"    A=P+Pt? {total_activos_net:,.0f} vs {total_pasivos+total_pt_base+utilidad:,.0f} -> {'OK' if total_activos_net==total_pasivos+total_pt_base+utilidad else 'DESCUADRADO!'}")
        print(f"    Deb: {total_deb:,.0f} | Cre: {total_cre:,.0f} | Dif: {total_deb-total_cre:,.0f}")


# ============================================================
# 6. CONCILIACION BANCARIA - Auxiliar + Extracto
# ============================================================
def gen_conciliacion():
    print(">>> Generando TEST_Conciliacion_Auxiliar_Bancos.xlsx ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auxiliar Bancos"

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value="AUXILIAR CONTABLE - CUENTA 111005 BANCO DE BOGOTA CTA CTE 012-345678-90")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A2:H2')
    ws.cell(row=2, column=1, value="Periodo: Marzo 2026 | NIT: 901234567-1 TechServicios Colombia S.A.S.")

    headers = ["Fecha", "Comprobante", "NIT", "Nombre Tercero", "Descripcion", "Debito", "Credito", "Saldo"]
    h_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=h_row, column=i, value=h)
    header_style(ws, h_row, len(headers))

    # (fecha, comp, nit, tercero, desc, deb, cre, saldo)
    saldo = 85432100
    movs = [
        ("01/03/2026","APERTURA","901234567","TechServicios","Apertura periodo marzo 2026",0,0,saldo),
        ("02/03/2026","EG-001","900123456","Proveedor ABC S.A.S.","Pago factura FV-2345",0,12500000,None),
        ("02/03/2026","EG-002","900234567","Servicios XYZ Ltda","Pago servicio mensual marzo",0,3800000,None),
        ("03/03/2026","RC-001","900345678","Cliente Grande S.A.","Recaudo factura FV-1001",25000000,0,None),
        ("05/03/2026","EG-003","860002964","Banco de Bogota S.A.","Cuota credito #45678 capital+intereses",0,5200000,None),
        ("05/03/2026","ND-001","860002964","Banco de Bogota S.A.","Nota debito comision transferencia ACH",0,45000,None),
        ("07/03/2026","RC-002","900456789","Tech Solutions S.A.S.","Recaudo factura FV-1002",8750000,0,None),
        ("08/03/2026","EG-004","800123456","DIAN","Pago retencion en la fuente febrero 2026",0,4350000,None),
        ("10/03/2026","EG-005","900567890","Inmobiliaria Centro S.A.","Pago canon arriendo oficina marzo",0,6000000,None),
        ("10/03/2026","RC-003","900678901","Distribuidora Norte S.A.","Recaudo factura FV-1003 transferencia",15200000,0,None),
        ("12/03/2026","EG-006","900789012","Suministros Industriales S.A.","Pago factura FV-3456 materiales",0,8900000,None),
        ("13/03/2026","NM-001","901234567","Nomina","Pago nomina primera quincena marzo",0,18500000,None),
        ("14/03/2026","RC-004","900890123","Comercial Sur Ltda","Recaudo factura FV-1004",12300000,0,None),
        ("15/03/2026","EG-007","800234567","EPS Sura","Pago aportes salud marzo",0,3200000,None),
        ("15/03/2026","EG-008","800345678","AFP Proteccion","Pago aportes pension marzo",0,4800000,None),
        ("17/03/2026","RC-005","900345678","Cliente Grande S.A.","Recaudo factura FV-1005",18500000,0,None),
        ("18/03/2026","EG-009","900901234","Publicidad Digital SAS","Pago campana Google Ads marzo",0,2500000,None),
        ("19/03/2026","EG-010","860012345","Asesoria Legal Ltda","Pago honorarios asesoria marzo",0,4500000,None),
        ("20/03/2026","RC-006","901012345","Gobierno Municipal Bogota","Recaudo contrato 2026-001 anticipo",35000000,0,None),
        ("21/03/2026","EG-011","900123456","Proveedor ABC S.A.S.","Pago factura FV-2400",0,9800000,None),
        ("22/03/2026","EG-012","899999999","EPM","Pago servicios publicos febrero",0,1850000,None),
        ("24/03/2026","RC-007","900456789","Tech Solutions S.A.S.","Recaudo factura FV-1006",6200000,0,None),
        ("25/03/2026","EG-013","900234567","Servicios XYZ Ltda","Pago servicio mensual abril anticipado",0,3800000,None),
        ("26/03/2026","NM-002","901234567","Nomina","Pago nomina segunda quincena marzo",0,19200000,None),
        ("27/03/2026","RC-008","900678901","Distribuidora Norte S.A.","Recaudo factura FV-1007 consignacion",11800000,0,None),
        ("28/03/2026","EG-014","800678901","Seguros Bolivar S.A.","Pago poliza todo riesgo trimestral",0,4500000,None),
        ("28/03/2026","EG-015","860002964","Banco de Bogota S.A.","Cuota credito #45678 capital+intereses",0,5200000,None),
        ("29/03/2026","CH-001","900789012","Suministros Industriales S.A.","Cheque #4567 pago factura FV-3500",0,7500000,None),
        ("30/03/2026","CH-002","900890123","Comercial Sur Ltda","Cheque #4568 anticipo pedido",0,5000000,None),
        ("31/03/2026","NC-001","860002964","Banco de Bogota S.A.","Nota credito intereses CDT marzo",850000,0,None),
        ("31/03/2026","ND-002","860002964","Banco de Bogota S.A.","Gravamen movimientos financieros 4x1000",0,520000,None),
    ]

    # Calcular saldos
    for i in range(1, len(movs)):
        f, comp, nit, terc, desc, deb, cre, s = movs[i]
        saldo = saldo + deb - cre
        movs[i] = (f, comp, nit, terc, desc, deb, cre, saldo)

    book_balance = saldo
    for i, row in enumerate(movs, h_row + 1):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        money_fmt(ws, i, [6, 7, 8])

    auto_width(ws)
    fp = os.path.join(OUT, "TEST_Conciliacion_Auxiliar_Bancos.xlsx")
    wb.save(fp)
    print(f"    OK: {len(movs)} movimientos, Saldo final libros: {book_balance:,.0f}")

    # --- EXTRACTO BANCARIO ---
    print(">>> Generando TEST_Conciliacion_Extracto_Bancario.xlsx ...")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Extracto Bancario"

    ws2.merge_cells('A1:F1')
    ws2.cell(row=1, column=1, value="EXTRACTO BANCARIO - BANCO DE BOGOTA")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.merge_cells('A2:F2')
    ws2.cell(row=2, column=1, value="Cuenta Corriente No. 012-345678-90")
    ws2.merge_cells('A3:F3')
    ws2.cell(row=3, column=1, value="Periodo: 01/03/2026 al 31/03/2026")
    ws2.merge_cells('A4:F4')
    ws2.cell(row=4, column=1, value="Cliente: TechServicios Colombia S.A.S. - NIT 901234567-1")

    headers2 = ["Fecha", "Descripcion", "Referencia", "Debito", "Credito", "Saldo"]
    h_row2 = 6
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=h_row2, column=i, value=h)
    header_style(ws2, h_row2, len(headers2))

    # Extracto: desde perspectiva del banco
    # Debito = sale dinero (pagos), Credito = entra dinero (recaudos)
    # NO incluye: CH-001, CH-002 (cheques pendientes), RC-008 (consignacion en transito)
    # SI incluye: cargos bancarios no registrados en libros
    bank_saldo = 85432100
    extracto = [
        ("01/03/2026","APERTURA PERIODO MARZO","APERTURA",0,0,bank_saldo),
        ("02/03/2026","PAGO PROVEEDOR TRANSF ACH","REF-90012345",12500000,0,None),
        ("02/03/2026","PAGO SERVICIO TRANSF ACH","REF-90023456",3800000,0,None),
        ("03/03/2026","CONSIGNACION NACIONAL","CONS-001",0,25000000,None),
        ("05/03/2026","PAGO CUOTA CREDITO 45678","AUTO-DEBITO",5200000,0,None),
        ("05/03/2026","COMISION TRANSFERENCIA ACH","ND-2026030501",45000,0,None),
        ("07/03/2026","CONSIGNACION CHEQUE LOCAL","CONS-002",0,8750000,None),
        ("08/03/2026","PAGO IMPUESTOS PSE DIAN","PSE-80012345",4350000,0,None),
        ("10/03/2026","PAGO ARRIENDO TRANSF ACH","REF-90056789",6000000,0,None),
        ("10/03/2026","TRANSFERENCIA RECIBIDA","TRANSF-003",0,15200000,None),
        ("12/03/2026","PAGO PROVEEDOR TRANSF ACH","REF-90078901",8900000,0,None),
        ("13/03/2026","PAGO NOMINA LOTE","NOM-2026031501",18500000,0,None),
        ("14/03/2026","CONSIGNACION NACIONAL","CONS-004",0,12300000,None),
        ("15/03/2026","PAGO APORTES EPS PILA","PILA-80023456",3200000,0,None),
        ("15/03/2026","PAGO APORTES PENSION PILA","PILA-80034567",4800000,0,None),
        ("15/03/2026","ND COMISION CHEQUERA 50 CHEQUES","ND-2026031502",125000,0,None),  # Solo en banco
        ("17/03/2026","CONSIGNACION NACIONAL","CONS-005",0,18500000,None),
        ("18/03/2026","PAGO PSE GOOGLE ADS","PSE-90090123",2500000,0,None),
        ("19/03/2026","PAGO PROVEEDOR TRANSF ACH","REF-86001234",4500000,0,None),
        ("20/03/2026","TRANSFERENCIA RECIBIDA GOBIERNO","TRANSF-006",0,35000000,None),
        ("21/03/2026","PAGO PROVEEDOR TRANSF ACH","REF-90012345B",9800000,0,None),
        ("22/03/2026","PAGO SERVICIOS PSE EPM","PSE-89999999",1850000,0,None),
        ("24/03/2026","CONSIGNACION CHEQUE LOCAL","CONS-007",0,6200000,None),
        ("25/03/2026","PAGO SERVICIO TRANSF ACH","REF-90023456B",3800000,0,None),
        ("26/03/2026","PAGO NOMINA LOTE","NOM-2026032601",19200000,0,None),
        # RC-008 (27/03) NO aparece - consignacion en transito
        ("28/03/2026","PAGO SEGUROS TRANSF ACH","REF-80067890",4500000,0,None),
        ("28/03/2026","PAGO CUOTA CREDITO 45678","AUTO-DEBITO",5200000,0,None),
        ("28/03/2026","NC DEVOLUCION COMISION COBRO","NC-2026032801",0,45000,None),  # Solo en banco
        # CH-001 y CH-002 NO aparecen - cheques pendientes de cobro
        ("31/03/2026","ABONO INTERESES CDT","NC-CDT-0326",0,850000,None),
        ("31/03/2026","GRAVAMEN MOV FINANCIEROS GMF","GMF-2026-03",520000,0,None),
        ("31/03/2026","ND CUOTA MANEJO TARJETA EMPRESARIAL","ND-2026033101",89000,0,None),  # Solo en banco
    ]

    # Calcular saldos del extracto
    for i in range(1, len(extracto)):
        f, desc, ref, deb, cre, s = extracto[i]
        bank_saldo = bank_saldo - deb + cre
        extracto[i] = (f, desc, ref, deb, cre, bank_saldo)

    for i, row in enumerate(extracto, h_row2 + 1):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
        money_fmt(ws2, i, [4, 5, 6])

    auto_width(ws2)
    fp2 = os.path.join(OUT, "TEST_Conciliacion_Extracto_Bancario.xlsx")
    wb2.save(fp2)
    print(f"    OK: {len(extracto)} movimientos, Saldo final banco: {bank_saldo:,.0f}")

    # Verificar conciliacion
    cheques_pend = 7500000 + 5000000
    consig_transito = 11800000
    nd_no_registradas = 125000 + 89000
    nc_no_registradas = 45000
    book_ajustado = book_balance + nc_no_registradas - nd_no_registradas
    bank_ajustado = bank_saldo - cheques_pend + consig_transito

    print(f"\n    === CONCILIACION ===")
    print(f"    Saldo libros:            {book_balance:>15,.0f}")
    print(f"    + NC banco no registrada:   +{nc_no_registradas:>12,.0f}")
    print(f"    - ND banco no registradas:  -{nd_no_registradas:>12,.0f}")
    print(f"    = Saldo libros ajustado: {book_ajustado:>15,.0f}")
    print(f"    Saldo banco:             {bank_saldo:>15,.0f}")
    print(f"    - Cheques pendientes:       -{cheques_pend:>12,.0f}")
    print(f"    + Consignacion transito:    +{consig_transito:>12,.0f}")
    print(f"    = Saldo banco ajustado:  {bank_ajustado:>15,.0f}")
    print(f"    Concilia? {'SI' if book_ajustado == bank_ajustado else 'NO - REVISAR!'}")


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 70)
    print("GENERADOR DE ARCHIVOS DE PRUEBA - EXÓGENA DIAN")
    print(f"Directorio: {OUT}")
    print("=" * 70)
    print()

    gen_exogena()
    print()
    gen_renta110()
    print()
    gen_iva300()
    print()
    gen_retencion350()
    print()
    gen_eeff()
    print()
    gen_conciliacion()

    print()
    print("=" * 70)
    print("ARCHIVOS GENERADOS:")
    for f in sorted(os.listdir(OUT)):
        if f.startswith("TEST_") and f.endswith(".xlsx"):
            size = os.path.getsize(os.path.join(OUT, f))
            print(f"  {f:<55} {size:>8,} bytes")
    print("=" * 70)
    print("LISTO! Todos los archivos de prueba han sido generados.")
